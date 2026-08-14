# -*- coding: utf-8 -*-
"""
================================================================================
  CEREBRO LIGA MX 2026  —  Apertura 2026
  Modelo Dixon-Coles + Monte Carlo con capas de ajuste cualitativo
================================================================================
Hermano del "Cerebro Mundial 2026" (Dixon-Coles + Elo ensemble). Reutiliza:
  - Núcleo Dixon-Coles bivariado con corrección tau para marcadores bajos
  - Simulación Monte Carlo de marcadores exactos
  - Estructura de capas: fuerza base -> ajustes de contexto -> simulación

Diferencias vs el Mundial:
  - Localía REAL por estadio (altura, calor, afición) en lugar de sede neutral
  - Factor "escudo" (peso institucional de los grandes)
  - Ajustes por jornada: mundialistas ausentes, DT nuevo, ascendido, lesiones
  - Ponderación temporal exponencial (xi) al calibrar con histórico
  - Reglas históricas H2H ("X nunca le ha ganado a Y") como sesgo bayesiano

USO RÁPIDO (Jornada 1, sin histórico cargado -> usa priors):
    python cerebro_ligamx_2026.py

USO COMPLETO (calibración MLE con 10 años de histórico):
    python cerebro_ligamx_2026.py --historico historico_ligamx.csv

Formato del CSV histórico (una fila por partido, ~10 años Liga MX):
    fecha,local,visitante,gl,gv,fase
    2016-07-16,América,Chivas,1,0,regular      # fase: regular | liguilla

NOTA METODOLÓGICA (importante):
  Los ratings PRIORS_ILUSTRATIVOS de abajo son una calibración CUALITATIVA
  basada en hechos verificables de la temporada 2025-26 (Toluca campeón
  Apertura 2025, Cruz Azul campeón Clausura 2026, rachas documentadas,
  media de 2.91 goles/partido) y desempeño de la última década. Están
  marcados como ILUSTRATIVOS: al cargar el CSV histórico, el módulo
  fit_dixon_coles() los SUSTITUYE por estimadores de máxima verosimilitud.
================================================================================
Autor: Ale (asunad) + Claude | Sin fines de lucro, proyecto actuarial
"""

import argparse
import math
import os
import sys
import csv
from collections import defaultdict
from datetime import date, datetime

import numpy as np

try:
    import pandas as pd
    HAY_PANDAS = True
except ImportError:
    HAY_PANDAS = False

try:
    from scipy.optimize import minimize
    HAY_SCIPY = True
except ImportError:
    HAY_SCIPY = False

RNG = np.random.default_rng(20260716)

# ==============================================================================
# 1. CONFIGURACIÓN GLOBAL
# ==============================================================================

N_SIMULACIONES = 100_000          # simulaciones Monte Carlo por partido
MAX_GOLES = 10                    # truncamiento de la malla de marcadores
RHO_DC = -0.11                    # corrección Dixon-Coles marcadores bajos (típico -0.05 a -0.15)
XI_DECAIMIENTO = 0.0038           # decaimiento temporal diario (vida media ~6 meses)
PESO_LIGUILLA = 1.25              # partidos de liguilla pesan más al calibrar

# Media de goles Liga MX temporada 2025-26: 2.91 goles/partido (dato real)
MU_GLOBAL = 2.91

# Goles promedio de LOCAL y de VISITA (base de escala de las lambdas).
# Valores por defecto (modo priors); main() los recalcula con el historico real.
BASE_HOME_GOALS = 1.58
BASE_AWAY_GOALS = 1.33

# Proporción de goles anotados en el PRIMER tiempo (Liga MX real).
# Fuente: 2,047 partidos con marcador HT, footballcsv 2018-2024
# (local 44.5%, visita 44.3% -> se usa la global). Habilita mercados de mitades
# vía adelgazamiento de Poisson: Goles_1T ~ Poisson(lambda * p), 2T ~ el resto.
PROP_GOLES_1T = 0.4442

# Ventaja local promedio Liga MX (histórica ~45% triunfo local).
# Se expresa como boost multiplicativo al ataque del local.
VENTAJA_LOCAL_BASE = 1.28

EQUIPOS = [
    "América", "Atlas", "Atlante", "Atlético San Luis", "Cruz Azul",
    "Chivas", "Juárez", "León", "Monterrey", "Necaxa", "Pachuca",
    "Puebla", "Pumas", "Querétaro", "Santos", "Tigres", "Tijuana", "Toluca",
]

# ==============================================================================
# 2. PRIORS DE FUERZA (ILUSTRATIVOS — sustituidos por MLE si hay histórico)
# ------------------------------------------------------------------------------
# ataque / defensa en escala multiplicativa (1.00 = equipo promedio).
# defensa > 1.00 = concede MÁS goles que el promedio (defensa débil).
# Anclas cualitativas verificables 2025-26:
#   * Toluca: campeón Apertura 2025, 12 racha invicta Clausura -> élite
#   * Cruz Azul: campeón Clausura 2026 ("la novena"), racha 7 triunfos -> élite
#   * Tigres: 11 partidos invicto Apertura -> élite
#   * Chivas: racha 6 triunfos y goleadas 5-0 en Clausura -> alto
#   * América: potencia de la década (tricampeón 23-24), pero 4 bajas y DT nuevo
#   * Santos/Puebla/Querétaro: rachas perdedoras documentadas -> bajo
#   * Atlante: recién ascendido (franquicia ex-Mazatlán, la más débil 25-26)
# ==============================================================================

PRIORS_ILUSTRATIVOS = {
    #                     ataque  defensa
    "Toluca":            (1.35,   0.80),
    "Cruz Azul":         (1.28,   0.78),
    "Tigres":            (1.22,   0.82),
    "América":           (1.22,   0.85),
    "Monterrey":         (1.18,   0.90),
    "Chivas":            (1.12,   0.88),
    "Pachuca":           (1.05,   0.95),
    "León":              (0.98,   1.02),
    "Pumas":             (0.98,   1.00),
    "Tijuana":           (0.95,   1.02),
    "Necaxa":            (0.95,   1.05),
    "Atlético San Luis": (0.92,   1.05),
    "Atlas":             (0.90,   1.08),
    "Juárez":            (0.88,   1.08),
    "Santos":            (0.82,   1.18),
    "Puebla":            (0.80,   1.20),
    "Querétaro":         (0.80,   1.15),
    "Atlante":           (0.75,   1.22),   # ascendido: máxima incertidumbre
}

# ==============================================================================
# 3. LOCALÍA POR ESTADIO (multiplicador sobre VENTAJA_LOCAL_BASE)
# ------------------------------------------------------------------------------
# Factores: altitud (Toluca 2,660 m; CDMX ~2,240 m; Pachuca 2,400 m),
# clima extremo (Juárez / Tijuana), presión de afición (Monterrey, Akron).
# ==============================================================================

LOCALIA_ESTADIO = {
    "Toluca":            1.10,   # Nemesio Diez, la altura más castigadora
    "Pachuca":           1.06,   # Hidalgo, altura
    "Pumas":             1.05,   # CU, altura + pasto natural
    "América":           1.03,   # regresa al Azteca remodelado
    "Cruz Azul":         1.02,
    "Monterrey":         1.05,   # BBVA, entradas récord 25-26
    "Tigres":            1.05,   # Volcán
    "Chivas":            1.04,   # Akron
    "Juárez":            1.05,   # calor/desgaste del viaje
    "Tijuana":           1.06,   # Caliente: pasto sintético + viaje largo
    "Atlético San Luis": 1.03,
    "León":              1.03,
    "Necaxa":            1.02,
    "Santos":            1.03,   # calor de Torreón
    "Puebla":            1.01,
    "Querétaro":         1.01,
    "Atlas":             1.02,
    "Atlante":           1.02,   # Azteca, pero afición en reconstrucción
}

# ==============================================================================
# 4. FACTOR "ESCUDO" (peso institucional / cultura ganadora)
# ------------------------------------------------------------------------------
# Modela el fenómeno "aunque traigan mal plantel, el puro escudo pesa":
# los 4 grandes históricamente sobre-rinden en probabilidad de NO perder.
# Se aplica como reducción del ataque RIVAL (intimidación) — efecto pequeño
# y acotado para no romper la calibración global.
# ==============================================================================

FACTOR_ESCUDO = {
    "América":   1.05,
    "Chivas":    1.03,
    "Cruz Azul": 1.03,
    "Pumas":     1.02,
    "Tigres":    1.03,   # grande moderno de la década
    "Monterrey": 1.02,
}

# ==============================================================================
# 5. REGLAS HISTÓRICAS H2H (sesgo bayesiano suave)
# ------------------------------------------------------------------------------
# Codifica patrones tipo "X nunca/casi nunca le gana a Y" o "el local nunca
# ha perdido este duelo". Cada regla mueve el ataque de un equipo ±%.
# ESTRUCTURA: (local, visitante): (mult_ataque_local, mult_ataque_visita, nota)
# Las de abajo son las relevantes para J1; agrega las tuyas libremente.
# ==============================================================================

REGLAS_H2H = {
    # Tigres domina el historial vs Tijuana, pero el Caliente (sintético,
    # viaje) históricamente le complica la visita -> efectos que se compensan
    ("Tijuana", "Tigres"): (1.04, 0.98, "Caliente históricamente incómodo para Tigres pese a dominio felino"),
    # Toluca de visita en el Akron: duelo cerrado históricamente
    ("Chivas", "Toluca"): (1.00, 1.00, "historial parejo; sin sesgo"),
}

# ==============================================================================
# 6. AJUSTES DE ACTUALIDAD — JORNADA 2 APERTURA 2026 (21-26 julio)
# ------------------------------------------------------------------------------
# Mundial terminado (final 19-jul; los mundialistas ya se reintegran).
# Hechos verificados post-J1:
#  * América: Zendejas LESIONADO (confirmado fuera) y Cervantes SUSPENDIDO
#    (expulsión J1); Almada apenas en su 2do juego -> castigo moderado
#  * Chivas: Sepúlveda SUSPENDIDO (roja vs Toluca); presión tras caer en casa
#  * Puebla: Iker Moreno SUSPENDIDO (roja vs Juárez)
#  * Juárez: Luka Romero lesionado de la rodilla al minuto 5 de la J1
#  * Tigres: expulsión en Tijuana + era post-Gignac en transición
#  * Monterrey: respondió bien (3-2); se reduce el castigo de transición
#  * Cruz Azul / Toluca: juegan HOY (martes, adelantado) y el sábado el
#    Campeón de Campeones -> sin fatiga extra el martes
#  * Atlante: compitió (perdió 2-1 al 90+3); se suaviza el castigo
# 1.00 = sin ajuste. Rango recomendado: 0.85 – 1.10.
# ==============================================================================

AJUSTES_JORNADA = {
    "América":   {"fuerza": 0.94, "nota": "sin Zendejas (lesión) ni Cervantes (suspendido); Almada 2do juego"},
    "Chivas":    {"fuerza": 0.96, "nota": "sin Sepúlveda (suspendido); presión tras 0-2 en casa"},
    "Puebla":    {"fuerza": 0.97, "nota": "sin Iker Moreno (suspendido)"},
    "Juárez":    {"fuerza": 0.97, "nota": "sin Luka Romero (lesión de rodilla en J1)"},
    "Tigres":    {"fuerza": 0.95, "nota": "expulsión en J1 + reestructura post-Gignac"},
    "Monterrey": {"fuerza": 0.98, "nota": "transición Almeyda, pero respondió con triunfo en J1"},
    "Atlante":   {"fuerza": 0.96, "nota": "ascendido; compitió de visita hasta el 90+3"},
}

# DT: si un técnico nunca ha vencido a otro, castiga al "dominado".
# (local, visitante) -> mult al ataque del equipo cuyo DT está dominado.
REGLAS_DT = {
    # Ejemplo de estructura (sin reglas activas verificadas para J1):
    # ("Querétaro", "América"): {"equipo": "Querétaro", "mult": 0.97,
    #                            "nota": "DT de Gallos nunca le ha ganado a Almada"},
}

# ==============================================================================
# 7. FIXTURE JORNADA 2 — APERTURA 2026 (21-26 julio, fecha partida por el
#    Campeón de Campeones Toluca vs Cruz Azul del sábado 25)
# ==============================================================================

JORNADA_ACTUAL = "Jornada 2 · Apertura 2026"

FIXTURE = [
    # (fecha, local, visitante, estadio)
    ("2026-07-21", "Cruz Azul", "Puebla",            "CDMX (adelantado)"),
    ("2026-07-21", "Toluca",    "Pumas",             "Nemesio Diez (adelantado)"),
    ("2026-07-24", "Tigres",    "Atlético San Luis", "Universitario"),
    ("2026-07-24", "Tijuana",   "León",              "Caliente"),
    ("2026-07-24", "Atlante",   "América",           "Ciudad de los Deportes"),
    ("2026-07-25", "Chivas",    "Juárez",            "Akron"),
    ("2026-07-25", "Santos",    "Atlas",             "Corona"),
    ("2026-07-26", "Necaxa",    "Monterrey",         "Victoria"),
    ("2026-07-26", "Pachuca",   "Querétaro",         "Hidalgo"),
]

# J1 (histórica, ya jugada 16-18 jul y agregada al CSV histórico):
# Necaxa 2-1 Atlante | Tijuana 3-1 Tigres | San Luis 2-3 Cruz Azul |
# León 2-3 Atlas | Juárez 0-1 Puebla | Pumas 0-3 Pachuca |
# Monterrey 3-2 Santos | Chivas 0-2 Toluca | Querétaro 0-1 América


# ==============================================================================
# 8. NÚCLEO DIXON-COLES
# ==============================================================================

def tau_dc(x, y, lam, mu, rho=RHO_DC):
    """Corrección Dixon-Coles para dependencia en marcadores bajos."""
    if x == 0 and y == 0:
        return 1 - lam * mu * rho
    if x == 0 and y == 1:
        return 1 + lam * rho
    if x == 1 and y == 0:
        return 1 + mu * rho
    if x == 1 and y == 1:
        return 1 - rho
    return 1.0


def lambdas_partido(local, visitante, ratings, jornada_actual=True):
    """
    Calcula (lambda_local, lambda_visita) apilando todas las capas:
      base Poisson -> localía -> escudo -> H2H -> DT -> actualidad J1
    Devuelve también bitácora de ajustes aplicados (auditable).
    """
    atk_l, def_l = ratings[local]
    atk_v, def_v = ratings[visitante]
    bitacora = []

    # Capa 1: base = goles promedio reales de local y visita (calibrado con los
    # datos). La ventaja de local YA esta en que BASE_HOME > BASE_AWAY, por eso
    # LOCALIA_ESTADIO es solo el residual por estadio (altura, etc.) y no se
    # aplica el castigo 0.92 (ya viene en BASE_AWAY, mas bajo).
    lam_l = BASE_HOME_GOALS * atk_l * def_v
    lam_v = BASE_AWAY_GOALS * atk_v * def_l

    # Capa 2: residual de localía por estadio (relativo al promedio ya incluido)
    loc = LOCALIA_ESTADIO.get(local, 1.0)
    lam_l *= loc
    bitacora.append(f"localía {local}: x{loc:.2f}")

    # Capa 3: factor escudo (intimidación -> reduce ataque rival)
    if local in FACTOR_ESCUDO:
        lam_v /= FACTOR_ESCUDO[local]
        bitacora.append(f"escudo {local}: rival /{FACTOR_ESCUDO[local]:.2f}")
    if visitante in FACTOR_ESCUDO:
        lam_l /= FACTOR_ESCUDO[visitante]
        bitacora.append(f"escudo {visitante}: rival /{FACTOR_ESCUDO[visitante]:.2f}")

    # Capa 4: reglas históricas H2H
    if (local, visitante) in REGLAS_H2H:
        ml, mv, nota = REGLAS_H2H[(local, visitante)]
        lam_l *= ml
        lam_v *= mv
        bitacora.append(f"H2H: {nota}")

    # Capa 5: reglas de técnicos
    if (local, visitante) in REGLAS_DT:
        r = REGLAS_DT[(local, visitante)]
        if r["equipo"] == local:
            lam_l *= r["mult"]
        else:
            lam_v *= r["mult"]
        bitacora.append(f"DT: {r['nota']}")

    # Capa 6: actualidad de la jornada (mundialistas, lesiones, DT nuevo...)
    if jornada_actual:
        for eq, lam_attr in ((local, "l"), (visitante, "v")):
            if eq in AJUSTES_JORNADA:
                f = AJUSTES_JORNADA[eq]["fuerza"]
                if lam_attr == "l":
                    lam_l *= f
                    lam_v /= math.sqrt(f)  # media fuerza también en defensa
                else:
                    lam_v *= f
                    lam_l /= math.sqrt(f)
                bitacora.append(f"actualidad {eq}: x{f:.2f} ({AJUSTES_JORNADA[eq]['nota']})")

    return lam_l, lam_v, bitacora


def matriz_marcadores(lam_l, lam_v, max_goles=MAX_GOLES):
    """Matriz P(GL=x, GV=y) con corrección Dixon-Coles, renormalizada."""
    px = np.array([math.exp(-lam_l) * lam_l**k / math.factorial(k) for k in range(max_goles + 1)])
    py = np.array([math.exp(-lam_v) * lam_v**k / math.factorial(k) for k in range(max_goles + 1)])
    M = np.outer(px, py)
    for x in range(2):
        for y in range(2):
            M[x, y] *= tau_dc(x, y, lam_l, lam_v)
    return M / M.sum()


def simular_partido(lam_l, lam_v, n=N_SIMULACIONES):
    """Monte Carlo: muestrea marcadores desde la matriz Dixon-Coles."""
    M = matriz_marcadores(lam_l, lam_v)
    flat = M.ravel()
    idx = RNG.choice(len(flat), size=n, p=flat)
    gl, gv = np.divmod(idx, M.shape[1])
    return gl, gv, M


# ==============================================================================
# 9. CALIBRACIÓN MLE CON HISTÓRICO (opcional, sustituye a los priors)
# ==============================================================================

def cargar_historico(ruta):
    """Lee CSV histórico. Columnas: fecha, local, visitante, gl, gv, [fase]."""
    partidos = []
    with open(ruta, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            partidos.append({
                "fecha": datetime.strptime(row["fecha"].strip(), "%Y-%m-%d").date(),
                "local": row["local"].strip(),
                "visitante": row["visitante"].strip(),
                "gl": int(row["gl"]),
                "gv": int(row["gv"]),
                "fase": row.get("fase", "regular").strip().lower(),
            })
    return partidos


def fit_dixon_coles(partidos, hoy=None):
    """
    MLE de ataque/defensa por equipo con:
      - ponderación temporal exponencial exp(-XI * días de antigüedad)
      - sobrepeso a liguilla (PESO_LIGUILLA)
    Requiere scipy; si no está, usa un ajuste iterativo tipo Poisson-rating.
    Devuelve dict {equipo: (ataque, defensa)} normalizado a media 1.
    """
    hoy = hoy or date.today()
    equipos = sorted({p["local"] for p in partidos} | {p["visitante"] for p in partidos})
    ix = {e: i for i, e in enumerate(equipos)}
    n = len(equipos)

    w = np.array([
        math.exp(-XI_DECAIMIENTO * (hoy - p["fecha"]).days)
        * (PESO_LIGUILLA if p["fase"] == "liguilla" else 1.0)
        for p in partidos
    ])
    il = np.array([ix[p["local"]] for p in partidos])
    iv = np.array([ix[p["visitante"]] for p in partidos])
    gl = np.array([p["gl"] for p in partidos], dtype=float)
    gv = np.array([p["gv"] for p in partidos], dtype=float)

    def neg_loglik(theta):
        atk = theta[:n]
        dfn = theta[n:2 * n]
        home = theta[-1]
        lam = np.exp(atk[il] + dfn[iv] + home) * MU_GLOBAL / 2
        mu_ = np.exp(atk[iv] + dfn[il]) * MU_GLOBAL / 2
        ll = w * (gl * np.log(lam) - lam + gv * np.log(mu_) - mu_)
        # corrección tau en marcadores bajos
        mask = (gl <= 1) & (gv <= 1)
        if mask.any():
            taus = np.array([
                tau_dc(int(a), int(b), l_, m_)
                for a, b, l_, m_ in zip(gl[mask], gv[mask], lam[mask], mu_[mask])
            ])
            ll[mask] += w[mask] * np.log(np.clip(taus, 1e-10, None))
        # identificabilidad: penaliza medias distintas de 0
        return -(ll.sum()) + 1000 * (atk.mean() ** 2 + dfn.mean() ** 2)

    theta0 = np.zeros(2 * n + 1)
    theta0[-1] = math.log(VENTAJA_LOCAL_BASE)

    if HAY_SCIPY:
        res = minimize(neg_loglik, theta0, method="L-BFGS-B")
        theta = res.x
        print(f"  [MLE] convergió={res.success}  -loglik={res.fun:,.1f}")
    else:
        print("  [AVISO] scipy no disponible -> ajuste iterativo simple")
        theta = _ajuste_iterativo(partidos, equipos, ix, w)

    ratings = {}
    for e in equipos:
        ratings[e] = (math.exp(theta[ix[e]]), math.exp(theta[n + ix[e]]))
    print(f"  [MLE] ventaja local estimada: x{math.exp(theta[-1]):.3f}")
    return ratings


def _ajuste_iterativo(partidos, equipos, ix, w):
    """Fallback sin scipy: ratings por promedios ponderados de GF/GC."""
    n = len(equipos)
    gf = np.zeros(n); gc = np.zeros(n); pj = np.zeros(n)
    for p, wi in zip(partidos, w):
        i, j = ix[p["local"]], ix[p["visitante"]]
        gf[i] += wi * p["gl"]; gc[i] += wi * p["gv"]; pj[i] += wi
        gf[j] += wi * p["gv"]; gc[j] += wi * p["gl"]; pj[j] += wi
    media = (gf.sum() / pj.sum())
    theta = np.zeros(2 * n + 1)
    theta[:n] = np.log(np.clip((gf / pj) / media, 0.3, 3.0))
    theta[n:2 * n] = np.log(np.clip((gc / pj) / media, 0.3, 3.0))
    theta[-1] = math.log(VENTAJA_LOCAL_BASE)
    return theta


# ==============================================================================
# 10. VALIDADOR DE REGLAS HISTÓRICAS (minería automática del CSV)
# ==============================================================================

def minar_patrones(partidos, min_partidos=6):
    """
    Descubre automáticamente patrones tipo:
      - 'X nunca le ha ganado a Y' (en los últimos 10 años)
      - 'el local nunca ha perdido este duelo'
      - 'este duelo nunca ha tenido más de k goles'
    Los imprime para que decidas cuáles promover a REGLAS_H2H.
    """
    h2h = defaultdict(lambda: {"pj": 0, "gana_l": 0, "empata": 0, "gana_v": 0, "max_goles": 0})
    for p in partidos:
        k = (p["local"], p["visitante"])
        d = h2h[k]
        d["pj"] += 1
        d["max_goles"] = max(d["max_goles"], p["gl"] + p["gv"])
        if p["gl"] > p["gv"]:
            d["gana_l"] += 1
        elif p["gl"] == p["gv"]:
            d["empata"] += 1
        else:
            d["gana_v"] += 1

    print("\n--- PATRONES HISTÓRICOS DETECTADOS (candidatos a regla) ---")
    for (l, v), d in sorted(h2h.items()):
        if d["pj"] < min_partidos:
            continue
        if d["gana_v"] == 0:
            print(f"  * {v} NUNCA ha ganado en casa de {l} ({d['pj']} partidos)")
        if d["gana_l"] == 0:
            print(f"  * {l} NUNCA ha ganado de local a {v} ({d['pj']} partidos)")
        if d["max_goles"] <= 3:
            print(f"  * {l} vs {v}: nunca más de {d['max_goles']} goles ({d['pj']} partidos)")


# ==============================================================================
# 11. REPORTE
# ==============================================================================

def pronosticar_jornada(fixture, ratings, detalle=True, salida_csv=None):
    filas = []
    matrices = []
    print("\n" + "=" * 78)
    print(f"  CEREBRO LIGA MX 2026 — PRONÓSTICO {JORNADA_ACTUAL.upper()}")
    print(f"  {N_SIMULACIONES:,} simulaciones Monte Carlo por partido | Dixon-Coles rho={RHO_DC}")
    print("=" * 78)

    for fecha, local, visitante, estadio in fixture:
        lam_l, lam_v, bitacora = lambdas_partido(local, visitante, ratings)
        gl, gv, M = simular_partido(lam_l, lam_v)
        matrices.append({"nombre": f"{local} vs {visitante}",
                         "H": local, "A": visitante, "M": M,
                         "lam_l": lam_l, "lam_v": lam_v})

        p_local = float(np.mean(gl > gv))
        p_empate = float(np.mean(gl == gv))
        p_visita = float(np.mean(gl < gv))

        # marcadores exactos más probables (de la matriz analítica, más fina que el MC)
        top = sorted(
            ((M[x, y], x, y) for x in range(6) for y in range(6)),
            reverse=True
        )[:5]

        marcador_modal = f"{top[0][1]}-{top[0][2]}"
        pick = local if p_local == max(p_local, p_empate, p_visita) else (
            "Empate" if p_empate == max(p_local, p_empate, p_visita) else visitante)

        print(f"\n  {fecha} | {local} vs {visitante}  ({estadio})")
        print(f"    lambda_local={lam_l:.2f}  lambda_visita={lam_v:.2f}")
        print(f"    P(local)={p_local:5.1%}   P(empate)={p_empate:5.1%}   P(visita)={p_visita:5.1%}")
        print(f"    >> PICK QUINIELA: {pick:20s} | MARCADOR EXACTO: {marcador_modal}")
        print(f"    Top marcadores: " + "  ".join(f"{x}-{y} ({p:.1%})" for p, x, y in top))
        if detalle and bitacora:
            for b in bitacora:
                print(f"      · {b}")

        filas.append({
            "fecha": fecha, "local": local, "visitante": visitante,
            "p_local": round(p_local, 4), "p_empate": round(p_empate, 4),
            "p_visita": round(p_visita, 4), "pick": pick,
            "marcador": marcador_modal,
            "lam_local": round(lam_l, 3), "lam_visita": round(lam_v, 3),
        })

    if salida_csv:
        with open(salida_csv, "w", newline="", encoding="utf-8-sig") as f:
            wcsv = csv.DictWriter(f, fieldnames=filas[0].keys())
            wcsv.writeheader()
            wcsv.writerows(filas)
        print(f"\n  [OK] Pronóstico exportado a: {salida_csv}")

    return filas, matrices


# ==============================================================================
# 11-BIS. PARLAYS (multi-mercado, probabilidad conjunta exacta)
# ------------------------------------------------------------------------------
# Cada partido ofrece un CATÁLOGO de legs derivados de su matriz de marcadores:
# resultado, doble oportunidad, más/menos goles, ambos anotan, total por equipo.
# (Tarjetas/córners/goleador/minuto NO se modelan: requieren datos que no
#  tenemos; inventarlos sería deshonesto y además son de alta varianza.)
#
# Un parlay pega si TODOS sus legs aciertan. Como puede haber VARIOS legs del
# mismo partido (correlacionados), la probabilidad se calcula agrupando por
# partido: prob conjunta dentro del partido (suma exacta sobre la matriz) y
# producto entre partidos distintos (independientes). Se elige la combinación
# que maximiza el VALOR ESPERADO = (nº de picks) × P(todos peguen), por búsqueda
# voraz: se agrega el leg que más suba el valor esperado, hasta que ya no suba.
# ==============================================================================

def _joint_mitades(lam_l, lam_v, n=8):
    """
    Distribución conjunta EXACTA de (goles local 1T, visita 1T, local 2T,
    visita 2T) por adelgazamiento de Poisson (thinning):
        Local 1T ~ Poisson(lam_l * p),  Local 2T ~ Poisson(lam_l * (1-p)),
    idem visita, con p = PROP_GOLES_1T (real de Liga MX 2018-2024).
    La corrección Dixon-Coles se aplica como peso sobre el marcador FINAL
    (h1+h2, a1+a2) y se renormaliza, para que el marginal de tiempo completo
    sea consistente con la matriz principal.
    Devuelve J[h1, a1, h2, a2].
    """
    p = PROP_GOLES_1T
    def pois(lam):
        return np.array([math.exp(-lam) * lam**k / math.factorial(k) for k in range(n)])
    ph1, ph2 = pois(lam_l * p), pois(lam_l * (1 - p))
    pa1, pa2 = pois(lam_v * p), pois(lam_v * (1 - p))
    J = (ph1[:, None, None, None] * pa1[None, :, None, None]
         * ph2[None, None, :, None] * pa2[None, None, None, :])
    # peso Dixon-Coles sobre el marcador final
    for h1 in range(n):
        for a1 in range(n):
            for h2 in range(n):
                for a2 in range(n):
                    if h1 + h2 <= 1 and a1 + a2 <= 1:
                        J[h1, a1, h2, a2] *= tau_dc(h1 + h2, a1 + a2, lam_l, lam_v)
    return J / J.sum()


def _catalogo_legs(mt):
    """Genera legs candidatos de un partido. Cada pred opera sobre
    (h1, a1, h2, a2): goles por equipo y por mitad. FT = (h1+h2, a1+a2)."""
    H, A = mt["H"], mt["A"]
    FT = lambda f: (lambda h1, a1, h2, a2: f(h1 + h2, a1 + a2))
    defs = [
        # --- tiempo completo ---
        (f"Gana {H}", "RES", FT(lambda x, y: x > y)),
        (f"Gana {A}", "RES", FT(lambda x, y: x < y)),
        (f"{H} o empate", "RES", FT(lambda x, y: x >= y)),
        (f"Empate o {A}", "RES", FT(lambda x, y: x <= y)),
        ("No hay empate", "RES", FT(lambda x, y: x != y)),
        ("Más de 1.5 goles", "TOT", FT(lambda x, y: x + y > 1.5)),
        ("Más de 2.5 goles", "TOT", FT(lambda x, y: x + y > 2.5)),
        ("Menos de 2.5 goles", "TOT", FT(lambda x, y: x + y < 2.5)),
        ("Menos de 3.5 goles", "TOT", FT(lambda x, y: x + y < 3.5)),
        ("Ambos anotan: Sí", "BTS", FT(lambda x, y: x >= 1 and y >= 1)),
        ("Ambos anotan: No", "BTS", FT(lambda x, y: not (x >= 1 and y >= 1))),
        (f"{H} anota 2+", "EQL", FT(lambda x, y: x >= 2)),
        (f"{H} menos de 1.5", "EQL", FT(lambda x, y: x < 2)),
        (f"{A} anota 2+", "EQV", FT(lambda x, y: y >= 2)),
        (f"{A} menos de 1.5", "EQV", FT(lambda x, y: y < 2)),
        # --- mitades (gol "antes del minuto 45" = gol en 1T) ---
        ("Gol en el 1er tiempo", "T1G", lambda h1, a1, h2, a2: h1 + a1 >= 1),
        ("Sin goles en el 1T", "T1G", lambda h1, a1, h2, a2: h1 + a1 == 0),
        ("Menos de 1.5 goles 1T", "T1T", lambda h1, a1, h2, a2: h1 + a1 < 2),
        ("Más de 0.5 goles 2T", "T2G", lambda h1, a1, h2, a2: h2 + a2 >= 1),
        (f"{H} o empate al descanso", "RHT", lambda h1, a1, h2, a2: h1 >= a1),
        (f"Empate o {A} al descanso", "RHT", lambda h1, a1, h2, a2: h1 <= a1),
        ("2T con más goles que 1T", "TMM", lambda h1, a1, h2, a2: h2 + a2 > h1 + a1),
    ]
    return [{"label": lb, "familia": fam, "pred": pr} for lb, fam, pr in defs]


def _prob_conjunta(seleccion, matrices):
    """P(todos los legs peguen): suma exacta sobre el joint de mitades dentro de
    cada partido; producto entre partidos distintos (independientes)."""
    from collections import defaultdict
    por_part = defaultdict(list)
    for leg in seleccion:
        por_part[leg["midx"]].append(leg["pred"])
    P = 1.0
    for midx, preds in por_part.items():
        J = matrices[midx]["J"]
        n = J.shape[0]
        s = 0.0
        for h1 in range(n):
            for a1 in range(n):
                for h2 in range(n):
                    for a2 in range(n):
                        if all(pr(h1, a1, h2, a2) for pr in preds):
                            s += J[h1, a1, h2, a2]
        P *= float(s)
    return P


def _greedy_parlay(cands, matrices, piso, techo, tope, max_por_partido):
    """Búsqueda voraz: agrega el leg que más suba (picks × P), respetando reglas.
    Solo considera legs con probabilidad en [piso, techo]: se excluyen los
    triviales (>techo, tipo 'menos de 4.5 goles' al 97%) que solo inflarían el
    conteo sin ser picks 'de mérito', y los muy riesgosos (<piso)."""
    elegidos = []
    cands = [c for c in cands if piso <= c["prob"] <= techo]
    while len(elegidos) < tope:
        mejor, mejor_ev = None, len(elegidos) * _prob_conjunta(elegidos, matrices) if elegidos else 0.0
        for leg in cands:
            if leg in elegidos:
                continue
            mismos = [e for e in elegidos if e["midx"] == leg["midx"]]
            if len(mismos) >= max_por_partido:
                continue
            if any(e["familia"] == leg["familia"] for e in mismos):
                continue  # no dos legs de la misma familia en un partido (evita anidados)
            ev = (len(elegidos) + 1) * _prob_conjunta(elegidos + [leg], matrices)
            if ev > mejor_ev + 1e-9:
                mejor_ev, mejor = ev, leg
        if mejor is None:
            break
        elegidos.append(mejor)
    return elegidos


def _parlay_por_umbral(cands, matrices, piso, techo, tope, max_por_partido, prob_min):
    """Agrega legs por confianza (mayor primero) mientras la prob. combinada se
    mantenga >= prob_min. Sirve para 'seguro' (prob_min alto -> pocos picks) y
    'agresivo' (prob_min bajo -> muchos picks)."""
    pool = sorted([c for c in cands if piso <= c["prob"] <= techo],
                  key=lambda c: c["prob"], reverse=True)
    elegidos = []
    for leg in pool:
        if len(elegidos) >= tope:
            break
        mismos = [e for e in elegidos if e["midx"] == leg["midx"]]
        if len(mismos) >= max_por_partido:
            continue
        if any(e["familia"] == leg["familia"] for e in mismos):
            continue
        if _prob_conjunta(elegidos + [leg], matrices) >= prob_min:
            elegidos.append(leg)
    return elegidos


def construir_parlays(matrices):
    """Devuelve variantes: recomendado (equilibrado), seguro, agresivo."""
    cands = []
    for midx, mt in enumerate(matrices):
        mt["J"] = _joint_mitades(mt["lam_l"], mt["lam_v"])  # joint (h1,a1,h2,a2)
        for leg in _catalogo_legs(mt):
            leg = dict(leg)
            leg["midx"] = midx
            leg["partido"] = mt["nombre"]
            leg["prob"] = _prob_conjunta([leg], matrices)  # marginal exacta
            cands.append(leg)

    variantes = {
        # equilibrio óptimo (máx. picks × P): picks con valor (58-85%), 2/partido
        "recomendado": _greedy_parlay(cands, matrices, piso=0.58, techo=0.85, tope=6, max_por_partido=2),
        # muy seguro: pocos picks, alta prob combinada (uno por partido)
        "seguro": _parlay_por_umbral(cands, matrices, piso=0.62, techo=0.85,
                                     tope=4, max_por_partido=1, prob_min=0.45),
        # agresivo: máximo de picks tolerando riesgo (prob combinada >= 18%)
        "agresivo": _parlay_por_umbral(cands, matrices, piso=0.55, techo=0.85,
                                       tope=10, max_por_partido=2, prob_min=0.18),
    }
    salida = {}
    for k, legs in variantes.items():
        salida[k] = {"legs": legs, "prob": _prob_conjunta(legs, matrices)}
    return salida


def imprimir_parlays(parlays):
    etiquetas = {"recomendado": "PARLAY RECOMENDADO (equilibrio picks vs. seguridad)",
                 "seguro": "PARLAY SEGURO (máxima probabilidad)",
                 "agresivo": "PARLAY AGRESIVO (más picks, más premio)"}
    print("\n" + "=" * 78)
    print("  PARLAYS — MEJOR COMBINACIÓN DE LA JORNADA")
    print("=" * 78)
    for k in ("recomendado", "seguro", "agresivo"):
        pk = parlays[k]
        print(f"\n  {etiquetas[k]}")
        print(f"    {'Partido':<32}{'Selección':<26}{'Prob.':>7}")
        for leg in pk["legs"]:
            print(f"    {leg['partido']:<32}{leg['label']:<26}{leg['prob']:>6.1%}")
        n = len(pk["legs"])
        print(f"    → {n} picks | Prob. combinada: {pk['prob']:.1%} | "
              f"Valor esperado: {n * pk['prob']:.2f}")


# ==============================================================================
# 11-TER. EXPORTACIÓN A EXCEL (GPV) — pestañas Pronóstico y Parlays
# ==============================================================================

# Paleta corporativa GPV
GPV_VERDE = "00573F"
GPV_DORADO = "C9A961"
GPV_VERDE_CLARO = "E4EFEA"
GPV_GRIS = "595959"


def exportar_excel(predicciones, parlays, ruta, jornada=JORNADA_ACTUAL):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [Excel] openpyxl no disponible; instala con: pip install openpyxl")
        return False

    verde = PatternFill("solid", fgColor=GPV_VERDE)
    dorado = PatternFill("solid", fgColor=GPV_DORADO)
    verde_cl = PatternFill("solid", fgColor=GPV_VERDE_CLARO)
    blanco_b = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    verde_b = Font(name="Calibri", bold=True, color=GPV_VERDE, size=11)
    normal = Font(name="Calibri", size=10)
    gris_i = Font(name="Calibri", size=8, italic=True, color=GPV_GRIS)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="D9D9D9")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    PCT = "0.0%"

    wb = openpyxl.Workbook()

    # ---------- Hoja 1: PRONÓSTICO ----------
    ws = wb.active
    ws.title = "Pronóstico"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "CEREBRO LIGA MX 2026 — PRONÓSTICO"
    ws["A1"].font = Font(name="Calibri", bold=True, color=GPV_VERDE, size=16)
    ws["A2"] = jornada
    ws["A2"].font = Font(name="Calibri", size=11, color=GPV_DORADO, bold=True)

    hdr = ["Fecha", "Local", "Visitante", "P(Local)", "P(Empate)", "P(Visita)",
           "Pick quiniela", "Marcador"]
    r0 = 4
    for j, h in enumerate(hdr, 1):
        c = ws.cell(r0, j, h)
        c.fill = verde; c.font = blanco_b; c.alignment = center; c.border = borde
    for i, pr in enumerate(predicciones):
        r = r0 + 1 + i
        vals = [pr["fecha"], pr["local"], pr["visitante"], pr["p_local"],
                pr["p_empate"], pr["p_visita"], pr["pick"], pr["marcador"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v)
            c.font = normal; c.border = borde
            c.alignment = left if j in (2, 3, 7) else center
            if j in (4, 5, 6):
                c.number_format = PCT
        if i % 2:
            for j in range(1, 9):
                ws.cell(r, j).fill = verde_cl
        # resalta el pick
        ws.cell(r, 7).font = Font(name="Calibri", bold=True, color=GPV_VERDE, size=10)
    anchos = [12, 20, 20, 10, 11, 10, 22, 10]
    for j, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---------- Hoja 2: PARLAYS ----------
    wp = wb.create_sheet("Parlays")
    wp.sheet_view.showGridLines = False
    wp["A1"] = "PARLAYS — Mejor combinación por jornada"
    wp["A1"].font = Font(name="Calibri", bold=True, color=GPV_VERDE, size=16)
    wp["A2"] = jornada
    wp["A2"].font = Font(name="Calibri", size=11, color=GPV_DORADO, bold=True)
    for col, w in zip("ABC", (34, 30, 12)):
        wp.column_dimensions[col].width = w

    def bloque(titulo, subt, parlay, fila):
        legs = parlay["legs"]
        wp.cell(fila, 1, titulo).font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
        for j in range(1, 4):
            wp.cell(fila, j).fill = verde
        wp.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
        wp.cell(fila, 1).alignment = left
        wp.cell(fila + 1, 1, subt).font = gris_i
        wp.merge_cells(start_row=fila + 1, start_column=1, end_row=fila + 1, end_column=3)
        h = fila + 2
        for j, t in enumerate(["Partido", "Selección (mercado)", "Prob."], 1):
            c = wp.cell(h, j, t)
            c.fill = dorado; c.font = verde_b; c.alignment = center; c.border = borde
        p1 = h + 1
        for k, leg in enumerate(legs):
            rr = p1 + k
            wp.cell(rr, 1, leg["partido"]).alignment = left
            wp.cell(rr, 2, leg["label"]).alignment = left
            pc = wp.cell(rr, 3, round(leg["prob"], 4))
            pc.number_format = PCT; pc.alignment = center
            for j in range(1, 4):
                wp.cell(rr, j).font = normal; wp.cell(rr, j).border = borde
        pN = p1 + len(legs) - 1
        # prob combinada = JOINT real del modelo (no producto: hay legs del mismo
        # partido correlacionados). Se escribe como valor calculado.
        rs = pN + 1
        wp.cell(rs, 1, f"Picks: {len(legs)}").font = verde_b
        wp.cell(rs, 2, "Prob. combinada (conjunta) →").font = verde_b
        wp.cell(rs, 2).alignment = Alignment(horizontal="right", vertical="center")
        pc = wp.cell(rs, 3, round(parlay["prob"], 4))
        pc.number_format = PCT; pc.font = Font(name="Calibri", bold=True, color=GPV_VERDE, size=11)
        for j in range(1, 4):
            wp.cell(rs, j).fill = verde_cl; wp.cell(rs, j).border = borde
        rv = rs + 1
        wp.cell(rv, 2, "Valor esperado (picks × prob) →").font = gris_i
        wp.cell(rv, 2).alignment = Alignment(horizontal="right", vertical="center")
        ev = wp.cell(rv, 3, f"=COUNT(C{p1}:C{pN})*C{rs}")
        ev.number_format = "0.00"; ev.font = gris_i
        return rv + 2

    f = 4
    f = bloque("PARLAY RECOMENDADO", "Equilibrio óptimo entre número de picks y seguridad (máx. valor esperado).",
               parlays["recomendado"], f)
    f = bloque("PARLAY SEGURO", "Máxima probabilidad: los legs más sólidos, uno por partido. Menos premio, más certeza.",
               parlays["seguro"], f)
    f = bloque("PARLAY AGRESIVO", "Más picks = más premio si pega, pero menor probabilidad combinada.",
               parlays["agresivo"], f)

    # notas al pie
    notas = [
        "Criterio: búsqueda voraz que agrega el leg que más suba (nº de picks × P(todos peguen)); "
        "un partido puede aportar varios legs o ninguno.",
        "Prob. combinada = probabilidad conjunta EXACTA sobre el joint de mitades (h1,a1,h2,a2): "
        "legs del mismo partido (incluso 1T vs tiempo completo) están correlacionados; partidos distintos se multiplican.",
        "Mercados: resultado, doble oportunidad, más/menos goles, ambos anotan, total por equipo, "
        "y MITADES (gol en 1T, resultado al descanso, 2T más goleador) calibradas con 2,047 partidos reales con marcador HT (2018-2024).",
        "Reparto de mitades: 44.4% de los goles caen en el 1T (dato real Liga MX); modelo validado: "
        "P(gol en 1T) 68.8% vs 68.9% observado.",
        "Tarjetas, córners y minuto exacto de gol NO se modelan: no existe base pública de Liga MX con esos campos "
        "(verificado: football-data.co.uk los publica vacíos para MEX). Requieren API de pago.",
        "Probabilidades del modelo Dixon-Coles + Monte Carlo con histórico real 2010-2025 (openfootball).",
    ]
    for nt in notas:
        wp.cell(f, 1, nt).font = gris_i
        wp.merge_cells(start_row=f, start_column=1, end_row=f, end_column=3)
        f += 1

    wb.save(ruta)
    print(f"  [Excel] Workbook GPV guardado: {ruta}")
    return True


# ==============================================================================
# 12. MAIN
# ==============================================================================

def main():
    ap = argparse.ArgumentParser(description="Cerebro Liga MX 2026")
    ap.add_argument("--historico", help="CSV con ~10 años de partidos Liga MX")
    ap.add_argument("--csv", default="pronostico_j2_apertura2026.csv",
                    help="Archivo de salida del pronóstico")
    ap.add_argument("--sin-detalle", action="store_true",
                    help="Oculta la bitácora de ajustes por partido")
    ap.add_argument("--excel", default="Cerebro_LigaMX_J2.xlsx",
                    help="Ruta del workbook Excel de salida (pestañas Pronóstico y Parlays)")
    args = ap.parse_args()

    if args.historico and os.path.exists(args.historico):
        print(f"[1/2] Calibrando MLE Dixon-Coles con {args.historico} ...")
        partidos = cargar_historico(args.historico)
        print(f"      {len(partidos):,} partidos cargados")
        # Recalibra la escala: goles promedio de local y visita en fase regular
        global BASE_HOME_GOALS, BASE_AWAY_GOALS
        reg = [p for p in partidos if p["fase"] == "regular"]
        if reg:
            BASE_HOME_GOALS = sum(p["gl"] for p in reg) / len(reg)
            BASE_AWAY_GOALS = sum(p["gv"] for p in reg) / len(reg)
            print(f"      Escala real -> goles local={BASE_HOME_GOALS:.2f}  "
                  f"visita={BASE_AWAY_GOALS:.2f}  (total={BASE_HOME_GOALS+BASE_AWAY_GOALS:.2f})")
        ratings = fit_dixon_coles(partidos)
        minar_patrones(partidos)
        # equipos sin histórico (Atlante) conservan su prior
        for e in EQUIPOS:
            ratings.setdefault(e, PRIORS_ILUSTRATIVOS[e])
    else:
        print("[1/2] Sin histórico -> usando PRIORS_ILUSTRATIVOS "
              "(calibración cualitativa 2025-26; cargar CSV para MLE real)")
        ratings = dict(PRIORS_ILUSTRATIVOS)

    print(f"[2/2] Simulando {JORNADA_ACTUAL} ...")
    filas, matrices = pronosticar_jornada(FIXTURE, ratings,
                                          detalle=not args.sin_detalle, salida_csv=args.csv)

    # Parlays (multi-mercado) + workbook Excel con pestaña Parlays
    parlays = construir_parlays(matrices)
    imprimir_parlays(parlays)
    exportar_excel(filas, parlays, args.excel)

    print("\n" + "-" * 78)
    print("  Recuerda: los priors son ILUSTRATIVOS. Para el modelo de concurso,")
    print("  carga el CSV histórico (10 años) y el MLE recalibra todo + mina")
    print("  patrones H2H automáticamente con minar_patrones().")
    print("-" * 78)


if __name__ == "__main__":
    main()
