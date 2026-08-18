# -*- coding: utf-8 -*-
"""
Hoja de resumen de retrocesiones de America del Sur.

Arma una pestana ejecutiva con lo que pide la linea para la Junta de Planeacion:
asegurado, pais, cedente, corredor, prima al 100%, % de retrocesion y fee para
Patria, y atras las columnas de soporte para poder amarrar cada cifra.

Se usa desde los scripts de consulta y desde el convertidor:
    from resumen_sudamerica import construir_resumen, agregar_hoja_resumen
"""

import pandas as pd
from openpyxl.styles import Alignment, Color, Font, PatternFill, colors
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paises que se consideran America del Sur
# ---------------------------------------------------------------------------
# 'Quito' esta dado de alta como pais en el catalogo; se toma como Ecuador.
PAISES_SUDAMERICA = [
    'Argentina', 'Bolivia', 'Brasil', 'Chile', 'Colombia', 'Ecuador', 'Quito',
    'Guyana', 'Guyana Francesa', 'Paraguay', 'Perú', 'Surinam', 'Suriname',
    'Uruguay', 'Venezuela',
]


def _sin_acentos(texto):
    reemplazos = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ü': 'u',
                  'ñ': 'n', 'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
                  'Ü': 'U', 'Ñ': 'N'}
    texto = str(texto)
    for viejo, nuevo in reemplazos.items():
        texto = texto.replace(viejo, nuevo)
    return texto.strip().lower()


_SUDAMERICA = {_sin_acentos(p) for p in PAISES_SUDAMERICA}


def es_sudamerica(paises):
    """True si la celda de paises (uno o varios, separados por coma) trae algun
    pais de America del Sur. Trabaja sobre la version con nombres del archivo."""
    if paises is None:
        return False
    try:
        if pd.isna(paises):
            return False
    except (TypeError, ValueError):
        pass
    partes = [_sin_acentos(p) for p in str(paises).split(',')]
    return any(p in _SUDAMERICA for p in partes)


# ---------------------------------------------------------------------------
# Campos que se toman de la consulta
# nombre interno -> (encabezado en el Excel, bloque en el que vive)
# ---------------------------------------------------------------------------
CAMPOS = {
    # Bloque del tomado (el negocio original)
    'Tipo_Reaseguro':             ('Tipo Reaseguro', 'tomado'),
    'Cedente':                    ('Cedente', 'tomado'),
    'NoCedente':                  ('No. Cedente', 'tomado'),
    'Corredor_':                  ('Corredor', 'tomado'),
    'NoContrato_':                ('No. Contrato', 'tomado'),
    'NoOferta':                   ('No. Oferta', 'tomado'),
    'Endoso':                     ('Endoso', 'tomado'),
    'Año_Vigencia_':              ('Año Vigencia', 'tomado'),
    'Referencia_Original':        ('Referencia Original', 'tomado'),
    'Inicio_Vigencia_':           ('Inicio Vigencia', 'tomado'),
    'Fin_Vigencia_':              ('Fin Vigencia', 'tomado'),
    'PrcRetro':                   ('PrcRetro', 'tomado'),
    'Ramos_Cubiertos_':           ('Ramos Cubiertos', 'tomado'),
    'Paises_Cubiertos':           ('Paises Cubiertos', 'tomado'),
    'Asegurado_Nombre_OFD':       ('Asegurado Nombre (Daños)', 'tomado'),
    'Afianzado_Nombre_OFF':       ('Afianzado Nombre (Fianzas)', 'tomado'),
    'Asegurado_Fisico_Nombre_OFV': ('Asegurado Físico Nombre (Vida Ind)', 'tomado'),
    'Grupo_Asegurado_Nombre_OFVG': ('Grupo Asegurado Nombre (Vida Gpo)', 'tomado'),
    'Prc_Patria_TP':              ('% Patria (Prop)', 'tomado'),
    'PriEst_100_TNP':             ('Prima Esperada 100% (No Prop)', 'tomado'),
    'PriEst_Patria_TNP':          ('Prima Esperada Patria (No Prop)', 'tomado'),
    'SA_Unica_100_OFD':           ('Suma Asegurada Única 100% (Daños)', 'tomado'),
    'Aceptacion_Patria_Prc_OFD':  ('% Aceptación Patria (Daños)', 'tomado'),
    'Aceptacion_Patria_Mnt_OFD':  ('Monto Aceptación Patria (Daños)', 'tomado'),
    'EPI_100_OFF':                ('EPI 100 (Fianzas)', 'tomado'),
    'Monedas_Movimiento':         ('Monedas del Movimiento', 'tomado'),
    'PRIMAS_DEL_REASEGURO_TOMADO': ('PRIMAS_DEL_REASEGURO_TOMADO', 'tomado'),
    'PRIMAS_RETROCEDIDAS':        ('PRIMAS_RETROCEDIDAS', 'tomado'),
    # Bloque del cedido (el contrato de retrocesion)
    'Corredor':                   ('Corredor', 'cedido'),
    'NoContrato':                 ('No. Contrato', 'cedido'),
    'Ident_Contrato':             ('Ident Contrato', 'cedido'),
    'Año_Vigencia':               ('Año Vigencia', 'cedido'),
    'Nombre':                     ('Nombre', 'cedido'),
    'ComPrimB':                   ('ComPrimBruta', 'cedido'),
    'ComPrimNeta':                ('ComPrimNeta', 'cedido'),
}


def _normalizar(df):
    """Deja el DataFrame con los nombres internos, venga de la consulta (nombres
    internos) o de leer el Excel ya generado (encabezados de la hoja).

    Los encabezados se repiten entre cedido y tomado (Corredor, No. Contrato,
    Año Vigencia...), asi que se distingue por su posicion: el bloque del tomado
    empieza en la columna 'Tipo Reaseguro'.
    """
    columnas = list(df.columns)
    corte = 0
    for etiqueta in ('Tipo_Reaseguro', 'Tipo Reaseguro'):
        if etiqueta in columnas:
            corte = columnas.index(etiqueta)
            break

    salida = pd.DataFrame(index=df.index)
    for interno, (encabezado, bloque) in CAMPOS.items():
        # Primero se busca por el nombre interno de la consulta y, si no esta,
        # por el encabezado que lleva la columna en el Excel.
        posiciones = [i for i, c in enumerate(columnas) if c == interno]
        if not posiciones:
            posiciones = [i for i, c in enumerate(columnas) if c == encabezado]
        if len(posiciones) > 1:
            # Encabezado repetido en cedido y tomado: se toma el del bloque que toca.
            candidatas = [i for i in posiciones if (i >= corte) == (bloque == 'tomado')]
            posiciones = candidatas or posiciones
        salida[interno] = df.iloc[:, posiciones[0]] if posiciones else None
    return salida


def _numero(valor):
    """Convierte a float lo que se pueda; lo demas queda como None."""
    if valor is None:
        return None
    try:
        if pd.isna(valor):
            return None
    except (TypeError, ValueError):
        pass
    try:
        return float(valor)
    except (TypeError, ValueError):
        return None


def _primero_con_texto(*valores):
    for valor in valores:
        if valor is None:
            continue
        try:
            if pd.isna(valor):
                continue
        except (TypeError, ValueError):
            pass
        texto = str(valor).strip()
        if texto and texto.lower() != 'nan':
            return texto
    return None


def _prima_al_100(fila, prima_patria):
    """Prima al 100% y de donde salio.

    Orden de preferencia:
      1. No proporcional: la prima esperada al 100% viene en la consulta.
      2. Fianzas: EPI al 100%.
      3. Facultativo danos: se estima con la prima contable y el % de aceptacion.
      4. Proporcional: se estima con la prima contable y el % de participacion.
    """
    prima_np = _numero(fila.get('PriEst_100_TNP'))
    if prima_np:
        return prima_np, 'Prima esperada 100% (No Prop)'

    epi = _numero(fila.get('EPI_100_OFF'))
    if epi:
        return epi, 'EPI 100% (Fianzas)'

    aceptacion = _numero(fila.get('Aceptacion_Patria_Prc_OFD'))
    if aceptacion and prima_patria:
        return prima_patria * 100.0 / aceptacion, 'Estimada: prima contable / % aceptación Patria'

    participacion = _numero(fila.get('Prc_Patria_TP'))
    if participacion and prima_patria:
        return prima_patria * 100.0 / participacion, 'Estimada: prima contable / % Patria (Prop)'

    if prima_patria:
        return None, 'Sin % de participación para escalar a 100%'
    return None, 'Sin prima registrada'


COLUMNAS_RESUMEN = [
    # Lo que pidio la linea
    'Asegurado', 'País', 'Cedente', 'Corredor (Tomado)',
    'Prima al 100%', '% Retrocesión', '% Fee Patria', 'Fee Patria',
    # Soporte
    'Base de la Prima al 100%', 'Prima Tomada Patria', 'Base de la Prima Patria',
    'Prima Retrocedida (estimada)', 'Prima Retrocedida (contable, por contrato)',
    'Tipo Reaseguro', 'Referencia Original', 'Ramos Cubiertos',
    'Año Vigencia', 'Inicio Vigencia', 'Fin Vigencia',
    'Suma Asegurada 100% (Daños)', '% Aceptación Patria (Daños)',
    'Monto Aceptación Patria (Daños)', '% Patria (Prop)',
    'Prima Esperada Patria (No Prop)',
    'No. Cedente', 'No. Contrato (Tomado)', 'No. Oferta', 'Endoso',
    'Contrato de Retro', 'Corredor (Retro)', 'No. Contrato (Retro)',
    'Ident Contrato (Retro)', 'Año Vigencia (Retro)',
    'Monedas del Movimiento', 'Llave contable', 'Primer renglón de la llave',
    'Riesgos en la llave contable',
]

# Columnas con formato de importe y de porcentaje en la hoja
COLUMNAS_IMPORTE = {
    'Prima al 100%', 'Fee Patria', 'Prima Tomada Patria',
    'Prima Retrocedida (estimada)', 'Prima Retrocedida (contable, por contrato)',
    'Suma Asegurada 100% (Daños)', 'Monto Aceptación Patria (Daños)',
    'Prima Esperada Patria (No Prop)',
}
COLUMNAS_PORCENTAJE = {
    '% Retrocesión', '% Fee Patria', '% Aceptación Patria (Daños)', '% Patria (Prop)',
}
COLUMNAS_FECHA = {'Inicio Vigencia', 'Fin Vigencia'}


def construir_resumen(df, solo_facultativo=False):
    """Regresa el DataFrame del resumen de America del Sur.

    df: la consulta ya traducida a nombres (CedF_Nombres) o el contenido de la
        hoja del archivo _nombres ya generado.
    solo_facultativo: si es True deja unicamente los renglones cuyo negocio
        tomado es facultativo. Por omision se dejan todos los renglones del
        archivo, que de por si son retrocesiones facultativas, y se puede
        filtrar con la columna 'Tipo Reaseguro'.
    """
    base = _normalizar(df)
    base = base[base['Paises_Cubiertos'].apply(es_sudamerica)]

    if solo_facultativo:
        tipos = base['Tipo_Reaseguro'].astype(str).str.upper()
        base = base[tipos.str.startswith('FACULTATIVO')]

    renglones = []
    llaves_vistas = set()
    riesgos_por_llave = {}

    for _, fila in base.iterrows():
        # La prima de Patria sale del movimiento contable; si el contrato todavia
        # no tiene movimientos (tipico del no proporcional recien renovado), se
        # usa la prima esperada de Patria que trae la consulta.
        prima_contable = _numero(fila.get('PRIMAS_DEL_REASEGURO_TOMADO'))
        prima_contable = abs(prima_contable) if prima_contable else None
        prima_esperada_patria = _numero(fila.get('PriEst_Patria_TNP'))

        if prima_contable:
            prima_tomada, base_patria = prima_contable, 'Movimiento contable'
        elif prima_esperada_patria:
            prima_tomada, base_patria = prima_esperada_patria, 'Prima esperada Patria (No Prop)'
        else:
            prima_tomada, base_patria = None, 'Sin dato'

        prima_100, base_prima = _prima_al_100(fila, prima_tomada)

        prc_retro = _numero(fila.get('PrcRetro'))
        prima_retro = prima_tomada * prc_retro / 100.0 if (prima_tomada and prc_retro) else None

        fee_prc = _numero(fila.get('ComPrimB'))
        fee_monto = prima_retro * fee_prc / 100.0 if (prima_retro and fee_prc) else None

        llave = '-'.join(str(fila.get(c)) for c in
                         ('Corredor_', 'NoCedente', 'NoContrato_', 'Año_Vigencia_'))
        primera = llave not in llaves_vistas
        llaves_vistas.add(llave)

        # El movimiento contable esta a nivel contrato-anio: si la llave agrupa
        # varias ofertas o asegurados, la prima NO es de un solo riesgo.
        riesgos_por_llave.setdefault(llave, set()).add(
            (fila.get('NoOferta'), fila.get('Endoso'),
             _primero_con_texto(fila.get('Asegurado_Nombre_OFD'),
                                fila.get('Afianzado_Nombre_OFF'),
                                fila.get('Asegurado_Fisico_Nombre_OFV'),
                                fila.get('Grupo_Asegurado_Nombre_OFVG'),
                                fila.get('Referencia_Original'))))

        renglones.append({
            'Asegurado': _primero_con_texto(fila.get('Asegurado_Nombre_OFD'),
                                            fila.get('Afianzado_Nombre_OFF'),
                                            fila.get('Asegurado_Fisico_Nombre_OFV'),
                                            fila.get('Grupo_Asegurado_Nombre_OFVG')),
            'País': fila.get('Paises_Cubiertos'),
            'Cedente': fila.get('Cedente'),
            'Corredor (Tomado)': fila.get('Corredor_'),
            'Prima al 100%': prima_100,
            '% Retrocesión': prc_retro,
            '% Fee Patria': fee_prc,
            'Fee Patria': fee_monto,
            'Base de la Prima al 100%': base_prima,
            'Prima Tomada Patria': prima_tomada,
            'Base de la Prima Patria': base_patria,
            'Prima Retrocedida (estimada)': prima_retro,
            'Prima Retrocedida (contable, por contrato)': _numero(fila.get('PRIMAS_RETROCEDIDAS')),
            'Tipo Reaseguro': fila.get('Tipo_Reaseguro'),
            'Referencia Original': fila.get('Referencia_Original'),
            'Ramos Cubiertos': fila.get('Ramos_Cubiertos_'),
            'Año Vigencia': fila.get('Año_Vigencia_'),
            'Inicio Vigencia': fila.get('Inicio_Vigencia_'),
            'Fin Vigencia': fila.get('Fin_Vigencia_'),
            'Suma Asegurada 100% (Daños)': _numero(fila.get('SA_Unica_100_OFD')),
            '% Aceptación Patria (Daños)': _numero(fila.get('Aceptacion_Patria_Prc_OFD')),
            'Monto Aceptación Patria (Daños)': _numero(fila.get('Aceptacion_Patria_Mnt_OFD')),
            '% Patria (Prop)': _numero(fila.get('Prc_Patria_TP')),
            'Prima Esperada Patria (No Prop)': _numero(fila.get('PriEst_Patria_TNP')),
            'No. Cedente': fila.get('NoCedente'),
            'No. Contrato (Tomado)': fila.get('NoContrato_'),
            'No. Oferta': fila.get('NoOferta'),
            'Endoso': fila.get('Endoso'),
            'Contrato de Retro': fila.get('Nombre'),
            'Corredor (Retro)': fila.get('Corredor'),
            'No. Contrato (Retro)': fila.get('NoContrato'),
            'Ident Contrato (Retro)': fila.get('Ident_Contrato'),
            'Año Vigencia (Retro)': fila.get('Año_Vigencia'),
            'Monedas del Movimiento': fila.get('Monedas_Movimiento'),
            'Llave contable': llave,
            'Primer renglón de la llave': 'Sí' if primera else 'No',
            'Riesgos en la llave contable': None,   # se llena al final
        })

    for renglon in renglones:
        renglon['Riesgos en la llave contable'] = len(
            riesgos_por_llave.get(renglon['Llave contable'], ()))

    resumen = pd.DataFrame(renglones, columns=COLUMNAS_RESUMEN)
    if len(resumen):
        resumen = resumen.sort_values(
            by=['Año Vigencia', 'Prima al 100%'],
            ascending=[False, False], na_position='last').reset_index(drop=True)
    return resumen


NOTA = ("RETROCESIONES DE AMÉRICA DEL SUR  |  Importes en moneda nacional, acumulados de los movimientos "
        "contables.  |  La prima al 100% del facultativo y del proporcional es estimada: prima contable ÷ % de "
        "participación de Patria (ver 'Base de la Prima al 100%').  |  El movimiento contable es por contrato-año, "
        "no por riesgo: cuando 'Riesgos en la llave contable' es mayor a 1, la prima corresponde a todos esos "
        "riesgos juntos.  |  Para sumar sin duplicar, filtrar 'Primer renglón de la llave' = Sí.")


def agregar_hoja_resumen(libro, resumen, titulo="América del Sur"):
    """Agrega la hoja del resumen al libro de openpyxl, con el mismo formato
    visual del resto del archivo."""
    if titulo in libro.sheetnames:
        del libro[titulo]
    hoja = libro.create_sheet(title=titulo)

    verde = Color(indexed=17)
    estilo = Font(name='Arial', size=9, bold=True, color=colors.WHITE)
    relleno = PatternFill(fill_type="solid", start_color=verde, end_color=verde)

    hoja.cell(row=1, column=1).value = NOTA
    hoja.cell(row=1, column=1).font = Font(name='Arial', size=9, bold=True, color=colors.WHITE)
    hoja.cell(row=1, column=1).fill = relleno
    hoja.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(COLUMNAS_RESUMEN), 1))
    hoja.row_dimensions[1].height = 28

    for columna, encabezado in enumerate(COLUMNAS_RESUMEN, start=1):
        celda = hoja.cell(row=2, column=columna)
        celda.value = encabezado
        celda.font = estilo
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="left", wrap_text=True)

    for i, (_, fila) in enumerate(resumen.iterrows(), start=3):
        for columna, encabezado in enumerate(COLUMNAS_RESUMEN, start=1):
            valor = fila[encabezado]
            if valor is not None and not isinstance(valor, str):
                try:
                    if pd.isna(valor):
                        valor = None
                except (TypeError, ValueError):
                    pass
            celda = hoja.cell(row=i, column=columna)
            celda.value = valor
            if encabezado in COLUMNAS_IMPORTE:
                celda.number_format = '#,##0.00'
            elif encabezado in COLUMNAS_PORCENTAJE:
                celda.number_format = '0.00"%"'
            elif encabezado in COLUMNAS_FECHA:
                celda.number_format = 'DD/MM/YYYY'

    anchos = {'Asegurado': 38, 'País': 26, 'Cedente': 34, 'Corredor (Tomado)': 30,
              'Prima al 100%': 18, '% Retrocesión': 13, '% Fee Patria': 12,
              'Fee Patria': 16, 'Base de la Prima al 100%': 34,
              'Prima Tomada Patria': 18, 'Base de la Prima Patria': 26,
              'Prima Retrocedida (estimada)': 20,
              'Prima Retrocedida (contable, por contrato)': 22,
              'Tipo Reaseguro': 20, 'Referencia Original': 34, 'Ramos Cubiertos': 26,
              'Inicio Vigencia': 11, 'Fin Vigencia': 11, 'Contrato de Retro': 30,
              'Corredor (Retro)': 30, 'Monedas del Movimiento': 16,
              'Llave contable': 22, 'Primer renglón de la llave': 12,
              'Riesgos en la llave contable': 14}
    for columna, encabezado in enumerate(COLUMNAS_RESUMEN, start=1):
        hoja.column_dimensions[get_column_letter(columna)].width = anchos.get(encabezado, 14)

    ultima = get_column_letter(len(COLUMNAS_RESUMEN))
    hoja.auto_filter.ref = f"A2:{ultima}2"
    hoja.freeze_panes = 'E3'
    hoja.sheet_view.showGridLines = False
    return hoja


# ---------------------------------------------------------------------------
# Hoja de America del Sur: copia de la hoja de datos con el filtro de Excel
# ---------------------------------------------------------------------------
def _indice_columna_pais(hoja, encabezado_pais, fila_encabezados):
    encabezados = [celda.value for celda in hoja[fila_encabezados]]
    if encabezado_pais not in encabezados:
        raise ValueError(f"No se encontro la columna '{encabezado_pais}' en la hoja {hoja.title}.")
    return encabezados.index(encabezado_pais)   # base 0, igual que el filtro


def _copia_compacta(libro, origen, titulo, columna_pais, fila_encabezados):
    """Copia la hoja quedandose solo con los renglones de America del Sur.

    Se usa cuando la hoja es muy grande y duplicarla completa haria el archivo
    inmanejable. El formato queda igual; lo que cambia es que los renglones de
    otras regiones no se copian (siguen en la hoja original).
    """
    copia = libro.create_sheet(title=titulo)

    for letra, dimension in origen.column_dimensions.items():
        destino = copia.column_dimensions[letra]
        destino.width = dimension.width
        destino.hidden = dimension.hidden

    # max_row / max_column se recalculan recorriendo la hoja, asi que se leen
    # una sola vez y no dentro del ciclo.
    ultima_fila = origen.max_row
    ultima_columna = origen.max_column

    fila_destino = 0
    copiadas = 0
    for fila_origen in range(1, ultima_fila + 1):
        es_encabezado = fila_origen <= fila_encabezados
        if not es_encabezado:
            valor_pais = origen.cell(row=fila_origen, column=columna_pais + 1).value
            if not es_sudamerica(valor_pais):
                continue
            copiadas += 1
        fila_destino += 1
        for columna in range(1, ultima_columna + 1):
            celda_origen = origen.cell(row=fila_origen, column=columna)
            celda_destino = copia.cell(row=fila_destino, column=columna)
            celda_destino.value = celda_origen.value
            if celda_origen.has_style:
                celda_destino._style = celda_origen._style
        if origen.row_dimensions[fila_origen].height is not None:
            copia.row_dimensions[fila_destino].height = origen.row_dimensions[fila_origen].height

    # Las bandas de encabezado se combinan al final: una celda combinada es de
    # solo lectura y no dejaria copiar los valores.
    for rango in origen.merged_cells.ranges:
        if rango.max_row <= fila_encabezados:
            copia.merge_cells(str(rango))

    return copia, copiadas


def agregar_hoja_filtrada(libro, hoja_origen='Sheet1', titulo='América del Sur',
                          encabezado_pais='Paises Cubiertos', fila_encabezados=2,
                          solo_renglones_filtrados=False):
    """Copia la hoja de la consulta y le deja aplicado el filtro de Excel por
    los paises de America del Sur.

    Queda exactamente igual que la hoja original (mismos encabezados, colores,
    anchos y formatos); lo unico que cambia es que la columna de paises trae el
    filtro puesto y los renglones que no son de America del Sur salen ocultos.
    El filtro es de Excel, asi que se puede quitar o ampliar desde la flechita
    de la columna sin perder informacion.

    solo_renglones_filtrados=True copia unicamente los renglones de la region en
    lugar de ocultar el resto. Se ve igual y el archivo pesa mucho menos, pero en
    esa hoja ya no se puede quitar el filtro para ver otras regiones (para eso
    esta la hoja original, que sigue completa). Conviene para el proporcional,
    que trae decenas de miles de renglones.
    """
    if titulo in libro.sheetnames:
        del libro[titulo]

    origen = libro[hoja_origen] if hoja_origen in libro.sheetnames else libro.worksheets[0]
    columna = _indice_columna_pais(origen, encabezado_pais, fila_encabezados)

    if solo_renglones_filtrados:
        copia, visibles = _copia_compacta(libro, origen, titulo, columna, fila_encabezados)
    else:
        copia = libro.copy_worksheet(origen)
        copia.title = titulo
        visibles = None

    # copy_worksheet no arrastra la vista ni el filtro: se replican a mano
    copia.sheet_view.showGridLines = origen.sheet_view.showGridLines
    copia.freeze_panes = origen.freeze_panes
    copia.auto_filter.ref = origen.auto_filter.ref

    valores = set()
    contadas = 0
    for fila in range(fila_encabezados + 1, copia.max_row + 1):
        valor = copia.cell(row=fila, column=columna + 1).value
        if es_sudamerica(valor):
            valores.add(str(valor))
            contadas += 1
        elif not solo_renglones_filtrados:
            copia.row_dimensions[fila].hidden = True
    if visibles is None:
        visibles = contadas

    # Filtro de Excel sobre la columna de paises, con los valores de la region
    copia.auto_filter.add_filter_column(columna, sorted(valores), blank=False)
    return copia, visibles
