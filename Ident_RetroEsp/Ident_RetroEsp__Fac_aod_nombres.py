#%% Librerías

import pyodbc
import pandas as pd
import time
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Color, fills, Alignment, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime
import numpy as np
import os
import sys

# --- Traduccion de claves a nombres --------------------------------------
# catalogo_nombres.py debe estar en la misma carpeta que este script.
_carpeta_script = os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd()
if _carpeta_script not in sys.path:
    sys.path.insert(0, _carpeta_script)
from catalogo_nombres import (cargar_catalogos, traducir_dataframe,
                              reporte_claves_sin_catalogo, COLUMNAS_A_CATALOGO)
from resumen_sudamerica import agregar_hoja_filtrada

warnings.filterwarnings('ignore')
start_time = time.perf_counter()
territorios = {0: 0, 1: 1, 2: 3, 3: 2, 4: 4, 5: 3, 6: 4, 7: 3, 9: 3, 10: 2, 11: 2, 12: 3, 13: 3, 14: 2, 15: 3, 16: 2, 17: 3, 19: 2, 23: 3, 24: 2, 25: 3, 26: 3, 27: 2, 28: 3, 29: 3, 30: 4, 31: 3, 32: 4, 33: 4, 34: 3, 35: 4, 36: 3, 37: 3, 38: 3, 41: 3, 42: 3, 43: 2, 44: 3, 45: 3, 46: 2, 47: 3, 49: 3, 50: 4, 51: 3, 52: 3, 53: 3, 54: 3, 55: 3, 59: 3, 60: 3, 61: 3, 63: 2, 64: 2, 65: 2, 66: 3, 67: 3, 68: 4, 69: 4, 72: 3, 73: 3, 74: 3, 75: 3, 76: 3, 77: 4, 78: 4, 79: 3, 80: 3, 81: 3, 82: 2, 83: 2, 84: 3, 85: 4, 86: 3, 87: 3, 89: 3, 90: 2, 91: 3, 92: 3, 93: 3, 95: 3, 96: 3, 97: 3, 98: 3, 99: 3, 100: 3, 101: 3, 102: 4, 103: 4, 104: 4, 105: 4, 106: 4, 107: 4, 108: 4, 109: 4, 110: 4, 111: 4, 112: 4, 113: 4, 114: 4, 115: 4, 116: 3, 117: 3, 118: 3, 119: 3, 120: 3, 121: 3, 122: 3, 123: 3, 124: 3, 125: 3, 126: 3, 127: 3, 128: 3, 129: 3, 130: 3, 131: 3, 132: 3, 133: 3, 134: 3, 135: 3, 136: 3, 137: 3, 138: 3, 139: 3, 140: 3, 141: 3, 142: 3, 143: 3, 144: 3, 145: 3, 146: 3, 147: 3, 148: 3, 149: 2, 150: 3, 151: 3, 152: 3, 153: 3, 154: 3, 155: 3, 156: 3, 157: 3, 158: 3, 159: 3, 160: 3, 161: 3, 162: 3, 163: 3, 164: 3, 165: 3, 166: 3, 167: 3}

#xFolder = r"C:\Users\aburtona\OneDrive - GPV\Documentos - Transf. Operativa RPAT\Reporting y Consultas\Consulta Identificación RetroEsp"
xFolder = r"C:\Users\asunad\OneDrive - GPV\Planeación Financiera RPAT - Reporting y Consultas\Consulta Identificación RetroEsp"

xCedentes = pd.read_excel(f"{xFolder}\\Catálogo consulta ident_retroesp.xlsx")

#%% Parámetros de conexión
server = 'adsPeru05'
database = 'pDBSirec2'
username = 'sqlDWH'
password = 'PatriaDWH.2023#'

# Crea la cadena de conexión
conn_str = f"DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}"

# Establece la conexión
conn = pyodbc.connect(conn_str)


#%% Cedido Facultativo

Cedido = f"SELECT T1.cCOR_Id as Corredor, T1.aHCF_NumContrato as NoContrato, T1.aHCF_IdentContrato as Ident_Contrato, T1.aHCF_Vig as Año_Vigencia, T1.aHCF_FecRenovacion as Inicio_Vigencia, T1.aHCF_FecFinVigencia as Fin_Vigencia, T1.aHCF_Nombre as Nombre, T1.aHCF_NombreIng as NombreIng, T18.cTRE_Id as DescComision, T18.rSobCom_PorcSobCom as ComPrimB, T18.rSobCom_PorcSobComPrimNeta as ComPrimNeta, " \
    f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cRAM_Id AS varchar(50)) FROM rRamRegXHCF T0 WHERE T1.cCOR_Id=T0.cCOR_Id AND T1.cCOR_Version=T0.cCOR_Version AND T1.aHCF_NumContrato=T0.aHCF_NumContrato AND T1.aHCF_IdentContrato=T0.aHCF_IdentContrato AND T1.aHCF_Vig=T0.aHCF_Vig AND T1.aHCF_Estatus=T0.aHCF_Estatus FOR XML PATH ('')), 1, 2, '') AS Ramos_Cubiertos, " \
    f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cSBP_Id AS varchar(50)) FROM rSubExcXHCF T0 WHERE T1.cCOR_Id=T0.cCOR_Id AND T1.cCOR_Version=T0.cCOR_Version AND T1.aHCF_NumContrato=T0.aHCF_NumContrato AND T1.aHCF_IdentContrato=T0.aHCF_IdentContrato AND T1.aHCF_Vig=T0.aHCF_Vig AND T1.aHCF_Estatus=T0.aHCF_Estatus FOR XML PATH ('')), 1, 2, '') AS Subramos_Excluidos, " \
    f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cTER_Id AS varchar(50)) FROM rTerrCubXHCF T0 WHERE T1.cCOR_Id=T0.cCOR_Id AND T1.cCOR_Version=T0.cCOR_Version AND T1.aHCF_NumContrato=T0.aHCF_NumContrato AND T1.aHCF_IdentContrato=T0.aHCF_IdentContrato AND T1.aHCF_Vig=T0.aHCF_Vig AND T1.aHCF_Estatus=T0.aHCF_Estatus FOR XML PATH ('')), 1, 2, '') AS Territorios_Cubiertos, " \
    f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cPAI_Id AS varchar(50)) FROM rPaiExcXHCF T0 WHERE T1.cCOR_Id=T0.cCOR_Id AND T1.cCOR_Version=T0.cCOR_Version AND T1.aHCF_NumContrato=T0.aHCF_NumContrato AND T1.aHCF_IdentContrato=T0.aHCF_IdentContrato AND T1.aHCF_Vig=T0.aHCF_Vig AND T1.aHCF_Estatus=T0.aHCF_Estatus FOR XML PATH ('')), 1, 2, '') AS Paises_Excluidos, " \

Tomado = f"COALESCE(" \
f"CASE WHEN T4.cCOR_Id IS NOT NULL THEN 'PROPORCIONAL' END, " \
f"CASE WHEN T5.cCOR_Id IS NOT NULL THEN 'NO PROPORCIONAL' END, " \
f"CASE WHEN T12.cCOR_Id IS NOT NULL THEN 'NO PROPORCIONAL' END, " \
f"CASE WHEN T6.cCOR_Id IS NOT NULL THEN 'FACULTATIVO DAÑOS' END, " \
f"CASE WHEN T7.cCOR_Id IS NOT NULL THEN 'FACULTATIVO FIANZAS' END, " \
f"CASE WHEN T8.cCOR_Id IS NOT NULL THEN 'FACULTATIVO VIDA IND' END, " \
f"CASE WHEN T9.cCOR_Id IS NOT NULL THEN 'FACULTATIVO VIDA GRUPO' END " \
f") as Tipo_Reaseguro, " \
f"COALESCE(T4.cCOR_Id, T5.cCOR_Id, T12.cCOR_Id, T6.cCOR_Id, T7.cCOR_Id, T8.cCOR_Id, T9.cCOR_Id) as Corredor_, " \
f"COALESCE(T4.cCIA_Id, T5.cCIA_Id, T12.cCIA_Id, T6.cCIA_Id, T7.cCIA_Id, T8.cCIA_Id, T9.cCIA_Id) as NoCedente, " \
f"COALESCE(T4.aHTP_NoContr, T5.aHTNP_NoContr, T12.aHTNP_NoContr, T6.aOFD_NoContr, T7.aOFF_Contrato, T8.aOFV_NumContr, T9.aOFVG_NumContr) as NoContrato_, " \
f"COALESCE(T6.cTER_Id, T7.cTER_Id, T8.cTER_Id, T8.cTER_Id, T9.cTER_Id, ' ') as cTER_Id, " \
f"COALESCE(T6.aOFD_NoOferta, T7.aOFF_NumOferta, T8.aOFV_NumOfer, T9.aOFVG_NumOfer, ' ') as NoOferta, " \
f"COALESCE(T6.aOFD_NoEndoso, T7.aOFF_NumEndoso, T8.aOFV_NumEnd, T9.aOFVG_NumEnd, ' ') as Endoso, " \
f"COALESCE(T4.aHTP_AnioVig, T5.aHTNP_AnioVig, T12.aHTNP_AnioVig, T6.aOFD_AnioVig, T7.aOFF_Año, T8.aOFV_AñoVig, T9.aOFVG_AñoVig) as Año_Vigencia_, " \
f"COALESCE(T4.aHTP_RefOrig, T5.aHTNP_RefOrig, T12.aHTNP_RefOrig, T14.cAMO_Nombre, T15.cADO_Nombre, T16.cAFI_NombreAsegurado, T17.cGADO_Nombre) as Referencia_Original, " \
f"COALESCE(T4.aHTP_FecIniPerCub, T5.aHTNP_FecIniPerCub, T12.aHTNP_FecIniPerCub, T6.aOFD_FecIniVig, T7.aOFF_FecVigIni, T8.aOFV_FecIniVig, T9.aOFVG_FecIniVig) as Inicio_Vigencia_, " \
f"COALESCE(T4.aHTP_FecFinPerCub, T5.aHTNP_FecFinPerCub, T12.aHTNP_FecFinPerCub, T6.aOFD_FecFinVig, T7.aOFF_FecVigFin, T8.aOFV_FecFinVig, T9.aOFVG_FecFinVig) as Fin_Vigencia_, " \
f"COALESCE(T5.cTCT_Id, T12.cTCT_Id, ' ') as cTCT_Id, " \
f"COALESCE(T5.rCAP_Consec, T12.rCAP_Consec, ' ') as Capa, " \
f"COALESCE(T5.aHTNP_Renglon, T12.aHTNP_Renglon, ' ') as Renglón, " \
f"COALESCE(T2.rCRHD_PrcRetro, T3.rCROF_PrcRetro, 0) AS PrcRetro, " \
f"COALESCE(" \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cRAM_Id AS varchar(50)) FROM rSBRXHTP T0 WHERE T4.cCOR_Id=T0.cCOR_Id AND T4.cCOR_Version=T0.cCOR_Version AND T4.cCIA_Id=T0.cCIA_Id AND T4.cCIA_Version=T0.cCIA_Version AND T4.aHTP_NoContr=T0.aHTP_NoContr AND T4.aHTP_AnioVig=T0.aHTP_AnioVig FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cRAM_Id AS varchar(50)) FROM rSBRXHTNP T0 WHERE T13.cCOR_Id=T0.cCOR_Id AND T13.cCOR_Version=T0.cCOR_Version AND T13.cCIA_Id=T0.cCIA_Id AND T13.cCIA_Version=T0.cCIA_Version AND T13.aHTNP_NoContr=T0.aHTNP_NoContr AND T13.aHTNP_AnioVig=T0.aHTNP_AnioVig FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cRAM_Id AS varchar(50)) FROM rSBRXOFD T0 WHERE T6.aOFD_NoContr=T0.aOFD_NoContr AND T6.cTER_Id=T0.cTER_Id AND T6.aOFD_AnioVig=T0.aOFD_AnioVig AND T6.aOFD_NoOferta=T0.aOFD_NoOferta AND T6.aOFD_NoEndoso=T0.aOFD_NoEndoso FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cRAM_Id AS varchar(50)) FROM rSubCubXOFF T0 WHERE T7.aOFF_Contrato=T0.aOFF_Contrato AND T7.cTER_Id=T0.cTER_Id AND T7.aOFF_Año=T0.aOFF_Año AND T7.aOFF_NumOferta=T0.aOFF_NumOferta AND T7.aOFF_NumEndoso=T0.aOFF_NumEndoso FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cRAM_Id AS varchar(50)) FROM rSubCubxOFV T0 WHERE T8.aOFV_NumContr=T0.aOFV_NumContr AND T8.cTER_Id=T0.cTER_Id AND T8.aOFV_AñoVig=T0.aOFV_AñoVig AND T8.aOFV_NumOfer=T0.aOFV_NumOfer AND T8.aOFV_NumEnd=T0.aOFV_NumEnd FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cRAM_Id AS varchar(50)) FROM rSubCubxOFVG T0 WHERE T9.aOFVG_NumContr=T0.aOFVG_NumContr AND T9.cTER_Id=T0.cTER_Id AND T9.aOFVG_AñoVig=T0.aOFVG_AñoVig AND T9.aOFVG_NumOfer=T0.aOFVG_NumOfer AND T9.aOFVG_NumEnd=T0.aOFVG_NumEnd For XML PATH ('')), 1, 2, '') " \
f") as Ramos_Cubiertos_, " \
f"COALESCE(" \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cSBP_Id AS varchar(50)) FROM rSBRXHTP T0 WHERE T4.cCOR_Id=T0.cCOR_Id AND T4.cCOR_Version=T0.cCOR_Version AND T4.cCIA_Id=T0.cCIA_Id AND T4.cCIA_Version=T0.cCIA_Version AND T4.aHTP_NoContr=T0.aHTP_NoContr AND T4.aHTP_AnioVig=T0.aHTP_AnioVig FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cSBP_Id AS varchar(50)) FROM rSBRXHTNP T0 WHERE T13.cCOR_Id=T0.cCOR_Id AND T13.cCOR_Version=T0.cCOR_Version AND T13.cCIA_Id=T0.cCIA_Id AND T13.cCIA_Version=T0.cCIA_Version AND T13.aHTNP_NoContr=T0.aHTNP_NoContr AND T13.aHTNP_AnioVig=T0.aHTNP_AnioVig FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cSBP_Id AS varchar(50)) FROM rSBRXOFD T0 WHERE T6.aOFD_NoContr=T0.aOFD_NoContr AND T6.cTER_Id=T0.cTER_Id AND T6.aOFD_AnioVig=T0.aOFD_AnioVig AND T6.aOFD_NoOferta=T0.aOFD_NoOferta AND T6.aOFD_NoEndoso=T0.aOFD_NoEndoso FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cSBR_Id AS varchar(50)) FROM rSubCubXOFF T0 WHERE T7.aOFF_Contrato=T0.aOFF_Contrato AND T7.cTER_Id=T0.cTER_Id AND T7.aOFF_Año=T0.aOFF_Año AND T7.aOFF_NumOferta=T0.aOFF_NumOferta AND T7.aOFF_NumEndoso=T0.aOFF_NumEndoso FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cSBR_Id AS varchar(50)) FROM rSubCubxOFV T0 WHERE T8.aOFV_NumContr=T0.aOFV_NumContr AND T8.cTER_Id=T0.cTER_Id AND T8.aOFV_AñoVig=T0.aOFV_AñoVig AND T8.aOFV_NumOfer=T0.aOFV_NumOfer AND T8.aOFV_NumEnd=T0.aOFV_NumEnd FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cSBR_Id AS varchar(50)) FROM rSubCubxOFVG T0 WHERE T9.aOFVG_NumContr=T0.aOFVG_NumContr AND T9.cTER_Id=T0.cTER_Id AND T9.aOFVG_AñoVig=T0.aOFVG_AñoVig AND T9.aOFVG_NumOfer=T0.aOFVG_NumOfer AND T9.aOFVG_NumEnd=T0.aOFVG_NumEnd FOR XML PATH ('')), 1, 2, '')" \
f") as Subramos_Cubiertos, " \
f"COALESCE(" \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cPAI_Id AS varchar(50)) FROM rPACXHTP T0 WHERE T4.cCOR_Id=T0.cCOR_Id AND T4.cCOR_Version=T0.cCOR_Version AND T4.cCIA_Id=T0.cCIA_Id AND T4.cCIA_Version=T0.cCIA_Version AND T4.aHTP_NoContr=T0.aHTP_NoContr AND T4.aHTP_AnioVig=T0.aHTP_AnioVig FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cPAI_Id AS varchar(50)) FROM rPACXHTNP T0 WHERE T13.cCOR_Id=T0.cCOR_Id AND T13.cCOR_Version=T0.cCOR_Version AND T13.cCIA_Id=T0.cCIA_Id AND T13.cCIA_Version=T0.cCIA_Version AND T13.aHTNP_NoContr=T0.aHTNP_NoContr AND T13.aHTNP_AnioVig=T0.aHTNP_AnioVig FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cPais_Id AS varchar(50)) FROM rPAISCUBXOFD T0 WHERE T6.aOFD_NoContr=T0.aOFD_NoContr AND T6.cTER_Id=T0.cTER_Id AND T6.aOFD_AnioVig=T0.aOFD_AnioVig AND T6.aOFD_NoOferta=T0.aOFD_NoOferta AND T6.aOFD_NoEndoso=T0.aOFD_NoEndoso FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cPAI_Id AS varchar(50)) FROM aOFD_OfertaD T0 WHERE T6.aOFD_NoContr=T0.aOFD_NoContr AND T6.cTER_Id=T0.cTER_Id AND T6.aOFD_AnioVig=T0.aOFD_AnioVig AND T6.aOFD_NoOferta=T0.aOFD_NoOferta AND T6.aOFD_NoEndoso=T0.aOFD_NoEndoso FOR XML PATH ('')), 1, 2, ''), " \
f"STUFF((SELECT DISTINCT ', ' + CAST(T0.cPAI_Id AS varchar(50)) FROM aOFF_OfertaFianzas T0 WHERE T7.aOFF_Contrato=T0.aOFF_Contrato AND T7.cTER_Id=T0.cTER_Id AND T7.aOFF_Año=T0.aOFF_Año AND T7.aOFF_NumOferta=T0.aOFF_NumOferta AND T7.aOFF_NumEndoso=T0.aOFF_NumEndoso FOR XML PATH ('')), 1, 2, '')" \
f") AS Paises_Cubiertos, " \
f"T6.cAMO_Id as Asegurado_Id_OFD, " \
f"T14.cAMO_Nombre as Asegurado_Nombre_OFD, " \
f"T7.cADO_Id as Afianzado_Id_OFF, " \
f"T15.cADO_Nombre as Afianzado_Nombre_OFF, " \
f"T8.cAFI_Id as Asegurado_Fisico_Id_OFV, " \
f"T16.cAFI_NombreAsegurado as Asegurado_Fisico_Nombre_OFV, " \
f"T9.cGADO_Id as Grupo_Asegurado_Id_OFVG, " \
f"T17.cGADO_Nombre as Grupo_Asegurado_Nombre_OFVG, " \
f"T4.aHTP_EsNegMGA as Negocio_MGA_TP, " \
f"T4.aHTP_MntLimRes as Limite_100_TP, " \
f"T4.aHTP_PrcLimResPt as Prc_Patria_TP, " \
f"T4.aHTP_MntLimRes * T4.aHTP_PrcLimResPt / 100 as Monto_Patria_TP, " \
f"COALESCE(T5.aHTNP_MntLimRes, T12.aHTNP_MntLimRes) as Limite_Resp_100_TNP, " \
f"COALESCE(T5.aHTNP_MntLimResPt, T12.aHTNP_MntLimResPt) as Limite_Resp_Patria_TNP, " \
f"COALESCE(T5.aHTNP_MntPriMinDep, T12.aHTNP_MntPriMinDep) as PMD_100_TNP, " \
f"COALESCE(T5.aHTNP_MntPriMinDepPt, T12.aHTNP_MntPriMinDepPt) as PMD_Patria_TNP, " \
f"COALESCE(T5.aHTNP_MntPriEst, T12.aHTNP_MntPriEst) as PriEst_100_TNP, " \
f"COALESCE(T5.aHTNP_MntPriEstPt, T12.aHTNP_MntPriEstPt) as PriEst_Patria_TNP, " \
f"T6.aOFD_MntSumAse as SA_Unica_100_OFD, " \
f"T6.aOFD_PrcAceptPt as Aceptacion_Patria_Prc_OFD, " \
f"T6.aOFD_MntAceptPt as Aceptacion_Patria_Mnt_OFD, " \
f"T6.aOFD_PrcCorretaje as Corretaje_Prc_OFD, " \
f"(SELECT TOP 1 R4.aOFD_IVASinCorretaje FROM pDBSirec2.dbo.aOFD_OfertaDCompl as R4 WHERE R4.aOFD_NoContr=T6.aOFD_NoContr AND R4.cTER_Id=T6.cTER_Id AND R4.aOFD_AnioVig=T6.aOFD_AnioVig AND R4.aOFD_NoOferta=T6.aOFD_NoOferta AND R4.aOFD_NoEndoso=T6.aOFD_NoEndoso) as IVA_sCorretaje_OFD, " \
f"(SELECT TOP 1 R4.aOFD_ISR FROM pDBSirec2.dbo.aOFD_OfertaDCompl as R4 WHERE R4.aOFD_NoContr=T6.aOFD_NoContr AND R4.cTER_Id=T6.cTER_Id AND R4.aOFD_AnioVig=T6.aOFD_AnioVig AND R4.aOFD_NoOferta=T6.aOFD_NoOferta AND R4.aOFD_NoEndoso=T6.aOFD_NoEndoso) as ISR_OFD, " \
f"T7.aOFF_SumaAfi as SA_100_OFF, " \
f"T7.aOFF_EPI100 as EPI_100_OFF, " \
f"T7.aOFF_SumaPt as SA_Patria_OFF, " \
f"T7.aOFF_MontoPt as Monto_Patria_OFF, " \
f"T7.aOFF_Comision as Comision_OFF, " \
f"T7.aOFF_Cuota as Cuota_OFF, " \
f"(SELECT SUM(R1.rPCO_MntComision) FROM pDBSirec2.dbo.rPCOXHTP as R1 WHERE R1.cCOR_Id=T4.cCOR_Id AND R1.cCOR_Version=T4.cCOR_Version AND R1.cCIA_Id=T4.cCIA_Id AND R1.cCIA_Version=T4.cCIA_Version AND R1.aHTP_NoContr=T4.aHTP_NoContr AND R1.aHTP_AnioVig=T4.aHTP_AnioVig) as Comision_Original_TP, " \
f"(SELECT SUM(R2.rHTP_ISR) FROM pDBSirec2.dbo.rHTP_ImpuestosPaisHP as R2 WHERE R2.cCor_Id=T4.cCOR_Id AND R2.cCia_Id=T4.cCIA_Id AND R2.rHTP_NoContr=T4.aHTP_NoContr AND R2.rHTP_AnioVig=T4.aHTP_AnioVig) as ISR_TP, " \
f"(SELECT SUM(R2.rHTP_Interes) FROM pDBSirec2.dbo.rHTP_ImpuestosPaisHP as R2 WHERE R2.cCor_Id=T4.cCOR_Id AND R2.cCia_Id=T4.cCIA_Id AND R2.rHTP_NoContr=T4.aHTP_NoContr AND R2.rHTP_AnioVig=T4.aHTP_AnioVig) as Interes_TP, " \
f"(SELECT SUM(R2.rHTP_Impuestos) FROM pDBSirec2.dbo.rHTP_ImpuestosPaisHP as R2 WHERE R2.cCor_Id=T4.cCOR_Id AND R2.cCia_Id=T4.cCIA_Id AND R2.rHTP_NoContr=T4.aHTP_NoContr AND R2.rHTP_AnioVig=T4.aHTP_AnioVig) as Impuestos_TP, " \
f"(SELECT SUM(R2.rHTP_ImpuestosSInteres) FROM pDBSirec2.dbo.rHTP_ImpuestosPaisHP as R2 WHERE R2.cCor_Id=T4.cCOR_Id AND R2.cCia_Id=T4.cCIA_Id AND R2.rHTP_NoContr=T4.aHTP_NoContr AND R2.rHTP_AnioVig=T4.aHTP_AnioVig) as Impuestos_sInteres_TP, " \
f"STUFF((SELECT ', ' + R3.rCEC_Nombre + ': ' + CAST(R3.rCEC_MntPrc AS varchar(50)) FROM pDBSirec2.dbo.rCECXOFD as R3 WHERE R3.aOFD_NoContr=T6.aOFD_NoContr AND R3.cTER_Id=T6.cTER_Id AND R3.aOFD_AnioVig=T6.aOFD_AnioVig AND R3.aOFD_NoOferta=T6.aOFD_NoOferta AND R3.aOFD_NoEndoso=T6.aOFD_NoEndoso FOR XML PATH('')), 1, 2, '') as SA_Componentes_OFD " \

#rSobComXHCF
xFROM = f" FROM pDBSirec2.dbo.aHCF_CedFacultativo as T1 " \
            f" LEFT JOIN rCRHD_CuadroRetroHD as T2 on T1.cCOR_Id=T2.cCOR_IdCed AND T1.cCOR_Version=T2.cCOR_VersionCed AND T1.aHCF_NumContrato=T2.aHC_NumContrato AND T1.aHCF_IdentContrato=T2.aHC_IdentContrato AND T1.aHCF_Vig=T2.aHC_Vig " \
            f" LEFT JOIN rCROF_CuadroRetroOF as T3 on T1.cCOR_Id=T3.cCOR_IdCed AND T1.cCOR_Version=T3.cCOR_VersionCed AND T1.aHCF_NumContrato=T3.aHC_NumContrato AND T1.aHCF_IdentContrato=T3.aHC_IdentContrato AND T1.aHCF_Vig=T3.aHC_Vig " \
            f" LEFT JOIN pDBSirec2.dbo.aHTP_HojDesTP as T4 on T2.cCOR_Id=T4.cCOR_Id AND T2.cCOR_Version=T4.cCOR_Version AND T2.cCIA_Id=T4.cCIA_Id AND T2.cCIA_Version=T4.cCIA_Version AND T2.aHT_NoContr=T4.aHTP_NoContr AND T2.aHT_AnioVig=T4.aHTP_AnioVig " \
            f" LEFT JOIN (SELECT T10.cCOR_Id, T10.cCOR_Version, T10.cCIA_Id, T10.cCIA_Version, T10.aHTNP_NoContr, T10.aHTNP_AnioVig, T10.aHTNP_FecIniPerCub, T10.aHTNP_FecFinPerCub, T10.cTCT_Id, T10.aHTNP_RefOrig, T10.aHTNP_MntLimRes, T10.aHTNP_MntLimResPt, T10.aHTNP_MntPriMinDep, T10.aHTNP_MntPriMinDepPt, T10.aHTNP_MntPriEst, T10.aHTNP_MntPriEstPt, T11.rCAP_Consec, T11.aHTNP_Renglon FROM pDBSirec2.dbo.aHTNP_HojDesTNP as T10 RIGHT JOIN rCAPXHTNP as T11 on T10.cCOR_Id=T11.cCOR_Id AND T10.cCOR_Version=T11.cCOR_Version AND T10.cCIA_Id=T11.cCIA_Id AND T10.cCIA_Version=T11.cCIA_Version AND T10.aHTNP_NoContr=T11.aHTNP_NoContr AND T10.aHTNP_AnioVig=T11.aHTNP_AnioVig) " \
            f" as T5 on T2.cCOR_Id=T5.cCOR_Id AND T2.cCOR_Version=T5.cCOR_Version AND T2.cCIA_Id=T5.cCIA_Id AND T2.cCIA_Version=T5.cCIA_Version AND T2.aHT_NoContr=T5.aHTNP_NoContr AND T2.aHT_AnioVig=T5.aHTNP_AnioVig AND T2.aHT_CapConsec=T5.rCAP_Consec " \
            f" LEFT JOIN (SELECT T10.cCOR_Id, T10.cCOR_Version, T10.cCIA_Id, T10.cCIA_Version, T10.aHTNP_NoContr, T10.aHTNP_AnioVig, T10.aHTNP_FecIniPerCub, T10.aHTNP_FecFinPerCub, T10.cTCT_Id, T10.aHTNP_RefOrig, T10.aHTNP_MntLimRes, T10.aHTNP_MntLimResPt, T10.aHTNP_MntPriMinDep, T10.aHTNP_MntPriMinDepPt, T10.aHTNP_MntPriEst, T10.aHTNP_MntPriEstPt, T11.rCAP_Consec, T11.aHTNP_Renglon FROM pDBSirec2.dbo.aHTNP_HojDesTNP as T10 RIGHT JOIN rCAPXHTNP as T11 on T10.cCOR_Id=T11.cCOR_Id AND T10.cCOR_Version=T11.cCOR_Version AND T10.cCIA_Id=T11.cCIA_Id AND T10.cCIA_Version=T11.cCIA_Version AND T10.aHTNP_NoContr=T11.aHTNP_NoContr AND T10.aHTNP_AnioVig=T11.aHTNP_AnioVig) " \
            f" as T12 on T2.cCOR_Id=T12.cCOR_Id AND T2.cCOR_Version=T12.cCOR_Version AND T2.cCIA_Id=T12.cCIA_Id AND T2.cCIA_Version=T12.cCIA_Version AND T2.aHT_NoContr=T12.aHTNP_NoContr AND T2.aHT_AnioVig=T12.aHTNP_AnioVig AND T2.rCRHD_Renglon=T12.aHTNP_Renglon " \
            f" LEFT JOIN pDBSirec2.dbo.aOFD_OfertaD as T6 on T3.aOF_NoContr=T6.aOFD_NoContr AND T3.cTER_Id=T6.cTER_Id AND T3.aOF_AnioVig=T6.aOFD_AnioVig AND T3.aOF_NoOferta=T6.aOFD_NoOferta AND T3.aOF_NoEndoso=T6.aOFD_NoEndoso "\
            f" LEFT JOIN pDBSirec2.dbo.aOFF_OfertaFianzas as T7 on  T3.aOF_NoContr=T7.aOFF_Contrato AND T3.cTER_Id=T7.cTER_Id AND T3.aOF_AnioVig=T7.aOFF_Año AND T3.aOF_NoOferta=T7.aOFF_NumOferta AND T3.aOF_NoEndoso=T7.aOFF_NumEndoso "\
            f" LEFT JOIN pDBSirec2.dbo.aOFV_OferVida as T8 on  T3.aOF_NoContr=T8.aOFV_NumContr AND T3.cTER_Id=T8.cTER_Id AND T3.aOF_AnioVig=T8.aOFV_AñoVig AND T3.aOF_NoOferta=T8.aOFV_NumOfer AND T3.aOF_NoEndoso=T8.aOFV_NumEnd "\
            f" LEFT JOIN pDBSirec2.dbo.aOFVG_OferVidaGpo as T9 on T3.aOF_NoContr=T9.aOFVG_NumContr AND T3.cTER_Id=T9.cTER_Id AND T3.aOF_AnioVig=T9.aOFVG_AñoVig AND T3.aOF_NoOferta=T9.aOFVG_NumOfer AND T3.aOF_NoEndoso=T9.aOFVG_NumEnd "\
            f" LEFT JOIN pDBSirec2.dbo.aHTNP_HojDesTNP as T13 on T2.cCOR_Id=T13.cCOR_Id AND T2.cCOR_Version=T13.cCOR_Version AND T2.cCIA_Id=T13.cCIA_Id AND T2.cCIA_Version=T13.cCIA_Version AND T2.aHT_NoContr=T13.aHTNP_NoContr AND T2.aHT_AnioVig=T13.aHTNP_AnioVig "\
            f" LEFT JOIN cAMO_AseguradoMoral as T14 on T6.cAMO_Id=T14.cAMO_Id " \
            f" LEFT JOIN cADO_Afianzado as T15 on T7.cADO_Id=T15.cADO_Id " \
            f" LEFT JOIN cAFI_AseguradoFisico as T16 on T8.cAFI_Id=T16.cAFI_Id " \
            f" LEFT JOIN cGADO_GpoAseg as T17 on T9.cGADO_Id=T17.cGADO_Id " \
            f" LEFT JOIN rSobComXHCF as T18 on T1.cCOR_Id=T18.cCOR_Id AND T1.cCOR_Version=T18.cCOR_Version AND T1.aHCF_NumContrato=T18.aHCF_NumContrato AND T1.aHCF_IdentContrato=T18.aHCF_IdentContrato AND T1.aHCF_Vig=T18.aHCF_Vig  "
xWHERE = f" WHERE T1.aHCF_Estatus = 1 AND T1.aHCF_Vig >= 2021 " 
xORDERBY = f" ORDER BY T1.aHCF_NumContrato "

xSQL = " ".join([Cedido, Tomado, xFROM, xWHERE, xORDERBY])

CedF = pd.read_sql(xSQL, conn)


# Cierra la conexión
conn.close()

#%%Consulta movimientos contables del tomado
conn = pyodbc.connect(conn_str)

xSELECT = ' SELECT cCOR_IdOrig, cCIA_IdOrig, aMOV_IdentContrOrig, aMOV_AñoOrig, ' \
            ' sum(case when cCTA_IdCta in (27) then aMOV_MntNal else 0 end) as DEUDORES_POR_PRIMAS_DE_ACCIDENTES_Y_ENFERMEDADES_Y_DAÑOS, ' \
            ' sum(case when cCTA_IdCta in (46) then aMOV_MntNal else 0 end) as INSTITUCIONES_DE_SEGUROS_CUENTA_CORRIENTE, ' \
            ' sum(case when cCTA_IdCta in (48) then aMOV_MntNal else 0 end) as INSTITUCIONES_DE_FIANZAS_CUENTA_CORRIENTE, ' \
            ' sum(case when cCTA_IdCta in (51) then aMOV_MntNal else 0 end) as PARTICIPACION_DE_REASEGURADORES_POR_SINIESTROS_PENDIENTES, ' \
            ' sum(case when cCTA_IdCta in (55) then aMOV_MntNal else 0 end) as PARTICIPACION_DE_REAFIANZADORAS_POR_RECLAMACIONES_PAGADAS, ' \
            ' sum(case when cCTA_IdCta in (60) then aMOV_MntNal else 0 end) as PARTICIPACION_DE_REASEGURADORAS_POR_COBERTURAS_DE_REASEGURO_Y_REAFIANZAMIENTO_NO_PROPORCIONAL, ' \
            ' sum(case when cCTA_IdCta in (67) then aMOV_MntNal else 0 end) as IMPUESTOS_PAGADOS_POR_ANTICIPADO, ' \
            ' sum(case when cCTA_IdCta in (89) then aMOV_MntNal else 0 end) as RESERVA_PARA_OBLIGACIONES_PENDIENTES_DE_CUMPLIR_POR_SINIESTROS_OCURRIDOS, ' \
            ' sum(case when cCTA_IdCta in (135) then aMOV_MntNal else 0 end) as COMISIONES_POR_PAGAR_DEL_REASEGURO_TOMADO, ' \
            ' sum(case when cCTA_IdCta in (221) then aMOV_MntNal else 0 end) as PRIMAS_RETROCEDIDAS, ' \
            ' sum(case when cCTA_IdCta in (224) then aMOV_MntNal else 0 end) as PRIMAS_RETROCEDIDAS_EN_REAFIANZAMIENTO, ' \
            ' sum(case when cCTA_IdCta in (240) then aMOV_MntNal else 0 end) as COSTO_DE_COBERTURAS_DE_REASEGURO_Y_REAFIANZAMIENTO_NO_PROPORCIONAL, ' \
            ' sum(case when cCTA_IdCta in (249) then aMOV_MntNal else 0 end) as COMISIONES_POR_REASEGURO_Y_REAFIANZAMIENTO_TOMADO, ' \
            ' sum(case when cCTA_IdCta in (251) then aMOV_MntNal else 0 end) as CORRETAJE_A_FAVOR_DE_INTERMEDIARIOS_DE_REASEGURO_Y_REAFIANZAMIENTO, ' \
            ' sum(case when cCTA_IdCta in (254) then aMOV_MntNal else 0 end) as SINIESTROS_DEL_REASEGURO_Y_RECLAMACIONES_DE_REAFIANZAMIENTO, ' \
            ' sum(case when cCTA_IdCta in (318) then aMOV_MntNal else 0 end) as PRIMAS_DEL_REASEGURO_TOMADO, ' \
            ' sum(case when cCTA_IdCta in (321) then aMOV_MntNal else 0 end) as PRIMAS_DEL_REAFIANZAMIENTO_TOMADO, ' \
            ' sum(case when cCTA_IdCta in (326) then aMOV_MntNal else 0 end) as COMISIONES_POR_REASEGURO_Y_REAFIANZAMIENTO_RETROCEDIDO ' \

xFROM = ' FROM pDBSirec2.dbo.aMOV_Mov '

xWHERE= ' WHERE aMOV_AñoOrig >= 2020'

xGROUPBY = ' GROUP BY cCOR_IdOrig, cCIA_IdOrig, aMOV_IdentContrOrig, aMOV_AñoOrig '

xSQL = " ".join([xSELECT, xFROM, xWHERE, xGROUPBY])

MovTom = pd.read_sql(xSQL, conn)

#%% Moneda y póliza: consulta ligera (sin agregación pesada) y concatenación en pandas
# Se traen solo las columnas necesarias y se concatenan por contrato en Python,
# evitando las subconsultas correlacionadas que saturaban a SQL Server.
xSQL_MonPol = ' SELECT DISTINCT cCOR_IdOrig, cCIA_IdOrig, aMOV_IdentContrOrig, aMOV_AñoOrig, ' \
              ' cMON_IdOrig, aPOG_Num ' \
              ' FROM pDBSirec2.dbo.aMOV_Mov ' \
              ' WHERE aMOV_AñoOrig >= 2020 '

MonPol = pd.read_sql(xSQL_MonPol, conn)

conn.close()

# Claves de agrupación (las mismas 4 del movimiento)
_llaves_mp = ['cCOR_IdOrig', 'cCIA_IdOrig', 'aMOV_IdentContrOrig', 'aMOV_AñoOrig']

# Concatenar monedas y pólizas distintas por contrato (ordenadas, separadas por coma)
def _concat_unicos(serie):
    vals = []
    for v in serie.dropna().unique():
        try:
            vals.append(str(int(v)))   # entero sin decimales (.0)
        except (ValueError, TypeError):
            vals.append(str(v))
    return ', '.join(sorted(vals, key=lambda z: (len(z), z)))

MonPol_agg = MonPol.groupby(_llaves_mp).agg(
    Monedas_Movimiento=('cMON_IdOrig', _concat_unicos),
    Polizas_Asociadas=('aPOG_Num', _concat_unicos)
).reset_index()

# Unir las dos columnas concatenadas al resultado de movimientos
MovTom = MovTom.merge(MonPol_agg, how='left', on=_llaves_mp)

MovTom['LLAVE'] = MovTom[['cCOR_IdOrig', 'cCIA_IdOrig', 'aMOV_IdentContrOrig', 'aMOV_AñoOrig']].apply(lambda x: '-'.join(x.astype(int).astype(str)), axis=1)
CedF['LLAVE'] = CedF[['Corredor_', 'NoCedente', 'NoContrato_', 'Año_Vigencia_']].apply(lambda x: '-'.join(x.fillna(0).astype(int).astype(str)), axis=1)

CedF = CedF.merge(MovTom.drop_duplicates(), how="left", on="LLAVE")


#xFolder = r"C:\Users\aburtona\OneDrive - GPV\Documentos"
xFolder = r"C:\Users\asunad\OneDrive - GPV\Documents"
fileName = f"{xFolder}\\CedF.xlsx"
CedF.to_excel(fileName, index=False)

def to_set(value):
    """
    Convierte algo como '1, 2, 3' a set({'1','2','3'})
    Maneja: NaN, None, floats, strings sin espacios, strings con espacios, listas.
    """
    if pd.isna(value):
        return set()
    if isinstance(value, list):
        return {str(x).strip() for x in value if str(x).strip()}
    if isinstance(value, float):
        return set()
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return set()

    return {item.strip() for item in s.split(",") if item.strip()}

def comparar_ramos(row):
    ramos_cedidos = to_set(row.get('Ramos_Cubiertos'))
    ramos_tomados = to_set(row.get('Ramos_Cubiertos_'))
    return ramos_tomados.issubset(ramos_cedidos)

def comparar_subramos(row):
    subramos_cedidos = to_set(row.get('Subramos_Excluidos'))
    subramos_tomados = to_set(row.get('Subramos_Cubiertos'))
    
    if not subramos_cedidos:
        return True

    return subramos_tomados.isdisjoint(subramos_cedidos)

def comparar_paises(row):
    paises_cedidos = to_set(row.get('Paises_Excluidos'))
    paises_tomados = to_set(row.get('Paises_Cubiertos'))

    if not paises_cedidos:
        return True

    return paises_tomados.isdisjoint(paises_cedidos)

def comparar_territorios(row):
    terr_ced = to_set(row.get('Territorios_Cubiertos'))
    terr_tom = to_set(row.get('Territorio'))
    return terr_tom.issubset(terr_ced)

def comparar_vigencias(row):
    fecha_inicio_cedido = row['Inicio_Vigencia']
    fecha_inicio_tomado = row['Inicio_Vigencia_']

    return fecha_inicio_tomado >= fecha_inicio_cedido

def Columna_territorio(paises):
    paises_set = to_set(paises)
    paises_list = [int(x) for x in paises_set if x.isdigit()]
    terr = {territorios[p] for p in paises_list if p in territorios}
    return ", ".join(map(str, sorted(terr)))

#%%Funciones de validación y territorios

    ramos_cedidos = row['Ramos_Cubiertos'] if row['Ramos_Cubiertos'] is not None else ''
    ramos_tomados = row['Ramos_Cubiertos_'] if row['Ramos_Cubiertos_'] is not None else ''
   
    ramos_cedidos = sorted(set(ramos_cedidos.split(', '))) if ramos_cedidos else []
    ramos_tomados = sorted(set(ramos_tomados.split(', '))) if ramos_tomados else []

    return set(ramos_tomados) <= set(ramos_cedidos)



    Subramos_cedidos = row['Subramos_Excluidos'] if row['Subramos_Excluidos'] is not None else ''
    Subramos_tomados = row['Subramos_Cubiertos'] if row['Subramos_Cubiertos'] is not None else ''
    
    Subramos_cedidos = set(Subramos_cedidos.split(', ')) if Subramos_cedidos else set()
    Subramos_tomados = set(Subramos_tomados.split(', ')) if Subramos_tomados else set()

    if not Subramos_cedidos:
        return True

    return Subramos_tomados.isdisjoint(Subramos_cedidos)


    Paises_cedidos = row['Paises_Excluidos'] if row['Paises_Excluidos'] is not None else ''
    Paises_tomados = row['Paises_Cubiertos'] if row['Paises_Cubiertos'] is not None else ''
    
    Paises_cedidos = set(Paises_cedidos.split(', ')) if Paises_cedidos else set()
    Paises_tomados = set(Paises_tomados.split(', ')) if Paises_tomados else set()

    if not Paises_cedidos:
        return True
    
    return Paises_tomados.isdisjoint(Paises_cedidos)


    fecha_inicio_cedido = row['Inicio_Vigencia'] if row['Inicio_Vigencia'] is not None else None
    fecha_inicio_tomado = row['Inicio_Vigencia_'] if row['Inicio_Vigencia_'] is not None else None

    if isinstance(fecha_inicio_cedido, str) and isinstance(fecha_inicio_tomado, str):
        fecha_inicio_cedido = datetime.strptime(fecha_inicio_cedido, '%Y-%m-%d')
        fecha_inicio_tomado = datetime.strptime(fecha_inicio_tomado, '%Y-%m-%d')

        
    return fecha_inicio_tomado >= fecha_inicio_cedido


    if not paises:  
        return ''
    paises_list = [int(pais.strip()) for pais in paises.split(',') if pais.strip().isdigit()]
    territorios_set = {territorios[pais] for pais in paises_list if pais in territorios}

    return ', '.join(map(str, sorted(territorios_set)))



#%% Aplica las funciones de validaciones y territorio a las columnas correspondientes
CedF['Validacion Ramos'] = CedF.apply(comparar_ramos, axis=1)
CedF['Validacion Subramos'] = CedF.apply(comparar_subramos, axis=1)
CedF['Validacion Paises'] = CedF.apply(comparar_paises, axis=1)
CedF['Validacion Vigencia'] = CedF.apply(comparar_vigencias, axis=1)
CedF['Territorio'] = CedF['Paises_Cubiertos'].apply(Columna_territorio)
CedF['Validacion Territorios'] = CedF.apply(comparar_territorios, axis=1)
CedF = CedF.merge(xCedentes[["No. Cedente","Nombre Cedente"]].drop_duplicates(),
                             how="left", left_on="NoCedente", right_on="No. Cedente")
CedF['Cedente'] = CedF['Nombre Cedente'] 

#%%Reacomodo de las columnas de la consulta y guardado de la tabla sin formato
Orden = ['Corredor', 'NoContrato', 'Ident_Contrato', 'Año_Vigencia', 'Inicio_Vigencia', 'Fin_Vigencia', 'Nombre', 'NombreIng', 'DescComision', 'ComPrimB', 'ComPrimNeta', 'Ramos_Cubiertos', 'Subramos_Excluidos', 'Territorios_Cubiertos', 'Paises_Excluidos', 'Tipo_Reaseguro', 'Corredor_', 'NoCedente', 'Cedente', 'NoContrato_', 'cTER_Id', 'NoOferta', 'Endoso', 'Año_Vigencia_', 'Referencia_Original', 'Inicio_Vigencia_', 'Fin_Vigencia_', 'cTCT_Id', 'Capa', 'Renglón', 'PrcRetro','Ramos_Cubiertos_', 'Subramos_Cubiertos', 'Territorio','Paises_Cubiertos', 'Asegurado_Id_OFD', 'Asegurado_Nombre_OFD', 'Afianzado_Id_OFF', 'Afianzado_Nombre_OFF', 'Asegurado_Fisico_Id_OFV', 'Asegurado_Fisico_Nombre_OFV', 'Grupo_Asegurado_Id_OFVG', 'Grupo_Asegurado_Nombre_OFVG', 'Negocio_MGA_TP', 'Limite_100_TP', 'Prc_Patria_TP', 'Monto_Patria_TP', 'Limite_Resp_100_TNP', 'Limite_Resp_Patria_TNP', 'PMD_100_TNP', 'PMD_Patria_TNP', 'PriEst_100_TNP', 'PriEst_Patria_TNP', 'SA_Unica_100_OFD', 'Aceptacion_Patria_Prc_OFD', 'Aceptacion_Patria_Mnt_OFD', 'Corretaje_Prc_OFD', 'IVA_sCorretaje_OFD', 'ISR_OFD', 'SA_100_OFF', 'EPI_100_OFF', 'SA_Patria_OFF', 'Monto_Patria_OFF', 'Comision_OFF', 'Cuota_OFF', 'Comision_Original_TP', 'ISR_TP', 'Interes_TP', 'Impuestos_TP', 'Impuestos_sInteres_TP', 'SA_Componentes_OFD', 'Validacion Vigencia', 'Validacion Ramos', 'Validacion Subramos', 'Validacion Territorios', 'Validacion Paises',
         'Monedas_Movimiento', 'Polizas_Asociadas', 'DEUDORES_POR_PRIMAS_DE_ACCIDENTES_Y_ENFERMEDADES_Y_DAÑOS', 'INSTITUCIONES_DE_SEGUROS_CUENTA_CORRIENTE', 'INSTITUCIONES_DE_FIANZAS_CUENTA_CORRIENTE', 'PARTICIPACION_DE_REASEGURADORES_POR_SINIESTROS_PENDIENTES', 'PARTICIPACION_DE_REAFIANZADORAS_POR_RECLAMACIONES_PAGADAS', 'PARTICIPACION_DE_REASEGURADORAS_POR_COBERTURAS_DE_REASEGURO_Y_REAFIANZAMIENTO_NO_PROPORCIONAL', 'IMPUESTOS_PAGADOS_POR_ANTICIPADO', 'RESERVA_PARA_OBLIGACIONES_PENDIENTES_DE_CUMPLIR_POR_SINIESTROS_OCURRIDOS', 'COMISIONES_POR_PAGAR_DEL_REASEGURO_TOMADO', 'PRIMAS_RETROCEDIDAS', 'PRIMAS_RETROCEDIDAS_EN_REAFIANZAMIENTO', 'COSTO_DE_COBERTURAS_DE_REASEGURO_Y_REAFIANZAMIENTO_NO_PROPORCIONAL', 'COMISIONES_POR_REASEGURO_Y_REAFIANZAMIENTO_TOMADO', 'CORRETAJE_A_FAVOR_DE_INTERMEDIARIOS_DE_REASEGURO_Y_REAFIANZAMIENTO', 'SINIESTROS_DEL_REASEGURO_Y_RECLAMACIONES_DE_REAFIANZAMIENTO', 'PRIMAS_DEL_REASEGURO_TOMADO', 'PRIMAS_DEL_REAFIANZAMIENTO_TOMADO', 'COMISIONES_POR_REASEGURO_Y_REAFIANZAMIENTO_RETROCEDIDO']
CedF_Ordenado = CedF[Orden]



# Ruta de la carpeta donde se guardan los archivos de Excel
xFolder = r"C:\Users\asunad\OneDrive - GPV\Planeación Financiera RPAT - Reporting y Consultas\Consulta Identificación RetroEsp"

#%% Traduccion de claves a nombres con la hoja de Catalogo
# Se lee el mismo catalogo que ya se pega como pestana en el archivo final y se
# sustituyen corredor, pais, ramo, subramo y territorio por su descripcion.
catalogos = cargar_catalogos(f"{xFolder}\\Catálogo consulta ident_retroesp.xlsx")
CedF_Nombres = traducir_dataframe(CedF_Ordenado, catalogos)


#%% Generacion de los archivos de Excel (mismo formato para ambas versiones)
def generar_libro(tabla, fileName_Retro, version_nombres=False, hoja_sudamerica=False):
    """Escribe la tabla en Excel con el formato de siempre.

    version_nombres=True solo amplia las columnas que quedaron con texto largo
    despues de traducir las claves.
    hoja_sudamerica=True agrega una copia de la hoja con el filtro de Excel
    puesto en los paises de America del Sur.
    """
    tabla.to_excel(fileName_Retro, index=False)

    #%% Formato de la tabla en excel
    #Cargar el archivo recien creado con la consulta
    Ident_RetroEsp = load_workbook(fileName_Retro, keep_vba= True, keep_links=False)
    ws = Ident_RetroEsp.active

    #Formato general
    azul_oscuro = Color(indexed = 32) #Color para encabezados del cedido
    azul = Color(indexed = 48) #Color para encabezados del tomado
    morado = Color(indexed = 38)
    Estilo = Font(name='Arial', size = 9, bold = True, color = colors.WHITE) #Estilo y color letra de los encabezados
    alignment = Alignment(horizontal= "left") 

    #Formato encabezados del cedido
    for celdas in ['A1', 'B1', 'C1', 'D1', 'E1','F1', 'G1', 'H1', 'I1', 'J1', 'K1','L1','M1','N1','O1']:
        celda = ws[celdas]
        celda.font = Estilo
        celda.fill = PatternFill(fill_type="solid", start_color= azul_oscuro, end_color= azul_oscuro)
        celda.alignment = alignment

    #Formato encabezados del tomado (expandido a la columna BZ)
    Relleno = PatternFill(fill_type="solid", start_color= azul, end_color= azul)
    for celdas in ['P1', 'Q1','R1', 'S1', 'T1', 'U1', 'V1', 'W1','X1','Y1', 'Z1', 'AA1', 'AB1', 'AC1', 'AD1','AE1', 'AF1', 'AG1', 'AH1', 'AI1', 'AJ1','AK1','AL1','AM1','AN1', 'AO1', 'AP1', 'AQ1', 'AR1', 'AS1', 'AT1', 'AU1', 'AV1', 'AW1', 'AX1', 'AY1', 'AZ1', 'BA1', 'BB1', 'BC1', 'BD1', 'BE1', 'BF1', 'BG1', 'BH1', 'BI1', 'BJ1', 'BK1', 'BL1', 'BM1', 'BN1', 'BO1', 'BP1', 'BQ1', 'BR1', 'BS1', 'BT1', 'BU1', 'BV1', 'BW1', 'BX1', 'BY1', 'BZ1']:
        celda = ws[celdas]
        celda.font = Estilo
        celda.fill = Relleno
        celda.alignment = alignment

    #Formato encabezados de los movimientos contables (recorridos a CA:CR)
    Relleno = PatternFill(fill_type="solid", start_color= morado, end_color= morado)
    for celdas in ['CA1', 'CB1', 'CC1', 'CD1', 'CE1', 'CF1', 'CG1', 'CH1', 'CI1', 'CJ1', 'CK1', 'CL1', 'CM1', 'CN1', 'CO1', 'CP1', 'CQ1', 'CR1']:
        celda = ws[celdas]
        celda.font = Estilo
        celda.fill = Relleno
        celda.alignment = alignment

    #Formato de fechas cortas para los inicios y fines de vigencia del cedido y tomado
    num_filas = len(tabla) + 2

    for row in range (1,num_filas):
        ws[f'E{row}'].number_format = 'DD/MM/YYYY'
        ws[f'F{row}'].number_format = 'DD/MM/YYYY'
        ws[f'Z{row}'].number_format = 'DD/MM/YYYY'
        ws[f'AA{row}'].number_format = 'DD/MM/YYYY'

    #Formato visual de porcentaje para % Patria (Prop), % Aceptacion Patria (Daños) y % Corretaje (Daños).
    #El valor en la celda permanece como numero 0-100 (no se altera el dato ni el calculo de Monto Patria); solo se agrega el simbolo % a la vista.
    for row in range (3,num_filas):
        ws[f'AT{row}'].number_format = '0.00"%"'
        ws[f'BC{row}'].number_format = '0.00"%"'
        ws[f'BE{row}'].number_format = '0.00"%"'
        ws[f'BF{row}'].number_format = '0.00"%"'
        ws[f'BG{row}'].number_format = '0.00"%"'

    #Modifica ancho de columnas
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['P'].width = 15
    ws.column_dimensions['S'].width = 15
    ws.column_dimensions['Y'].width = 15
    ws.column_dimensions['Z'].width = 11
    ws.column_dimensions['AA'].width = 11
    ws.column_dimensions['BL'].width = 12
    ws.column_dimensions['BM'].width = 12
    ws.column_dimensions['BT'].width = 12
    ws.column_dimensions['BU'].width = 12
    ws.column_dimensions['BV'].width = 12
    ws.column_dimensions['BW'].width = 12
    ws.column_dimensions['BX'].width = 12

    #Modificar encabezados
    Nuevos_encabezados = [
        'Corredor', 'No. Contrato', 'Ident Contrato', 'Año Vigencia', 'Inicio Vigencia', 'Fin Vigencia', 'Nombre', 'Nombre Ing', 
        'DescComision', 'ComPrimBruta', 'ComPrimNeta', 'Ramos Cubiertos', 'Subramos Excluidos', 'Territorios Cubiertos', 'Paises Excluidos', 
        'Tipo Reaseguro', 'Corredor', 'No. Cedente', 'Cedente', 'No. Contrato', 'Territorio llave', 'No. Oferta', 'Endoso',
        'Año Vigencia', 'Referencia Original', 'Inicio Vigencia', 'Fin Vigencia', 'Tipo Contrato', 'Capa', 'Renglón', 'PrcRetro',
        'Ramos Cubiertos', 'Subramos Cubiertos', 'Territorios Cubiertos', 'Paises Cubiertos', 
        'Asegurado Id (Daños)', 'Asegurado Nombre (Daños)', 'Afianzado Id (Fianzas)', 'Afianzado Nombre (Fianzas)', 'Asegurado Físico Id (Vida Ind)', 'Asegurado Físico Nombre (Vida Ind)', 'Grupo Asegurado Id (Vida Gpo)', 'Grupo Asegurado Nombre (Vida Gpo)',
        'Negocio MGA (Prop)', 'Límite 100% (Prop)', '% Patria (Prop)', 'Monto Patria (Prop)', 'Límite Resp. 100% (No Prop)', 'Límite Resp. Patria (No Prop)', 'PMD 100% (No Prop)', 'PMD Patria (No Prop)', 'Prima Esperada 100% (No Prop)', 'Prima Esperada Patria (No Prop)', 'Suma Asegurada Única 100% (Daños)', '% Aceptación Patria (Daños)', 'Monto Aceptación Patria (Daños)', '% Corretaje (Daños)', 'IVA s/Corretaje (Daños)', 'ISR (Daños)', 'Suma Afianzada 100 (Fianzas)', 'EPI 100 (Fianzas)', 'Suma Asegurada Patria (Fianzas)', 'Monto Patria (Fianzas)', 'Comisión (Fianzas)', 'Cuota (Fianzas)', '% Comisión Original (Prop)', 'ISR (Prop)', 'Interés (Prop)', 'Impuestos (Prop)', 'Impuestos s/Interés (Prop)', 'Suma Asegurada por Componente (Daños)',
        'Validacion Vigencia', 'Validacion Ramos', 'Validacion Subramos', 'Validacion Territorios', 'Validacion Paises',
        'Monedas del Movimiento', 'Pólizas Asociadas', 'DEUDORES_POR_PRIMAS_DE_ACCIDENTES_Y_ENFERMEDADES_Y_DAÑOS', 'INSTITUCIONES_DE_SEGUROS_CUENTA_CORRIENTE', 'INSTITUCIONES_DE_FIANZAS_CUENTA_CORRIENTE',
        'PARTICIPACION_DE_REASEGURADORES_POR_SINIESTROS_PENDIENTES', 'PARTICIPACION_DE_REAFIANZADORAS_POR_RECLAMACIONES_PAGADAS', 'PARTICIPACION_DE_REASEGURADORAS_POR_COBERTURAS_DE_REASEGURO_Y_REAFIANZAMIENTO_NO_PROPORCIONAL',
        'IMPUESTOS_PAGADOS_POR_ANTICIPADO', 'RESERVA_PARA_OBLIGACIONES_PENDIENTES_DE_CUMPLIR_POR_SINIESTROS_OCURRIDOS', 
        'COMISIONES_POR_PAGAR_DEL_REASEGURO_TOMADO', 'PRIMAS_RETROCEDIDAS', 'PRIMAS_RETROCEDIDAS_EN_REAFIANZAMIENTO', 'COSTO_DE_COBERTURAS_DE_REASEGURO_Y_REAFIANZAMIENTO_NO_PROPORCIONAL',
        'COMISIONES_POR_REASEGURO_Y_REAFIANZAMIENTO_TOMADO', 'CORRETAJE_A_FAVOR_DE_INTERMEDIARIOS_DE_REASEGURO_Y_REAFIANZAMIENTO', 'SINIESTROS_DEL_REASEGURO_Y_RECLAMACIONES_DE_REAFIANZAMIENTO', 
        'PRIMAS_DEL_REASEGURO_TOMADO', 'PRIMAS_DEL_REAFIANZAMIENTO_TOMADO', 'COMISIONES_POR_REASEGURO_Y_REAFIANZAMIENTO_RETROCEDIDO'
    ]
    for col, valor in enumerate(Nuevos_encabezados, start=1):
        ws.cell(row=1, column=col).value = valor



    #Agrega una fila arriba de los encabezados para indicar los encabezados del cedido y del tomado con su formato de color correspondiente
    ws.insert_rows(idx=1, amount=1)
    celda_cedido = ws['A1']
    celda_cedido.value = "CEDIDO"
    celda_cedido.font = Estilo
    celda_cedido.fill = PatternFill(fill_type="solid", start_color= azul_oscuro, end_color= azul_oscuro)
    celda_cedido.alignment = Alignment(horizontal= "center") 

    celda_tomado = ws['P1']
    celda_tomado.value = "TOMADO"
    celda_tomado.font = Estilo
    celda_tomado.fill = Relleno = PatternFill(fill_type="solid", start_color= azul, end_color= azul)
    celda_tomado.alignment = Alignment(horizontal= "center") 

    celda_contables = ws['CA1']
    celda_contables.value = "MOVIMIENTOS CONTABLES"
    celda_contables.font = Estilo
    celda_contables.fill = Relleno = PatternFill(fill_type="solid", start_color= morado, end_color= morado)
    celda_contables.alignment = Alignment(horizontal= "center") 

    ws.merge_cells('A1:O1')
    ws.merge_cells('P1:BZ1')
    ws.merge_cells('CA1:CR1')


    ws.sheet_view.showGridLines = False #Quita las líneas de cuadricula
    ws.freeze_panes = 'A3' #Inmovilizar paneles
    ws.auto_filter.ref = 'A2:CR2' #Aplicar filtro a los encabezados
    #%%Obtener el catálogo del archivo de la misma carpeta

    #Crea una nueva hoja para el catálogo
    nueva_hoja = Ident_RetroEsp.create_sheet(title="Catálogo")

    #Ruta archivo catálogo
    xFolder2 = r"C:\Users\asunad\OneDrive - GPV\Planeación Financiera RPAT - Reporting y Consultas\Consulta Identificación RetroEsp"
    fileName_Catalogo = f"{xFolder2}\\Catálogo consulta ident_retroesp.xlsx"
    Catalogo = load_workbook(fileName_Catalogo)

    # Seleccionar la hoja
    hoja_catalogo = Catalogo['Catálogo']


    #Copiar y pegar los valores del catalogo 
    for fila in hoja_catalogo.iter_rows():
        for celda in fila:
            # Copiar el valor de la celda
            nueva_hoja[celda.coordinate].value = celda.value

    #Copiar formatos de ancho de columna
    for col in hoja_catalogo.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        nueva_hoja.column_dimensions[column].width = adjusted_width   

    #Formatos encabezado catálogo
    for celdas in ['A1', 'B1', 'D1', 'E1', 'G1', 'H1', 'J1', 'K1','M1','N1','O1','Q1','R1']:
        celda = nueva_hoja[celdas]
        celda.font = Estilo
        celda.fill = PatternFill(fill_type="solid", start_color= azul_oscuro, end_color= azul_oscuro)
        celda.alignment = alignment

    #Quita las líneas de cuadricula de la hoja del catálogo
    nueva_hoja.sheet_view.showGridLines = False


    #Ancho extra para las columnas que traen nombres en lugar de claves
    if version_nombres:
        for nombre_columna in COLUMNAS_A_CATALOGO:
            if nombre_columna in tabla.columns:
                letra = get_column_letter(tabla.columns.get_loc(nombre_columna) + 1)
                if (ws.column_dimensions[letra].width or 0) < 30:
                    ws.column_dimensions[letra].width = 30

    #Copia de la hoja con el filtro de America del Sur ya aplicado
    if hoja_sudamerica:
        _, visibles = agregar_hoja_filtrada(Ident_RetroEsp)
        print(f"Hoja 'América del Sur': {visibles:,} renglones visibles con el filtro")

    #%%Guarda el archivo nuevo
    Ident_RetroEsp.save(fileName_Retro)
    print(f"Archivo generado: {fileName_Retro}")
    return fileName_Retro


#%% Se generan las dos versiones: la de claves (la de siempre) y la de nombres
generar_libro(CedF_Ordenado, f"{xFolder}\\Ident_RetroEsp_Facv2.xlsm")
generar_libro(CedF_Nombres, f"{xFolder}\\Ident_RetroEsp_Facv2_nombres.xlsm",
              version_nombres=True, hoja_sudamerica=True)

#Claves que no estan dadas de alta en el catalogo (si las hay, conviene agregarlas)
print(reporte_claves_sin_catalogo())
end_time = time.perf_counter()
elapsed_time = end_time - start_time
print("Elapsed time: ", elapsed_time)

