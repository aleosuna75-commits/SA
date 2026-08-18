# -*- coding: utf-8 -*-
"""
Convierte un archivo Ident_RetroEsp ya generado (con claves) a su version
"_nombres", sin volver a correr la consulta a SIREC.

Toma el archivo tal cual esta, lee su propia hoja "Catalogo" (o el archivo de
catalogo de la carpeta, si se prefiere) y sustituye en la hoja de datos las
claves de corredor, pais, ramo, subramo y territorio por su descripcion.

Uso rapido (Spyder / linea de comandos):
    python Convertir_Ident_RetroEsp_a_nombres.py

Uso con rutas explicitas:
    python Convertir_Ident_RetroEsp_a_nombres.py "...\\Ident_RetroEsp_Facv2.xlsm"

El archivo original NO se modifica: se escribe uno nuevo con el sufijo _nombres.
"""

import os
import sys
import time

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    from catalogo_nombres import (cargar_catalogos_desde_filas, traducir,
                                  traducir_si_no, reporte_claves_sin_catalogo,
                                  ENCABEZADOS_A_CATALOGO, _normaliza_texto)
    from resumen_sudamerica import agregar_hoja_filtrada
except ImportError as detalle:
    sys.exit(f"Falta un modulo en la misma carpeta que este script ({detalle}).")

# ---------------------------------------------------------------------------
# Parametros
# ---------------------------------------------------------------------------
xFolder = r"C:\Users\asunad\OneDrive - GPV\Planeación Financiera RPAT - Reporting y Consultas\Consulta Identificación RetroEsp"

ARCHIVOS = [
    f"{xFolder}\\Ident_RetroEsp_Propv2.xlsm",
    f"{xFolder}\\Ident_RetroEsp_Facv2.xlsm",
]

SUFIJO = "_nombres"
HOJA_DATOS = "Sheet1"        # hoja con la consulta
HOJA_CATALOGO = "Catálogo"   # hoja con las claves y sus descripciones
FILA_ENCABEZADOS = 2         # la fila 1 trae las bandas CEDIDO / TOMADO / MOVIMIENTOS
ENCABEZADOS_SI_NO = ('Negocio MGA (Prop)',)
ANCHO_MINIMO_TRADUCIDAS = 28
AGREGAR_RESUMEN_SUDAMERICA = True   # copia de la hoja con el filtro de America del Sur
# True copia solo los renglones de la region (archivo mucho mas ligero); False
# copia la hoja completa y oculta los demas renglones, dejando el filtro reversible.
SOLO_RENGLONES_FILTRADOS = False


def ruta_salida(ruta, sufijo=SUFIJO):
    base, extension = os.path.splitext(ruta)
    return f"{base}{sufijo}{extension}"


def convertir(ruta_entrada, ruta_salida_final=None, hoja_datos=HOJA_DATOS,
              hoja_catalogo=HOJA_CATALOGO, fila_encabezados=FILA_ENCABEZADOS,
              agregar_resumen=AGREGAR_RESUMEN_SUDAMERICA,
              solo_renglones_filtrados=SOLO_RENGLONES_FILTRADOS):
    """Genera la version con nombres del archivo indicado y regresa su ruta."""
    ruta_salida_final = ruta_salida_final or ruta_salida(ruta_entrada)

    libro = load_workbook(ruta_entrada, keep_vba=ruta_entrada.lower().endswith('.xlsm'))

    if hoja_catalogo not in libro.sheetnames:
        raise ValueError(f"El archivo no trae la hoja '{hoja_catalogo}': {ruta_entrada}")
    filas_catalogo = [[celda.value for celda in fila]
                      for fila in libro[hoja_catalogo].iter_rows()]
    catalogos = cargar_catalogos_desde_filas(filas_catalogo)

    hoja = libro[hoja_datos] if hoja_datos in libro.sheetnames else libro.worksheets[0]

    # Encabezado -> catalogo, comparando sin acentos por si cambia el archivo.
    equivalencias = {_normaliza_texto(k): v for k, v in ENCABEZADOS_A_CATALOGO.items()}
    si_no = {_normaliza_texto(k) for k in ENCABEZADOS_SI_NO}

    columnas_traducidas = []
    for celda in hoja[fila_encabezados]:
        encabezado = _normaliza_texto(celda.value)
        if not encabezado:
            continue
        if encabezado in equivalencias:
            nombre_catalogo = equivalencias[encabezado]
            catalogo = catalogos.get(nombre_catalogo, {})
            if catalogo:
                columnas_traducidas.append((celda.column, catalogo, False, nombre_catalogo))
        elif encabezado in si_no:
            columnas_traducidas.append((celda.column, None, True, None))

    total = 0
    for indice_columna, catalogo, es_si_no, nombre_catalogo in columnas_traducidas:
        for fila in range(fila_encabezados + 1, hoja.max_row + 1):
            celda = hoja.cell(row=fila, column=indice_columna)
            if celda.value is None:
                continue
            if es_si_no:
                nuevo = traducir_si_no(celda.value)
            else:
                nuevo = traducir(celda.value, catalogo, nombre_catalogo=nombre_catalogo)
            if nuevo != celda.value:
                celda.value = nuevo
                total += 1
        # Los nombres son mas largos que las claves: se amplia la columna.
        letra = get_column_letter(indice_columna)
        ancho_actual = hoja.column_dimensions[letra].width or 0
        if ancho_actual < ANCHO_MINIMO_TRADUCIDAS:
            hoja.column_dimensions[letra].width = ANCHO_MINIMO_TRADUCIDAS

    #Copia de la hoja con el filtro de Excel de America del Sur ya aplicado
    if agregar_resumen:
        _, visibles = agregar_hoja_filtrada(
            libro, hoja.title, fila_encabezados=fila_encabezados,
            solo_renglones_filtrados=solo_renglones_filtrados)
        print(f"  Hoja 'América del Sur': {visibles:,} renglones visibles con el filtro")

    libro.save(ruta_salida_final)
    columnas = [get_column_letter(c) for c, _, _, _ in columnas_traducidas]
    print(f"  Columnas traducidas: {', '.join(columnas) if columnas else 'ninguna'}")
    print(f"  Celdas actualizadas: {total:,}")
    print(f"  Archivo generado:    {ruta_salida_final}")
    return ruta_salida_final


if __name__ == "__main__":
    inicio = time.perf_counter()
    archivos = sys.argv[1:] or ARCHIVOS
    for archivo in archivos:
        if not os.path.exists(archivo):
            print(f"[!] No se encontro: {archivo}")
            continue
        print(f"\nProcesando: {archivo}")
        convertir(archivo)
    print()
    print(reporte_claves_sin_catalogo())
    print(f"\nElapsed time: {time.perf_counter() - inicio:.1f} s")
