#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 ACTUALIZACIÓN DEL CATÁLOGO CENTRALIZADO (SIREC y SAP)
================================================================================

Toma los movimientos pendientes (#N/D) de las hojas ``ValidacionProp`` y
``ValidacionNoProp`` del archivo ``Consulta_IDEspeciales_v2_XXXX.xlsm`` y los
inserta en la hoja ``ID_Esp<AAAA>=TablaExtendida`` del
``CentralizadoCatálogos_SIRECySAP.xlsx``.

Reglas implementadas (las que pediste):

1. Pendientes  = llaves de la hoja de validación que NO están ni en la columna
   CENTRALIZADO ni en la columna ANTERIOR (exactamente lo que evalúa la fórmula
   ``=IFERROR(VLOOKUP(D5,O:O,1,0),VLOOKUP(D5,M:M,1,0))`` y que produce el #N/D).

2. Mes de ocurrencia = AAAAMM de la ``Fecha de Captura`` del contrato
   (``aHTP_FecCaptura`` / ``aHTNP_FecCaptura``) en las hojas Prop / NoProp.
   Ese valor va a la columna **ENTRA** del catálogo.

3. Se INSERTA una fila (no se sobreescribe nada): el resto de los registros se
   recorre hacia abajo. El punto de inserción es:
      a) debajo del ÚLTIMO movimiento de la MISMA CEDENTE dentro de ese mes;
      b) si no hay, debajo del último renglón del bloque de ese mes;
      c) si el mes todavía no existe, debajo del bloque del mes anterior más
         cercano (para un mes nuevo eso equivale al final de la tabla).

4. Los datos que no vienen en la Consulta (Identificación, DR, LN, SUBLN, MGA,
   País/Territorio de Riesgo, Binders, Cedente, Clasificación, MGA 2) se copian
   de un renglón "plantilla" del propio catálogo, elegido con esta jerarquía:
      0) MAPEO_MANUAL (lo que tú fijes a mano manda sobre todo)
      1) "ID anterior CCC-CIA-NN-AAAA" encontrado en Observaciones
      2) mismo Corredor + Compañía + Contrato  (año de suscripción más reciente)
      3) misma Compañía + Contrato
      4) mismo Corredor + Compañía          (contrato más cercano)
      5) misma Compañía                     (contrato más cercano)
   Cada renglón se reporta con su nivel de confianza para que puedas revisar.

5. Todas las OFERTAS de los renglones nuevos se ponen en CERO.

6. Las columnas de fórmula (A = Llave, Q = Llave 2) se escriben como fórmula,
   igual que el resto de la hoja, y se regeneran para TODOS los renglones para
   que sigan apuntando a su propio renglón después del recorrido.

FORMATO: el archivo de salida se genera editando el XML del .xlsx original, no
reconstruyéndolo. Se conservan estilos, anchos, filtros, etiqueta de
confidencialidad (customXml), comentarios, propiedades y todo lo demás. Sólo
cambian: la hoja del catálogo, el rango del autofiltro y el calcChain (que Excel
regenera solo).

--------------------------------------------------------------------------------
USO
--------------------------------------------------------------------------------
    pip install openpyxl
    python actualizar_catalogo_centralizado.py

Ajusta las rutas en el bloque CONFIGURACIÓN. Puedes también pasarlas por línea
de comandos:

    python actualizar_catalogo_centralizado.py --consulta "...xlsm" \
        --centralizado "...xlsx" --salida "...xlsx"

    # sólo ver qué haría, sin escribir el archivo:
    python actualizar_catalogo_centralizado.py --simular
================================================================================
"""

from __future__ import annotations

import argparse
import datetime as _dt
import html
import os
import re
import sys
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("Falta openpyxl.  Instálalo con:  pip install openpyxl")


# ==============================================================================
# CONFIGURACIÓN
# ==============================================================================

# Carpeta donde tienes los dos archivos (por default, la carpeta del script).
CARPETA = os.path.dirname(os.path.abspath(__file__))

RUTA_CONSULTA = os.path.join(CARPETA, "Consulta_IDEspeciales_v2_0826.xlsm")
RUTA_CENTRALIZADO = os.path.join(CARPETA, "CentralizadoCatálogos_SIRECySAP.xlsx")
# Si lo dejas en None se genera "<nombre>_actualizado.xlsx" junto al original.
RUTA_SALIDA: str | None = None

# Hoja destino dentro del centralizado. None = se detecta sola (la que termina
# en "=TablaExtendida" con el año más reciente).
HOJA_CATALOGO: str | None = None

# Fila del encabezado de la hoja del catálogo (los datos empiezan en la
# siguiente).
FILA_ENCABEZADO = 3

# Valor que se pone en la columna "Oferta" de los renglones nuevos.
OFERTA_NUEVOS = 0

# True  -> si dentro del mes ya hay movimientos de la MISMA CEDENTE, el renglón
#          nuevo se pega debajo de ellos (aunque estén más arriba en la hoja).
# False -> siempre se va al final del bloque del mes.
AGRUPAR_POR_CEDENTE = True

# Hojas de la Consulta.
HOJAS_VALIDACION = {"P": "ValidacionProp", "NP": "ValidacionNoProp"}
HOJAS_DETALLE = {"P": ("Prop", "Prop_02"), "NP": ("NoProp", "NoProp_02")}
FILA_ENCABEZADO_VALIDACION = 4
FILA_ENCABEZADO_DETALLE = 2


# ------------------------------------------------------------------------------
# MAPEO MANUAL  (aquí mandas tú)
# ------------------------------------------------------------------------------
# La búsqueda de la plantilla ("cuál se parece más") es, por naturaleza, un
# criterio humano. El script la resuelve automáticamente, pero si quieres forzar
# de qué renglón del catálogo se copian los datos, escríbelo aquí:
#
#     "llave_nueva": "llave_del_catalogo_que_se_usa_como_plantilla"
#
# Los que están precargados son los que ya revisé contra el catálogo del
# 2026-08; si el mes que viene aparecen otros, el script los resuelve solo y te
# los reporta con su nivel de confianza para que decidas.
MAPEO_MANUAL: dict[str, str] = {
    # --- PROPORCIONAL -----------------------------------------------------
    # Hamilton: hay varios "Hamilton"; el que se parece es el de Fianzas
    # (mismo contrato 12 del año anterior).
    "0-555-12-2026-1": "0-555-12-2025-1",      # Hamilton Insurance (Fianzas)
    "0-903-15-2026-1": "0-903-15-2025-1",      # Talanx / HDI Fianzas Chile
    # Estos dos traen "ID anterior ..." en Observaciones (cambio de corredor):
    "0-903-28-2026-1": "296-903-10-2025-1",    # Talanx / HDI Chile Agro
    "0-903-29-2026-1": "296-903-12-2025-1",    # Talanx / HDI Chile Agro
    "0-1144-1-2026-1": "0-1144-1-2025-1",      # TEMPO Property
    "0-1144-6-2026-1": "0-1144-6-2025-1",      # TEMPO Property
    # OJO: Echo entra por un corredor nuevo (75) y con un tratado de
    # Agricultura India; en el catálogo sólo existe Echo / Chile / Cono Sur.
    # Se usa el contrato 1 de Echo como plantilla, pero REVÍSALO: es probable
    # que País/Territorio de Riesgo y la Línea de Negocio deban cambiar.
    "75-1029-1-2026-1": "0-1029-1-2025-2",     # Echo Reinsurance  << revisar
    "413-1135-1-2026-1": "413-1135-1-2025-1",  # TRISURA
    "413-1135-2-2026-1": "413-1135-2-2025-1",  # TRISURA
    # --- NO PROPORCIONAL --------------------------------------------------
    "0-903-14-2026-2": "0-903-14-2025-2",      # Talanx / HDI Fianzas Chile
    "0-903-30-2026-2": "296-903-8-2025-2",     # Talanx / HDI Chile Agro
    "0-903-31-2026-2": "296-903-11-2025-2",    # Talanx / HDI Chile Agro
    "0-1029-4-2026-2": "0-1029-4-2025-2",      # Echo
    "204-946-1-2026-2": "204-946-1-2025-2",    # Defense Antilles Property
    "204-946-2-2026-2": "204-946-2-2025-2",    # Defense Antilles Property
    "204-946-4-2026-2": "204-946-4-2025-2",    # Defense Antilles Property
    "394-240-1-2026-2": "394-240-1-2025-2",    # Hannover
}


# ==============================================================================
# LAYOUT DE LA HOJA DEL CATÁLOGO  (1 = columna A)
# ==============================================================================

COL = {
    "LLAVE": 1,        # A  fórmula
    "IDENT": 2,        # B
    "COR": 3,          # C
    "CIA": 4,          # D
    "CONTRATO": 5,     # E
    "OFERTA": 6,       # F
    "SUSC": 7,         # G
    "TIPOREA": 8,      # H
    "DR_ID": 9,        # I
    "DR": 10,          # J
    "LN_ID": 11,       # K
    "LN": 12,          # L
    "SUBLN": 13,       # M
    "MGA": 14,         # N
    "PAIS": 15,        # O
    "TERR": 16,        # P
    "LLAVE2": 17,      # Q  fórmula
    "BINDER_SEG": 18,  # R
    "BINDER_GEN": 19,  # S
    "CEDENTE": 20,     # T
    "CLASIF": 21,      # U
    "MGA2": 22,        # V
    "ENTRA": 23,       # W
}

# Columnas que se toman de la Consulta.
COLS_DESDE_CONSULTA = ("COR", "CIA", "CONTRATO", "SUSC", "TIPOREA")
# Columnas que se copian del renglón plantilla del propio catálogo.
COLS_DESDE_PLANTILLA = (
    "IDENT", "DR_ID", "DR", "LN_ID", "LN", "SUBLN", "MGA", "PAIS", "TERR",
    "BINDER_SEG", "BINDER_GEN", "CEDENTE", "CLASIF", "MGA2",
)
COLS_FORMULA = ("LLAVE", "LLAVE2")
ULTIMA_COL_DATOS = COL["ENTRA"]

# Fórmulas tal como están guardadas en el archivo (TEXTJOIN se guarda con el
# prefijo _xlfn porque es función "nueva" de Excel).
F_LLAVE = '_xlfn.TEXTJOIN("-",,C{r}:E{r},G{r}:H{r})'
F_LLAVE2 = '_xlfn.TEXTJOIN(" ",,TEXT(C{r},"0000"),TEXT(D{r},"0000"),TEXT(E{r},"00"),F{r},G{r})'


# ==============================================================================
# UTILERÍAS
# ==============================================================================

from openpyxl.utils.datetime import to_excel as _to_excel


def a_entero(valor: Any) -> int | None:
    """Convierte a entero tolerando basura de Excel.

    En las hojas Prop/NoProp hay renglones donde 'Número Contrato' y 'Año
    Vigencia' quedaron con formato de fecha: openpyxl los devuelve como
    datetime (p. ej. 1905-07-18 en vez de 2026). Aquí se regresan al número de
    serie que realmente son.
    """
    if valor is None:
        return None
    if isinstance(valor, bool):
        return int(valor)
    if isinstance(valor, (_dt.datetime, _dt.date)):
        return int(round(_to_excel(valor)))
    if isinstance(valor, (int, float)):
        return int(round(valor))
    texto = str(valor).strip().replace(",", "")
    if not texto:
        return None
    try:
        return int(round(float(texto)))
    except ValueError:
        return None


def llave_de(cor, cia, contrato, susc, tiporea) -> str | None:
    partes = [a_entero(x) for x in (cor, cia, contrato, susc, tiporea)]
    if any(p is None for p in partes):
        return None
    return "-".join(str(p) for p in partes)


def a_aaaamm(valor: Any) -> int | None:
    if isinstance(valor, _dt.datetime):
        return valor.year * 100 + valor.month
    if isinstance(valor, _dt.date):
        return valor.year * 100 + valor.month
    return None


def col_letra(indice: int) -> str:
    letra = ""
    while indice:
        indice, resto = divmod(indice - 1, 26)
        letra = chr(65 + resto) + letra
    return letra


def indice_columna(letra: str) -> int:
    n = 0
    for ch in letra:
        n = n * 26 + (ord(ch) - 64)
    return n


def mes_anterior(aaaamm: int) -> int:
    anio, mes = divmod(aaaamm, 100)
    return (anio - 1) * 100 + 12 if mes == 1 else anio * 100 + (mes - 1)


# ==============================================================================
# 1) LECTURA DE LA CONSULTA
# ==============================================================================

@dataclass
class Pendiente:
    tipo: str              # "P" | "NP"
    llave: str
    cor: int
    cia: int
    contrato: int
    susc: int
    tiporea: int
    nombre_cia: str
    fila_validacion: int
    mes: int | None = None            # ENTRA (AAAAMM)
    referencia: str = ""
    observaciones: str = ""
    id_anterior: str | None = None    # llave sugerida por Observaciones
    # se llenan más adelante
    plantilla: int | None = None      # fila del catálogo usada como plantilla
    confianza: str = ""
    motivo: str = ""
    fila_destino: int | None = None
    orden: int = 0                    # orden de captura dentro de la misma ancla
    fila_final: Any = None            # fila que ocupó en el archivo de salida
    avisos: list[str] = field(default_factory=list)


_RE_ID_ANTERIOR = re.compile(
    r"ID\s+anterior\s+(\d+)\s*-\s*(\d+)\s*-\s*(\d+)\s*-\s*(\d{4})", re.IGNORECASE
)


def leer_pendientes(ruta_consulta: str) -> list[Pendiente]:
    print(f"[1/5] Leyendo Consulta: {os.path.basename(ruta_consulta)} ...")
    wb = openpyxl.load_workbook(ruta_consulta, data_only=True, read_only=True)
    try:
        pendientes: list[Pendiente] = []
        for tipo, hoja in HOJAS_VALIDACION.items():
            if hoja not in wb.sheetnames:
                print(f"      ! No existe la hoja '{hoja}', se omite.")
                continue
            pendientes += _pendientes_de_hoja(wb[hoja], tipo)

        detalle = {tipo: _leer_detalle(wb, tipo) for tipo in HOJAS_DETALLE}
        for p in pendientes:
            registro = detalle[p.tipo].get(p.llave)
            if registro is None:
                p.avisos.append("No se encontró el contrato en las hojas de detalle")
                continue
            p.mes = a_aaaamm(registro["fecha_captura"])
            p.referencia = registro["referencia"]
            p.observaciones = registro["observaciones"]
            if p.mes is None:
                p.avisos.append("Sin Fecha de Captura: no se pudo determinar el mes")
            m = _RE_ID_ANTERIOR.search(p.observaciones or "")
            if m:
                cor, cia, contrato, anio = (int(g) for g in m.groups())
                p.id_anterior = f"{cor}-{cia}-{contrato}-{anio}-{p.tiporea}"
        return pendientes
    finally:
        wb.close()


def _pendientes_de_hoja(ws, tipo: str) -> list[Pendiente]:
    """Replica la fórmula de la columna 'validación'.

    #N/D  <=>  la llave no está ni en ANTERIOR (col O) ni en CENTRALIZADO
    (col M). Se recalcula en vez de confiar en el valor en caché, para que
    funcione aunque el archivo no se haya recalculado.
    """
    filas = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=22, values_only=True))
    inicio = FILA_ENCABEZADO_VALIDACION  # 0-based -> primera fila de datos

    centralizado = {str(f[12]).strip() for f in filas[inicio:] if f[12] not in (None, "")}
    anterior = {str(f[14]).strip() for f in filas[inicio:] if f[14] not in (None, "")}
    conocidas = centralizado | anterior

    pendientes: list[Pendiente] = []
    for offset, fila in enumerate(filas[inicio:], start=inicio + 1):
        cor, cia = a_entero(fila[4]), a_entero(fila[5])
        contrato, susc, tre = a_entero(fila[8]), a_entero(fila[9]), a_entero(fila[10])
        llave = llave_de(cor, cia, contrato, susc, tre)
        if llave is None:
            continue
        if llave in conocidas:
            continue
        cacheado = fila[1]
        if not (isinstance(cacheado, str) and cacheado.startswith("#")) and cacheado is not None:
            # El valor en caché dice que sí existe pero el recálculo dice que no.
            print(f"      ! {llave}: la celda de validación tiene '{cacheado}' en caché "
                  f"pero la llave no aparece en CENTRALIZADO/ANTERIOR. Se trata como pendiente.")
        pendientes.append(Pendiente(
            tipo=tipo, llave=llave, cor=cor, cia=cia, contrato=contrato,
            susc=susc, tiporea=tre, nombre_cia=str(fila[7] or "").strip(),
            fila_validacion=offset,
        ))
    print(f"      {ws.title}: {len(pendientes)} pendiente(s)")
    return pendientes


def _leer_detalle(wb, tipo: str) -> dict[str, dict]:
    """Indexa Prop/Prop_02 (o NoProp/NoProp_02) por llave."""
    detalle: dict[str, dict] = {}
    for hoja in HOJAS_DETALLE[tipo]:
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        filas = list(ws.iter_rows(min_row=1, max_row=ws.max_row,
                                  max_col=ws.max_column, values_only=True))
        if len(filas) < FILA_ENCABEZADO_DETALLE:
            continue
        encabezado = filas[FILA_ENCABEZADO_DETALLE - 1]
        idx = _indices_detalle(encabezado)
        for fila in filas[FILA_ENCABEZADO_DETALLE:]:
            llave = llave_de(fila[0], fila[3], fila[6], fila[7], fila[8])
            if llave is None or llave in detalle:
                continue
            detalle[llave] = {
                "hoja": hoja,
                "fecha_captura": _en(fila, idx.get("captura")),
                "referencia": str(_en(fila, idx.get("referencia")) or "").strip(),
                "observaciones": str(_en(fila, idx.get("observaciones")) or "").strip(),
            }
    return detalle


def _indices_detalle(encabezado: tuple) -> dict[str, int]:
    """Localiza columnas por nombre técnico (aHTP_* / aHTNP_*)."""
    idx: dict[str, int] = {}
    for i, valor in enumerate(encabezado):
        nombre = str(valor or "")
        if nombre.endswith("_FecCaptura"):
            idx.setdefault("captura", i)
        elif nombre.endswith("_RefOrig"):
            idx.setdefault("referencia", i)
        elif nombre.endswith("_Observaciones"):
            idx.setdefault("observaciones", i)
    return idx


def _en(fila: tuple, i: int | None):
    return fila[i] if i is not None and i < len(fila) else None


# ==============================================================================
# 2) LECTURA DEL CATÁLOGO CENTRALIZADO
# ==============================================================================

@dataclass
class FilaCatalogo:
    fila: int
    cor: int | None
    cia: int | None
    contrato: int | None
    susc: int | None
    tiporea: int | None
    llave: str | None
    ident: str
    cedente: str
    clasif: str
    entra: int | None


def detectar_hoja(wb) -> str:
    candidatas = [h for h in wb.sheetnames if h.replace(" ", "").lower().endswith("=tablaextendida")]
    if not candidatas:
        raise SystemExit("No encontré ninguna hoja '...=TablaExtendida' en el centralizado.")
    def anio(nombre: str) -> int:
        m = re.search(r"(\d{4})", nombre)
        return int(m.group(1)) if m else 0
    return max(candidatas, key=anio)


def leer_catalogo(ruta: str, hoja: str | None) -> tuple[str, list[FilaCatalogo]]:
    print(f"[2/5] Leyendo Centralizado: {os.path.basename(ruta)} ...")
    wb = openpyxl.load_workbook(ruta, data_only=False)
    try:
        nombre = hoja or detectar_hoja(wb)
        ws = wb[nombre]
        filas: list[FilaCatalogo] = []
        for r in range(FILA_ENCABEZADO + 1, ws.max_row + 1):
            cia = a_entero(ws.cell(r, COL["CIA"]).value)
            if cia is None:
                continue
            cor = a_entero(ws.cell(r, COL["COR"]).value)
            contrato = a_entero(ws.cell(r, COL["CONTRATO"]).value)
            susc = a_entero(ws.cell(r, COL["SUSC"]).value)
            tiporea = a_entero(ws.cell(r, COL["TIPOREA"]).value)
            filas.append(FilaCatalogo(
                fila=r, cor=cor, cia=cia, contrato=contrato, susc=susc, tiporea=tiporea,
                llave=llave_de(cor, cia, contrato, susc, tiporea),
                ident=str(ws.cell(r, COL["IDENT"]).value or "").strip(),
                cedente=str(ws.cell(r, COL["CEDENTE"]).value or "").strip(),
                clasif=str(ws.cell(r, COL["CLASIF"]).value or "").strip(),
                entra=a_entero(ws.cell(r, COL["ENTRA"]).value),
            ))
        print(f"      Hoja '{nombre}': {len(filas)} registros "
              f"(filas {filas[0].fila}-{filas[-1].fila})")
        return nombre, filas
    finally:
        wb.close()


# ==============================================================================
# 3) BÚSQUEDA DE LA PLANTILLA ("cuál se parece más")
# ==============================================================================

def elegir_plantilla(p: Pendiente, catalogo: list[FilaCatalogo]) -> None:
    por_llave: dict[str, FilaCatalogo] = {}
    for f in catalogo:
        if f.llave:
            por_llave.setdefault(f.llave, f)

    # 0) Mapeo manual
    forzada = MAPEO_MANUAL.get(p.llave)
    if forzada:
        elegida = por_llave.get(forzada)
        if elegida is not None:
            _asignar(p, elegida, "MANUAL", f"MAPEO_MANUAL -> {forzada}")
            return
        p.avisos.append(f"MAPEO_MANUAL apunta a '{forzada}' pero esa llave no está en el catálogo")

    # 1) "ID anterior ..." en Observaciones
    if p.id_anterior:
        elegida = por_llave.get(p.id_anterior)
        if elegida is None:
            # el año exacto puede no existir: se toma el más reciente de ese contrato
            cor, cia, contrato = (int(x) for x in p.id_anterior.split("-")[:3])
            elegida = _mas_reciente([f for f in catalogo
                                     if f.cor == cor and f.cia == cia and f.contrato == contrato])
        if elegida is not None:
            _asignar(p, elegida, "ALTA", f"Observaciones: ID anterior {p.id_anterior}")
            return

    mismos_cia = [f for f in catalogo if f.cia == p.cia]
    if not mismos_cia:
        p.avisos.append("No hay ningún registro de esa compañía en el catálogo: "
                        "hay que capturarlo a mano")
        return

    # 2) mismo corredor + compañía + contrato
    elegida = _mas_reciente([f for f in mismos_cia
                             if f.cor == p.cor and f.contrato == p.contrato])
    if elegida is not None:
        _asignar(p, elegida, "ALTA", "Mismo corredor, compañía y contrato")
        return

    # 3) misma compañía + contrato (cambió el corredor)
    elegida = _mas_reciente([f for f in mismos_cia if f.contrato == p.contrato])
    if elegida is not None:
        _asignar(p, elegida, "MEDIA", "Misma compañía y contrato (cambió el corredor)")
        return

    # 4) mismo corredor + compañía, contrato más cercano
    elegida = _mas_cercano([f for f in mismos_cia if f.cor == p.cor], p.contrato)
    if elegida is not None:
        _asignar(p, elegida, "MEDIA", "Mismo corredor y compañía, contrato más cercano")
        return

    # 5) misma compañía, contrato más cercano
    elegida = _mas_cercano(mismos_cia, p.contrato)
    if elegida is not None:
        _asignar(p, elegida, "BAJA", "Sólo coincide la compañía: REVISAR")


def _mas_reciente(cands: list[FilaCatalogo]) -> FilaCatalogo | None:
    if not cands:
        return None
    return max(cands, key=lambda f: (f.susc or 0, f.entra or 0, f.fila))


def _mas_cercano(cands: list[FilaCatalogo], contrato: int) -> FilaCatalogo | None:
    if not cands:
        return None
    return min(cands, key=lambda f: (abs((f.contrato or 0) - contrato),
                                     -(f.susc or 0), -(f.entra or 0)))


def _asignar(p: Pendiente, f: FilaCatalogo, confianza: str, motivo: str) -> None:
    p.plantilla = f.fila
    p.confianza = confianza
    p.motivo = f"{motivo}  [fila {f.fila}: {f.llave} · {f.ident}]"
    if f.tiporea is not None and f.tiporea != p.tiporea:
        p.avisos.append(
            f"La plantilla es Tipo Rea {f.tiporea} y el nuevo es {p.tiporea}: "
            f"se ajustó la Clasificación"
        )


def clasificacion_ajustada(clasif_plantilla: str, tiporea_plantilla: int | None,
                           tiporea_nuevo: int) -> str | None:
    """Corrige 'Proporcional' / 'No Proporcional' si la plantilla es del otro tipo."""
    if tiporea_plantilla is None or tiporea_plantilla == tiporea_nuevo:
        return None
    texto = clasif_plantilla or ""
    if tiporea_nuevo == 2:  # no proporcional
        if "No Proporcional" in texto:
            return None
        return texto.replace("Proporcional", "No Proporcional") if "Proporcional" in texto else None
    if "No Proporcional" in texto:
        return texto.replace("No Proporcional", "Proporcional")
    return None


# ==============================================================================
# 4) PUNTO DE INSERCIÓN
# ==============================================================================

def calcular_destinos(pendientes: list[Pendiente], catalogo: list[FilaCatalogo]) -> None:
    """Decide, para cada pendiente, después de qué fila del catálogo original va.

    Se procesa de uno en uno y los renglones ya colocados se agregan a la lista
    de trabajo (con un número de fila fraccionario, p. ej. 795.1), de modo que
    dos movimientos de la misma cedente en el mismo mes quedan pegados uno
    debajo del otro.
    """
    trabajo: list[FilaCatalogo] = list(catalogo)
    listos = [p for p in pendientes if p.mes is not None and p.plantilla is not None]
    # Orden de captura: mes, cedente, compañía, contrato, tipo de reaseguro.
    listos.sort(key=lambda p: (p.mes, _cedente_de(p, catalogo), p.cia, p.contrato, p.tiporea))

    contador: dict[int, int] = defaultdict(int)
    for orden, p in enumerate(listos):
        cedente = _cedente_de(p, catalogo)
        ancla = _buscar_ancla(trabajo, p.mes, cedente)
        p.fila_destino = int(ancla)
        p.orden = orden
        contador[p.fila_destino] += 1
        trabajo.append(FilaCatalogo(
            fila=ancla + contador[p.fila_destino] / 1000.0,
            cor=p.cor, cia=p.cia, contrato=p.contrato, susc=p.susc, tiporea=p.tiporea,
            llave=p.llave, ident="", cedente=cedente, clasif="", entra=p.mes,
        ))


def _cedente_de(p: Pendiente, catalogo: list[FilaCatalogo]) -> str:
    if p.plantilla is not None:
        for f in catalogo:
            if f.fila == p.plantilla:
                return f.cedente or p.nombre_cia
    return p.nombre_cia


def _buscar_ancla(filas: list[FilaCatalogo], mes: int, cedente: str) -> float:
    """Regresa el número de fila (del catálogo original) DESPUÉS de la cual se
    inserta el renglón nuevo."""
    del_mes = [f for f in filas if f.entra == mes]

    # a) mismo mes + misma cedente -> debajo del último de ellos
    if AGRUPAR_POR_CEDENTE and cedente:
        mismos = [f for f in del_mes if f.cedente == cedente]
        if mismos:
            return max(f.fila for f in mismos)

    # b) mismo mes -> debajo del último renglón de ese mes
    if del_mes:
        return max(f.fila for f in del_mes)

    # c) el mes todavía no existe -> debajo del último renglón del mes anterior
    #    más cercano (para un mes nuevo esto cae al final de la tabla)
    anteriores = [f for f in filas if f.entra is not None and f.entra < mes]
    if anteriores:
        mes_previo = max(f.entra for f in anteriores)
        return max(f.fila for f in anteriores if f.entra == mes_previo)

    # d) no hay nada anterior -> al final de la tabla
    return max(f.fila for f in filas)


# ==============================================================================
# 5) ESCRITURA:  cirugía sobre el XML del .xlsx
# ==============================================================================

_RE_ROW = re.compile(r"<row\b[^>]*?(?:/>|>.*?</row>)", re.S)
_RE_CELL = re.compile(r"<c\b[^>]*?(?:/>|>.*?</c>)", re.S)
_RE_REF = re.compile(r'\br="([A-Z]+)(\d+)"')
_RE_ROWNUM = re.compile(r'<row\b[^>]*?\br="(\d+)"')
_RE_STYLE = re.compile(r'\bs="(\d+)"')


class Celda:
    __slots__ = ("col", "estilo", "tipo", "interior", "extras")

    def __init__(self, xml: str):
        m = _RE_REF.search(xml)
        self.col = indice_columna(m.group(1))
        s = _RE_STYLE.search(xml)
        self.estilo = s.group(1) if s else None
        t = re.search(r'\bt="([^"]+)"', xml)
        self.tipo = t.group(1) if t else None
        # atributos extra (cm, vm, ph) que hay que conservar
        self.extras = " ".join(
            f'{k}="{v}"' for k, v in re.findall(r'\b(cm|vm|ph)="([^"]*)"', xml)
        )
        if xml.rstrip().endswith("/>"):
            self.interior = None
        else:
            self.interior = xml[xml.index(">") + 1: xml.rindex("</c>")]

    def render(self, fila: int, estilo: str | None = None) -> str:
        s = estilo if estilo is not None else self.estilo
        attrs = f'r="{col_letra(self.col)}{fila}"'
        if s is not None:
            attrs += f' s="{s}"'
        if self.tipo:
            attrs += f' t="{self.tipo}"'
        if self.extras:
            attrs += " " + self.extras
        if self.interior is None:
            return f"<c {attrs}/>"
        return f"<c {attrs}>{self.interior}</c>"


class Renglon:
    def __init__(self, xml: str):
        self.num = int(_RE_ROWNUM.search(xml).group(1))
        fin_attrs = xml.index(">")
        self.attrs = xml[len("<row"):fin_attrs].rstrip("/").strip()
        self.attrs = re.sub(r'\br="\d+"\s*', "", self.attrs).strip()
        if xml.rstrip().endswith("/>"):
            self.celdas: dict[int, Celda] = {}
        else:
            interior = xml[fin_attrs + 1: xml.rindex("</row>")]
            self.celdas = {}
            for c in _RE_CELL.finditer(interior):
                celda = Celda(c.group(0))
                self.celdas[celda.col] = celda

    def estilo(self, col: int, default: str | None = None) -> str | None:
        c = self.celdas.get(col)
        return c.estilo if c is not None and c.estilo is not None else default

    def render(self, num: int) -> str:
        attrs = f'r="{num}"' + (f" {self.attrs}" if self.attrs else "")
        if not self.celdas:
            return f"<row {attrs}/>"
        cuerpo = "".join(self.celdas[c].render(num) for c in sorted(self.celdas))
        return f"<row {attrs}>{cuerpo}</row>"


class CadenasCompartidas:
    """Lector / escritor de xl/sharedStrings.xml."""

    def __init__(self, xml: str | None):
        self.xml = xml
        self.nuevas: list[str] = []
        self.indice: dict[str, int] = {}
        self.total = 0
        if xml is None:
            return
        cuerpo = xml[xml.index("<sst"):]
        items = re.findall(r"<si\b[^>]*?(?:/>|>.*?</si>)", cuerpo, re.S)
        self.total = len(items)
        for i, si in enumerate(items):
            texto = "".join(html.unescape(t) for t in re.findall(r"<t\b[^>]*>(.*?)</t>", si, re.S))
            self.indice.setdefault(texto, i)

    def idx(self, texto: str) -> int:
        if texto in self.indice:
            return self.indice[texto]
        nuevo = self.total + len(self.nuevas)
        self.nuevas.append(texto)
        self.indice[texto] = nuevo
        return nuevo

    def serializar(self, celdas_agregadas: int) -> str | None:
        if self.xml is None or not self.nuevas:
            return None
        extra = "".join(
            f'<si><t xml:space="preserve">{html.escape(t, quote=False)}</t></si>'
            for t in self.nuevas
        )
        salida = self.xml.replace("</sst>", extra + "</sst>")
        def _sube(m):
            attr, valor = m.group(1), int(m.group(2))
            if attr == "uniqueCount":
                return f'{attr}="{self.total + len(self.nuevas)}"'
            return f'{attr}="{valor + celdas_agregadas}"'
        return re.sub(r'\b(count|uniqueCount)="(\d+)"', _sube, salida, count=2)


def escribir_salida(ruta_origen: str, ruta_salida: str, hoja: str,
                    pendientes: list[Pendiente], catalogo: list[FilaCatalogo]) -> None:
    print(f"[5/5] Escribiendo {os.path.basename(ruta_salida)} ...")

    with zipfile.ZipFile(ruta_origen) as z:
        partes = {n: z.read(n) for n in z.namelist()}
        orden = list(z.namelist())

    ruta_hoja = _ruta_de_hoja(partes, hoja)
    sheet_xml = partes[ruta_hoja].decode("utf-8")

    cadenas = CadenasCompartidas(
        partes["xl/sharedStrings.xml"].decode("utf-8") if "xl/sharedStrings.xml" in partes else None
    )

    ini = sheet_xml.index("<sheetData")
    fin = sheet_xml.index("</sheetData>") + len("</sheetData>")
    cabeza, cuerpo, cola = sheet_xml[:ini], sheet_xml[ini:fin], sheet_xml[fin:]

    renglones = [Renglon(m.group(0)) for m in _RE_ROW.finditer(cuerpo)]
    por_num = {r.num: r for r in renglones}

    # --- agrupar inserciones por ancla -------------------------------------
    a_insertar: dict[int, list[Pendiente]] = defaultdict(list)
    for p in pendientes:
        if p.fila_destino is not None:
            a_insertar[p.fila_destino].append(p)
    for lista in a_insertar.values():
        lista.sort(key=lambda x: x.orden)
    anclas = sorted(a_insertar)
    total_nuevas = sum(len(v) for v in a_insertar.values())
    if total_nuevas == 0:
        print("      No hay nada que insertar.")
        return

    def desplazamiento(fila: int) -> int:
        return sum(len(a_insertar[a]) for a in anclas if a < fila)

    ultima_fila_datos = max(f.fila for f in catalogo)
    info_catalogo = {f.fila: f for f in catalogo}

    # --- construir la nueva secuencia de renglones -------------------------
    salida: list[str] = []
    celdas_texto_nuevas = 0
    for r in sorted(por_num):
        renglon = por_num[r]
        nuevo_num = r + desplazamiento(r)
        # A y Q se reescriben como fórmula normal apuntando a su propio renglón
        if FILA_ENCABEZADO < r <= ultima_fila_datos:
            _reescribir_formulas(renglon, nuevo_num, info_catalogo.get(r))
        salida.append(renglon.render(nuevo_num))

        for i, p in enumerate(a_insertar.get(r, [])):
            fila_nueva = nuevo_num + 1 + i
            xml, agregadas = _renglon_nuevo(p, renglon, por_num, fila_nueva, cadenas, info_catalogo)
            celdas_texto_nuevas += agregadas
            salida.append(xml)
            p.fila_final = fila_nueva

    nuevo_cuerpo = "<sheetData>" + "".join(salida) + "</sheetData>"

    # --- rangos que dependen del número de filas ---------------------------
    cabeza = _ajustar_rango(cabeza, r"<dimension ref=\"([A-Z]+)(\d+):([A-Z]+)(\d+)\"/>",
                            desplazamiento, solo_fin=True)
    cola = _ajustar_referencias(cola, desplazamiento)

    sheet_xml = cabeza + nuevo_cuerpo + cola
    partes[ruta_hoja] = sheet_xml.encode("utf-8")

    # --- sharedStrings -----------------------------------------------------
    nuevo_sst = cadenas.serializar(celdas_texto_nuevas)
    if nuevo_sst is not None:
        partes["xl/sharedStrings.xml"] = nuevo_sst.encode("utf-8")

    # --- workbook.xml: _FilterDatabase + recálculo al abrir ----------------
    partes["xl/workbook.xml"] = _ajustar_workbook(
        partes["xl/workbook.xml"].decode("utf-8"), hoja, desplazamiento
    ).encode("utf-8")

    # --- calcChain: se elimina, Excel lo reconstruye solo ------------------
    partes, orden = _quitar_calcchain(partes, orden)

    os.makedirs(os.path.dirname(os.path.abspath(ruta_salida)) or ".", exist_ok=True)
    with zipfile.ZipFile(ruta_salida, "w", zipfile.ZIP_DEFLATED) as z:
        for nombre in orden:
            if nombre in partes:
                z.writestr(nombre, partes[nombre])
    print(f"      {total_nuevas} renglón(es) insertado(s).")


def _ruta_de_hoja(partes: dict[str, bytes], hoja: str) -> str:
    wb = partes["xl/workbook.xml"].decode("utf-8")
    m = re.search(r'<sheet[^>]*name="%s"[^>]*r:id="([^"]+)"' % re.escape(hoja), wb)
    if not m:
        raise SystemExit(f"No encontré la hoja '{hoja}' en workbook.xml")
    rels = partes["xl/_rels/workbook.xml.rels"].decode("utf-8")
    m2 = re.search(r'Id="%s"[^>]*Target="([^"]+)"' % re.escape(m.group(1)), rels)
    if not m2:
        raise SystemExit("No pude resolver la relación de la hoja destino")
    destino = m2.group(1).lstrip("/")
    return destino if destino.startswith("xl/") else "xl/" + destino


def _reescribir_formulas(renglon: Renglon, num: int, info: FilaCatalogo | None) -> None:
    """Convierte las fórmulas compartidas de A y Q en fórmulas normales del
    renglón, para que sigan apuntando a sus propias celdas después del
    recorrido."""
    for clave, plantilla in (("LLAVE", F_LLAVE), ("LLAVE2", F_LLAVE2)):
        col = COL[clave]
        celda = renglon.celdas.get(col)
        if celda is None or celda.interior is None or "<f" not in celda.interior:
            continue
        valor = _valor_formula(clave, info)
        cache = f"<v>{html.escape(valor, quote=False)}</v>" if valor is not None else ""
        celda.tipo = "str"
        celda.interior = f"<f>{html.escape(plantilla.format(r=num), quote=False)}</f>{cache}"


def _valor_formula(clave: str, info: FilaCatalogo | None) -> str | None:
    if info is None:
        return None
    c, d, e, g, h = info.cor, info.cia, info.contrato, info.susc, info.tiporea
    if clave == "LLAVE":
        partes = [x for x in (c, d, e, g, h) if x is not None]
        return "-".join(str(x) for x in partes) if partes else None
    if None in (c, d, e, g):
        return None
    return f"{c:04d} {d:04d} {e:02d} 0 {g}"


def _renglon_nuevo(p: Pendiente, ancla: Renglon, por_num: dict[int, Renglon],
                   num: int, cadenas: CadenasCompartidas,
                   info_catalogo: dict[int, FilaCatalogo]) -> tuple[str, int]:
    """Arma el XML del renglón nuevo: formato del renglón de arriba, datos de la
    Consulta y de la plantilla."""
    plantilla = por_num[p.plantilla]
    info_plantilla = info_catalogo.get(p.plantilla)
    celdas: dict[int, str] = {}
    agregadas = 0

    # estilo por default del renglón ancla (el de arriba), como cuando insertas
    # una fila en Excel
    def estilo(col: int) -> str | None:
        return ancla.estilo(col, plantilla.estilo(col))

    valores_consulta = {
        "COR": p.cor, "CIA": p.cia, "CONTRATO": p.contrato,
        "SUSC": p.susc, "TIPOREA": p.tiporea,
    }

    for clave, col in COL.items():
        if clave in COLS_FORMULA:
            formula = (F_LLAVE if clave == "LLAVE" else F_LLAVE2).format(r=num)
            cache = (p.llave if clave == "LLAVE"
                     else f"{p.cor:04d} {p.cia:04d} {p.contrato:02d} {OFERTA_NUEVOS} {p.susc}")
            s = estilo(col)
            attrs = f'r="{col_letra(col)}{num}"' + (f' s="{s}"' if s else "") + ' t="str"'
            celdas[col] = (f"<c {attrs}><f>{html.escape(formula, quote=False)}</f>"
                           f"<v>{html.escape(cache, quote=False)}</v></c>")
        elif clave in COLS_DESDE_CONSULTA:
            s = estilo(col)
            attrs = f'r="{col_letra(col)}{num}"' + (f' s="{s}"' if s else "")
            celdas[col] = f"<c {attrs}><v>{valores_consulta[clave]}</v></c>"
        elif clave == "OFERTA":
            s = estilo(col)
            attrs = f'r="{col_letra(col)}{num}"' + (f' s="{s}"' if s else "")
            celdas[col] = f"<c {attrs}><v>{OFERTA_NUEVOS}</v></c>"
        elif clave == "ENTRA":
            s = estilo(col)
            attrs = f'r="{col_letra(col)}{num}"' + (f' s="{s}"' if s else "")
            celdas[col] = f"<c {attrs}><v>{p.mes}</v></c>"
        else:  # copiado de la plantilla
            origen = plantilla.celdas.get(col)
            if clave == "CLASIF" and info_plantilla is not None:
                nueva = clasificacion_ajustada(info_plantilla.clasif,
                                               info_plantilla.tiporea, p.tiporea)
                if nueva:
                    s = estilo(col)
                    attrs = f'r="{col_letra(col)}{num}"' + (f' s="{s}"' if s else "") + ' t="s"'
                    celdas[col] = f"<c {attrs}><v>{cadenas.idx(nueva)}</v></c>"
                    agregadas += 1
                    continue
            if origen is None:
                s = estilo(col)
                celdas[col] = f'<c r="{col_letra(col)}{num}"' + (f' s="{s}"' if s else "") + "/>"
            else:
                celdas[col] = origen.render(num, estilo(col))

    # columnas vacías a la derecha (X, Y...) para conservar la banda de formato
    for col, celda in ancla.celdas.items():
        if col > ULTIMA_COL_DATOS and celda.interior is None:
            celdas[col] = celda.render(num)

    attrs = f'r="{num}"' + (f" {ancla.attrs}" if ancla.attrs else "")
    cuerpo = "".join(celdas[c] for c in sorted(celdas))
    return f"<row {attrs}>{cuerpo}</row>", agregadas


def _ajustar_rango(texto: str, patron: str, desplazamiento, solo_fin: bool = True) -> str:
    def _sub(m):
        c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
        nr1 = r1 if solo_fin else r1 + desplazamiento(r1)
        nr2 = r2 + desplazamiento(r2 + 1)
        return m.group(0).replace(f"{c1}{r1}:{c2}{r2}", f"{c1}{nr1}:{c2}{nr2}")
    return re.sub(patron, _sub, texto)


def _ajustar_referencias(cola: str, desplazamiento) -> str:
    """Recorre autoFilter, sortState, mergeCells, formato condicional y demás
    rangos que viven después de </sheetData>."""
    def _rango(texto: str) -> str:
        m = re.fullmatch(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?", texto)
        if not m:
            return texto
        c1, r1 = m.group(1), int(m.group(2))
        nr1 = r1 + desplazamiento(r1)
        if m.group(3) is None:
            return f"{c1}{nr1}"
        c2, r2 = m.group(3), int(m.group(4))
        return f"{c1}{nr1}:{c2}{r2 + desplazamiento(r2 + 1)}"

    def _sub(m):
        attr, valor = m.group(1), m.group(2)
        return f'{attr}="' + " ".join(_rango(x) for x in valor.split()) + '"'

    return re.sub(r'\b(ref|sqref)="([A-Z0-9: ]+)"', _sub, cola)


def _ajustar_workbook(wb_xml: str, hoja: str, desplazamiento) -> str:
    nombre_ref = f"'{hoja}'" if re.search(r"[^A-Za-z0-9_]", hoja) else hoja

    def _sub(m):
        c1, r1, c2, r2 = m.group(2), int(m.group(3)), m.group(4), int(m.group(5))
        return (f"{m.group(1)}${c1}${r1 + desplazamiento(r1)}:"
                f"${c2}${r2 + desplazamiento(r2 + 1)}")

    patron = (r"(" + re.escape(nombre_ref) + r"!)\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)")
    wb_xml = re.sub(patron, _sub, wb_xml)

    # que Excel recalcule al abrir (las fórmulas de A y Q se reescribieron)
    if "<calcPr" in wb_xml:
        if "fullCalcOnLoad" not in wb_xml:
            wb_xml = re.sub(r"<calcPr\b([^>]*?)/>", r'<calcPr\1 fullCalcOnLoad="1"/>', wb_xml)
    else:
        wb_xml = wb_xml.replace("</workbook>", '<calcPr fullCalcOnLoad="1"/></workbook>')
    return wb_xml


def _quitar_calcchain(partes: dict[str, bytes], orden: list[str]):
    if "xl/calcChain.xml" not in partes:
        return partes, orden
    rels = partes["xl/_rels/workbook.xml.rels"].decode("utf-8")
    m = re.search(r'<Relationship[^>]*Target="calcChain\.xml"[^>]*/>', rels)
    if m:
        partes["xl/_rels/workbook.xml.rels"] = rels.replace(m.group(0), "").encode("utf-8")
    ct = partes["[Content_Types].xml"].decode("utf-8")
    ct = re.sub(r'<Override[^>]*PartName="/xl/calcChain\.xml"[^>]*/>', "", ct)
    partes["[Content_Types].xml"] = ct.encode("utf-8")
    del partes["xl/calcChain.xml"]
    return partes, [n for n in orden if n != "xl/calcChain.xml"]


# ==============================================================================
# REPORTE
# ==============================================================================

def imprimir_reporte(pendientes: list[Pendiente], catalogo: list[FilaCatalogo]) -> None:
    info = {f.fila: f for f in catalogo}
    print()
    print("=" * 110)
    print("RESUMEN DE MOVIMIENTOS")
    print("=" * 110)
    encabezado = f"{'LLAVE':<20}{'MES':<8}{'FILA':<7}{'CONF.':<8}{'IDENTIFICACIÓN':<38}ORIGEN"
    print(encabezado)
    print("-" * 110)
    for p in sorted(pendientes, key=lambda x: (x.mes or 0, x.llave)):
        fila = getattr(p, "fila_final", None)
        ident = info[p.plantilla].ident if p.plantilla in info else "—"
        print(f"{p.llave:<20}{p.mes or '—':<8}{fila or '—':<7}{p.confianza:<8}"
              f"{ident[:36]:<38}{p.motivo[:60]}")
        for aviso in p.avisos:
            print(f"{'':<20}  ⚠  {aviso}")
    print("-" * 110)
    revisar = [p for p in pendientes if p.confianza in ("BAJA", "") or p.avisos]
    if revisar:
        print(f"\n⚠  {len(revisar)} renglón(es) que conviene revisar a mano: "
              + ", ".join(p.llave for p in revisar))
    print()


# ==============================================================================
# MAIN
# ==============================================================================

def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--consulta", default=RUTA_CONSULTA)
    ap.add_argument("--centralizado", default=RUTA_CENTRALIZADO)
    ap.add_argument("--salida", default=RUTA_SALIDA)
    ap.add_argument("--hoja", default=HOJA_CATALOGO)
    ap.add_argument("--simular", action="store_true",
                    help="sólo muestra el reporte, no escribe el archivo")
    args = ap.parse_args(argv)

    for ruta in (args.consulta, args.centralizado):
        if not os.path.isfile(ruta):
            return _error(f"No encontré el archivo: {ruta}")

    salida = args.salida
    if not salida:
        base, ext = os.path.splitext(args.centralizado)
        salida = f"{base}_actualizado{ext}"

    pendientes = leer_pendientes(args.consulta)
    if not pendientes:
        print("\nNo hay movimientos pendientes: el catálogo ya está al corriente.")
        return 0

    hoja, catalogo = leer_catalogo(args.centralizado, args.hoja)

    print("[3/5] Buscando el registro más parecido en el catálogo ...")
    for p in pendientes:
        elegir_plantilla(p, catalogo)

    print("[4/5] Calculando el punto de inserción por mes y cedente ...")
    calcular_destinos(pendientes, catalogo)

    sin_lugar = [p for p in pendientes if p.fila_destino is None]
    for p in sin_lugar:
        p.avisos.append("No se insertó: falta mes o plantilla")

    if args.simular:
        for p in pendientes:
            if p.fila_destino is not None:
                p.fila_final = f"~{p.fila_destino}+"
        imprimir_reporte(pendientes, catalogo)
        print("Simulación: no se escribió ningún archivo.")
        return 0

    escribir_salida(args.centralizado, salida, hoja,
                    [p for p in pendientes if p.fila_destino is not None], catalogo)
    imprimir_reporte(pendientes, catalogo)
    print(f"Listo:  {salida}")
    return 1 if sin_lugar else 0


def _error(msg: str) -> int:
    print(f"ERROR: {msg}", file=sys.stderr)
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
