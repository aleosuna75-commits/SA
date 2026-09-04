# -*- coding: utf-8 -*-
"""
================================================================================
 verificar_excel_formulas.py · Comprueba las fórmulas vivas de la hoja Mensual
--------------------------------------------------------------------------------
 Planeación Financiera (BP&A) · Reaseguradora Patria (GPV)

 La hoja «Mensual» de Validacion_Prima_Devengada.xlsx trae columnas con fórmulas Excel
 (ámbar) para que cada número se pueda auditar celda por celda. Este script recalcula
 el libro con LibreOffice (si está instalado) y compara el resultado de cada fórmula
 con el valor que calculó Python en comparacion_mensual.csv. Si no hay LibreOffice,
 sólo revisa que las fórmulas apunten a las columnas correctas.

 Uso:  python verificar_excel_formulas.py        (desde la carpeta del proyecto)
================================================================================
"""
import os
import re
import shutil
import subprocess
import sys
import tempfile

import numpy as np
import pandas as pd

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, "salidas", "Validacion_Prima_Devengada.xlsx")
CSV = os.path.join(BASE, "salidas", "comparacion_mensual.csv")
TOL = 0.01          # USD: las fórmulas deben reproducir Python al centavo

# columna de la hoja -> columna de comparacion_mensual.csv (o cómo derivarla)
FORMULAS = {
    "PR": "PR", "BRUTO_CAL": "BRUTO_CAL", "NETO_CAL": "NETO_CAL",
    "dBRUTO_real": "dBRUTO_real", "dBRUTO_CAL": "dBRUTO_CAL", "dNETO_real": "dNETO_real", "dNETO_CAL": "dNETO_CAL",
    "PD_tom_real": "PD_tom_real", "PD_tom_CAL": "PD_tom_CAL", "PD_ret_real": "PD_ret_real", "PD_ret_CAL": "PD_ret_CAL",
    "ratio_CAL_real": None,   # = PND_CAL / PND_real
}


def _encabezados(ws):
    return {str(ws.cell(3, j).value).split("\n")[0].strip(): j for j in range(1, ws.max_column + 1) if ws.cell(3, j).value}


def revisar_referencias():
    """Sin recalcular: cada fórmula debe referenciar las columnas que dice la hoja Formulas."""
    import openpyxl
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Mensual"]
    H = _encabezados(ws)
    from openpyxl.utils import get_column_letter as L
    col = {k: L(v) for k, v in H.items()}
    esperado = {
        "PR": f"={col['PE']}4*(1-{col['ces']}4)",
        "BRUTO_CAL": f"={col['PND_CAL']}4*{col['IS_eff']}4*(1+{col['g']}4+{col['mr']}4)",
        "NETO_CAL": f"={col['BRUTO_CAL']}4-{col['PND_CAL']}4*{col['IS_eff']}4*{col['c']}4",
        "dBRUTO_real": f'=IF({col["Grupo"]}4={col["Grupo"]}3,{col["BRUTO_real"]}4-{col["BRUTO_real"]}3,"")',
        "PD_tom_CAL": f'=IF({col["dBRUTO_CAL"]}4="","",{col["PE"]}4-{col["dBRUTO_CAL"]}4)',
        "PD_ret_real": f'=IF({col["dNETO_real"]}4="","",{col["PR"]}4-{col["dNETO_real"]}4)',
    }
    malos = []
    for k, f in esperado.items():
        v = ws.cell(4, H[k]).value
        if str(v).replace(" ", "") != f.replace(" ", ""):
            malos.append(f"{k}: hoja={v!r} esperado={f!r}")
    return malos


def recalcular_con_libreoffice():
    """Abre el libro con LibreOffice Calc y lo vuelve a guardar: al guardarlo recalcula todas las fórmulas.
    Devuelve la ruta del libro recalculado, o None si no hay LibreOffice (o no trae el módulo Calc)."""
    exe = shutil.which("soffice") or shutil.which("libreoffice")
    if not exe:
        return None
    tmp = tempfile.mkdtemp(prefix="recalc_")
    perfil = "file:///" + os.path.join(tmp, "perfil").replace("\\", "/").lstrip("/")   # perfil limpio, no toca el del usuario
    r = subprocess.run([exe, f"-env:UserInstallation={perfil}", "--headless", "--calc", "--convert-to", "xlsx",
                        "--outdir", tmp, os.path.abspath(XLSX)], capture_output=True, text=True, timeout=900)
    salida = os.path.join(tmp, os.path.basename(XLSX))
    if r.returncode != 0 or not os.path.exists(salida):
        print("LibreOffice no pudo recalcular el libro. Mensaje:", (r.stdout + r.stderr).strip()[-400:])
        print("(si dice «source file could not be loaded», falta el módulo Calc: instala libreoffice-calc)")
        return None
    return salida


def comparar_valores(xlsx_recalc):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_recalc, data_only=True)
    ws = wb["Mensual"]
    H = _encabezados(ws)
    filas = []
    for i in range(4, ws.max_row + 1):
        filas.append({k: ws.cell(i, j).value for k, j in H.items()})
    x = pd.DataFrame(filas)
    x["PERIODO"] = pd.to_numeric(x["Mes contable"], errors="coerce").astype(int)
    py = pd.read_csv(CSV).sort_values(["Grupo", "PERIODO"]).reset_index(drop=True)
    py["ratio_CAL_real"] = py["PND_CAL"] / py["PND_real"]
    m = x.merge(py, on=["Grupo", "PERIODO"], suffixes=("_xl", "_py"))
    assert len(m) == len(py) == len(x), (len(m), len(py), len(x))
    res = []
    for k, src in FORMULAS.items():
        a = pd.to_numeric(m[f"{k}_xl"] if f"{k}_xl" in m else m[k], errors="coerce")
        b = m[(src or k) + ("_py" if (src or k) + "_py" in m else "")]
        b = pd.to_numeric(b, errors="coerce")
        ok = a.isna() == b.isna()
        d = (a - b).abs()
        res.append(dict(columna=k, filas=len(m), vacios_iguales=bool(ok.all()),
                        max_abs_dif=float(np.nanmax(d.to_numpy())) if d.notna().any() else 0.0))
    return pd.DataFrame(res)


if __name__ == "__main__":
    if not os.path.exists(XLSX) or not os.path.exists(CSV):
        raise SystemExit("Faltan salidas/Validacion_Prima_Devengada.xlsx o comparacion_mensual.csv; corre antes validar_prima_devengada.py")
    malos = revisar_referencias()
    print("Referencias de las fórmulas:", "OK" if not malos else "FALLA")
    for mm in malos:
        print("  ", mm)
    rec = recalcular_con_libreoffice()
    if rec is None:
        print("LibreOffice no está instalado: no puedo recalcular aquí. Abre el archivo en Excel y compara la hoja Mensual "
              "contra salidas/comparacion_mensual.csv (mismas columnas).")
        sys.exit(1 if malos else 0)
    r = comparar_valores(rec)
    pd.set_option("display.width", 200)
    print(r.to_string(index=False))
    ok = (r["max_abs_dif"] <= TOL).all() and r["vacios_iguales"].all() and not malos
    print("\nRESULTADO:", "OK · las fórmulas del Excel reproducen los valores de Python" if ok else "FALLA · revisar")
    sys.exit(0 if ok else 1)
