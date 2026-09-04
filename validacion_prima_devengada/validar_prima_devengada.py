# -*- coding: utf-8 -*-
"""
================================================================================
 validar_prima_devengada.py · Validación del FND (MEC) contra la prima devengada REAL
--------------------------------------------------------------------------------
 Planeación Financiera (BP&A) · Reaseguradora Patria (GPV)

 PREGUNTA QUE RESPONDE
   ¿La prima devengada que produce el modelo (RRC = PT · FND · IS + gastos + MR,
   IRR = BEL · %cesión) cuadra con la prima devengada REAL?

 DEFINICIÓN DEL REAL (misma que Integración Dim, hoja ER_2026)
   Prima devengada tomada     = Prima emitida  − Δ RRC bruta   (tomada)
   Prima devengada retenida   = Prima retenida − Δ RRC neta    (bruta − IRR)
   con los saldos de la base BEL-IRR-MR (BD_Montos_RRC_SONR, en USD).

 QUÉ AÍSLA
   Todo lo que NO es FND (índice de siniestralidad, gasto, MR, % cesión) se toma del
   REAL mes a mes; así la única diferencia modelo–real es el factor de devengamiento
   y la prima base. Ratio modelo/real de la RRC == ratio de prima no devengada.

 VARIANTES DE FND QUE SE RECONSTRUYEN (por registro de prima de la BD del MEC)
   NT_reg_m  : Nota Técnica (24-avos) por antigüedad de REGISTRO, cuentas mensuales
               (tabla xPND 'NA' de los reforecast); no proporcional -> prorrata por cohorte
   NT_reg_t  : idem con cuentas trimestrales para proporcional (xPND '3')
   MEC_pub   : TablaFND publicada del MEC v2 (PF+): vector por antigüedad desde INICIO
               DE VIGENCIA; Vida con vector propio, resto vector de cartera
   MEC_prop  : PF+ con vector propio de cada ramo
   MEC_reg   : curva PF+ de cartera aplicada sobre la antigüedad de REGISTRO
   CAL       : calibrado: FND = max(0, NT(k_reg) − δ_ramo) para proporcional/facultativo,
               prorrata por cohorte para no proporcional; δ_ramo ajustado por mínimos
               cuadrados contra la prima no devengada real (202301–202605)

 UNIDADES: todo en USD (la base BEL-IRR-MR está en USD; la prima MXN de la BD se
 convierte al TC de cierre del mes de registro). El TC de 2019–2021 es aproximado
 (promedios Banxico) y sólo afecta colas anteriores a 2023.
================================================================================
"""
from __future__ import annotations
import os, sys, json
import numpy as np
import pandas as pd

# El script se ancla a SU PROPIA CARPETA (ponlo en «…\OneDrive - GPV\Documents»).
# Lee los insumos de la subcarpeta insumos/ (la genera preparar_insumos.py) o, si no
# existe, de la misma carpeta; escribe todo en salidas/.
BASE = os.path.dirname(os.path.abspath(__file__))
INS = os.path.join(BASE, "insumos") if os.path.isdir(os.path.join(BASE, "insumos")) else BASE
OUT = os.path.join(BASE, "salidas")
os.makedirs(OUT, exist_ok=True)

T0 = 202201                      # primer mes con RRC real en la base
T1 = 202605                      # último mes con RRC real; se SOBREESCRIBE con el último
                                 # mes con saldo de real_rrc_long.csv al cargarlo
CAL0 = 202301                    # ventana de calibración / reporte (evita la rampa 2022)
CAL0_GRUPO = {"CAT": 202401}     # CAT: el índice TEV/Hidro sólo es consistente desde 2024
H = 72                           # horizonte de los vectores MEC

# ----------------------------------------------------------------------------
# Catálogo: ramo BD-MEC -> grupo de validación -> columnas RAM_ de la base real
# (40=RC, 50=MyT, 90=Autos, 110=Diversos: mapeo del script RRC / ER real; el
#  NOMBRE_RAMO de generar_output_mec.py tiene esos cuatro nombres cruzados)
# ----------------------------------------------------------------------------
GRUPOS = {
    "Vida":     dict(ramo=10,  ram=["RAM_10"],                  er=["Vida"]),
    "AyE":      dict(ramo=30,  ram=["RAM_30", "RAM_34", "RAM_37"], er=["Acc Per.", "GMM", "Salud"]),
    "RC":       dict(ramo=40,  ram=["RAM_40"],                  er=["Resp. Civil"]),
    "MyT":      dict(ramo=50,  ram=["RAM_50"],                  er=["MyT"]),
    "Incendio": dict(ramo=60,  ram=["RAM_60"],                  er=["Incendio"]),
    "CAT":      dict(ramo=70,  ram=["RAM_71", "RAM_73"],        er=["Terremoto", "HyORH"]),
    "Agro":     dict(ramo=80,  ram=["RAM_80"],                  er=["Agropecuario"]),
    "Autos":    dict(ramo=90,  ram=["RAM_90"],                  er=["Autos"]),
    "Credito":  dict(ramo=100, ram=["RAM_100"],                 er=["Crédito"]),
    "Diversos": dict(ramo=110, ram=["RAM_110"],                 er=["Diversos"]),
}
RAMO2GRUPO = {v["ramo"]: g for g, v in GRUPOS.items()}
# códigos de índice de siniestralidad por columna RAM (prioridad por fecha)
IS_CODES = {"RAM_10": ["10"], "RAM_30": ["31", "30"], "RAM_34": ["35", "34", "30"],
            "RAM_37": ["39", "37", "30"], "RAM_40": ["40"], "RAM_50": ["50"], "RAM_60": ["60"],
            "RAM_71": ["TEV"], "RAM_73": ["Hidro"], "RAM_80": ["80"], "RAM_90": ["90"],
            "RAM_100": ["100"], "RAM_110": ["110"]}
IS_CAT_DESDE = {"RAM_71": 202401, "RAM_73": 202401}   # primer mes con índice CAT en base consistente

# Nota Técnica / tabla xPND de los reforecast (antigüedad de registro 0..11, cuentas mensuales)
NT_M = np.array([0.95890411, 0.876712329, 0.791780822, 0.706849315, 0.624657534, 0.539726027,
                 0.457534247, 0.37260274, 0.295890411, 0.210958904, 0.126027397, 0.043835616])
DESP_TRIM = (3 - 1) / 2 * 30 / 365          # regla M4: (t−1)/2 · 30/365 → 0.0822 (verificada vs xPND '3')

# TC aproximado 2019–2021 (promedio mensual Banxico FIX, redondeado). Sólo afecta
# prima registrada antes de 2022 → reservas de 2022 y colas multianuales (Vida).
TC_APROX = {
    2019: [19.16, 19.20, 19.26, 18.97, 19.14, 19.30, 19.05, 19.65, 19.60, 19.36, 19.34, 19.11],
    2020: [18.83, 18.89, 22.37, 24.16, 23.42, 22.29, 22.44, 22.11, 21.71, 21.30, 20.42, 19.93],
    2021: [19.90, 20.30, 20.72, 20.02, 19.94, 20.05, 19.93, 20.07, 20.03, 20.51, 20.87, 20.85],
}


def midx(aaaamm) -> np.ndarray:
    a = np.asarray(aaaamm, dtype=np.int64)
    return (a // 100) * 12 + (a % 100 - 1)


def aaaamm(m: int) -> int:
    return (m // 12) * 100 + (m % 12 + 1)


def meses(t0, t1):
    return [aaaamm(m) for m in range(int(midx(t0)), int(midx(t1)) + 1)]


# ============================================================================
# 1 · INSUMOS
# ============================================================================
def cargar_tc() -> pd.Series:
    tc = pd.read_csv(os.path.join(INS, "tc_mensual_bd.csv")).set_index("Periodo")["TC"]
    extra = {y * 100 + i + 1: v for y, vals in TC_APROX.items() for i, v in enumerate(vals)}
    tc = pd.concat([pd.Series(extra), tc]).sort_index()
    return tc


def cargar_input(tc: pd.Series) -> pd.DataFrame:
    df = pd.read_csv(os.path.join(INS, "input_mec_bd.csv"))
    df = df[(df["Fuente"] == "BD") & (df["Origen"] == "Real") & (df["Periodo"] <= T1)].copy()
    df = df[df["Ramo"].isin(RAMO2GRUPO)].copy()          # fianzas (130–170) sin RRC en la base
    df["Grupo"] = df["Ramo"].map(RAMO2GRUPO)
    df["TC"] = df["Periodo"].map(tc)
    if df["TC"].isna().any():
        faltan = sorted(df.loc[df["TC"].isna(), "Periodo"].unique())
        raise SystemExit(f"Sin TC para los periodos {faltan}")
    df["P_USD"] = df["PrimaDevMes"] / df["TC"]
    df["m_reg"] = midx(df["Periodo"])
    df["m_coh"] = midx(df["CohorteAAAAMM"])
    return df.reset_index(drop=True)


def cargar_real() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Devuelve (real por grupo × mes, IS efectivo) con:
    BEL, GTO, MR, IRR, BRUTO, NETO (USD), PND_real = Σ_sub BEL/IS, IS_eff = BEL/PND_real,
    g=GTO/BEL, mr=MR/BEL, c=IRR/BEL."""
    r = pd.read_csv(os.path.join(INS, "real_rrc_long.csv"))
    r = r[r["PERIODO"] >= T0]
    # último mes con saldo real: la base trae los meses futuros en cero
    tot = r.groupby("PERIODO")["RRC BRUTO"].sum()
    t1 = int(tot[tot.abs() > 0].index.max())
    r = r[r["PERIODO"] <= t1]
    is_ = pd.read_csv(os.path.join(INS, "is_rrc_real.csv")).set_index("Fecha")
    is_.columns = [str(c).strip() for c in is_.columns]
    per = sorted(r["PERIODO"].unique())
    is_ = is_.reindex(per)
    # IS por columna RAM: primer código disponible por fecha, luego ffill/bfill
    is_ram = {}
    for ram, codes in IS_CODES.items():
        s = pd.Series(np.nan, index=per, dtype=float)
        for c in codes:
            if c in is_.columns:
                s = s.fillna(pd.to_numeric(is_[c], errors="coerce"))
        if ram in IS_CAT_DESDE:
            # TEV/Hidro: antes de 2024 el 'Ind Sin RRC' de HParametros está en otra base
            # (valores 0.5–8.6 frente a 0.11–0.27 desde 2024); se descarta y se rellena
            # hacia atrás con el primer valor consistente. El índice implícito
            # BEL/PND_modelo de CAT es estable (~0.30–0.33 en 2022), lo que confirma que
            # el problema está en el índice y no en el devengamiento.
            s[s.index < IS_CAT_DESDE[ram]] = np.nan
        s = s.replace(0, np.nan).ffill().bfill()
        is_ram[ram] = s
    is_ram = pd.DataFrame(is_ram)
    r = r.merge(is_ram.stack().rename("IS").reset_index().rename(columns={"level_0": "PERIODO", "level_1": "RAM"}),
                on=["PERIODO", "RAM"], how="left")
    r["PND_sub"] = r["RRC BEL"] / r["IS"]
    r["Grupo"] = r["RAM"].map({ram: g for g, v in GRUPOS.items() for ram in v["ram"]})
    r = r.dropna(subset=["Grupo"])
    g = (r.groupby(["Grupo", "PERIODO"])
          .agg(BEL=("RRC BEL", "sum"), GTO=("RRC GTO", "sum"), MR=("RRC MR", "sum"),
               IRR=("RRC IRR", "sum"), BRUTO=("RRC BRUTO", "sum"), NETO=("RRC NETO", "sum"),
               PND_real=("PND_sub", "sum")).reset_index())
    g["IS_eff"] = g["BEL"] / g["PND_real"]
    g["g"] = g["GTO"] / g["BEL"]
    g["mr"] = g["MR"] / g["BEL"]
    g["c"] = g["IRR"] / g["BEL"]
    return g, is_ram, t1


def cargar_vectores() -> tuple[dict, np.ndarray, list]:
    v = pd.read_csv(os.path.join(INS, "mec_vectores_h72.csv"), index_col=0)
    cart = v.loc["CARTERA"].to_numpy(dtype=float)
    vec = {str(k): v.loc[k].to_numpy(dtype=float) for k in v.index if k != "CARTERA"}
    abiertos = ["10"]                      # decisión M3 del MEC v2: sólo Vida abre
    return vec, cart, abiertos


def cargar_cesion_er() -> pd.DataFrame:
    """% de prima cedida por grupo y año (ER real): para construir la prima retenida."""
    er = pd.read_csv(os.path.join(INS, "er_real_primas.csv"))
    out = {}
    for per, anio in [(202312, 2023), (202412, 2024), (202510, 2025)]:
        e = er[er["Periodo"] == per].set_index("Concepto")
        for g, spec in GRUPOS.items():
            pe = e.loc["PRIMA EMITIDA", spec["er"]].sum()
            ced = e.loc["(-) PRIMA CEDIDA / RETROCEDIDA", spec["er"]].sum()
            out[(g, anio)] = ced / pe if pe else 0.0
    ces = pd.Series(out).rename("ces").reset_index()
    ces.columns = ["Grupo", "Año", "ces"]
    return ces


# ============================================================================
# 2 · MOTOR: prima no devengada por variante
# ============================================================================
def _vec_lookup(vec: np.ndarray, k: np.ndarray) -> np.ndarray:
    """Valor del vector en la antigüedad k (k<0 → 0; k≥len → 0, la cola ya es ~0)."""
    out = np.zeros(len(k))
    ok = (k >= 0) & (k < len(vec))
    out[ok] = vec[k[ok]]
    return out


def nt_desp(desp: float) -> np.ndarray:
    """Regla M4 del MEC: FND_t(k) = NT(k) − desp, acotado a [0, 1] (δ<0 no puede pasar de 100%)."""
    return np.clip(NT_M - desp, 0.0, 1.0)


class Motor:
    def __init__(self, inp: pd.DataFrame, vec: dict, cart: np.ndarray, abiertos: list):
        self.inp = inp
        self.vec, self.cart, self.abiertos = vec, cart, abiertos
        self.P = inp["P_USD"].to_numpy()
        self.mreg = inp["m_reg"].to_numpy()
        self.mcoh = inp["m_coh"].to_numpy()
        self.tipo = inp["TipoRea"].to_numpy()
        self.ramo = inp["Ramo"].to_numpy().astype(str)
        self.grupo = inp["Grupo"].to_numpy()
        self.grupos = sorted(GRUPOS)
        self.gidx = np.array([self.grupos.index(g) for g in self.grupo])
        self.vec_pub = {r: (vec[r] if r in abiertos else cart) for r in vec}

    def fnd(self, variante: str, t: int, params: dict | None = None) -> np.ndarray:
        k_reg = int(midx(t)) - self.mreg
        k_coh = int(midx(t)) - self.mcoh
        prop = self.tipo != 2                    # proporcional (1) y facultativo (3)
        f = np.zeros(len(self.P))
        if variante in ("NT_reg_m", "NT_reg_t", "CAL"):
            if variante == "NT_reg_m":
                f[prop] = _vec_lookup(NT_M, k_reg[prop])
            elif variante == "NT_reg_t":
                t1 = self.tipo == 1
                f[t1] = _vec_lookup(nt_desp(DESP_TRIM), k_reg[t1])
                t3 = self.tipo == 3
                f[t3] = _vec_lookup(NT_M, k_reg[t3])
            else:                                # CAL: δ por grupo
                for gi, g in enumerate(self.grupos):
                    m = prop & (self.gidx == gi)
                    f[m] = _vec_lookup(nt_desp(params["delta"][g]), k_reg[m])
            np_ = ~prop                          # no proporcional: prorrata por fechas ≈ PF+ cartera por cohorte
            f[np_] = _vec_lookup(self.cart, k_coh[np_])
        elif variante == "MEC_pub":
            for r, v in self.vec_pub.items():
                m = self.ramo == r
                f[m] = _vec_lookup(v, k_coh[m])
        elif variante == "MEC_prop":
            for r, v in self.vec.items():
                m = self.ramo == r
                f[m] = _vec_lookup(v, k_coh[m])
        elif variante == "MEC_reg":
            f = _vec_lookup(self.cart, k_reg)
        elif variante == "NT_reg_susc":
            # Diagnóstico del filtro de reforecastRRC: «año(mes registro) <= Susc» (la prima
            # positiva registrada en un año posterior al de suscripción queda fuera de la RRC).
            # Proxy: año de suscripción ≈ año de inicio de vigencia (la BD del MEC no trae Susc).
            f[prop] = _vec_lookup(NT_M, k_reg[prop])
            f[~prop] = _vec_lookup(self.cart, k_coh[~prop])
            tarde = (self.mreg // 12 > self.mcoh // 12) & (self.P > 0)
            f[tarde] = 0.0
        else:
            raise ValueError(variante)
        f[k_reg < 0] = 0.0                       # prima aún no registrada a la fecha de valuación
        return f

    def pnd(self, variante: str, params: dict | None = None) -> pd.DataFrame:
        rows = []
        for t in meses(T0, T1):
            f = self.fnd(variante, t, params)
            s = np.bincount(self.gidx, weights=self.P * f, minlength=len(self.grupos))
            rows.append(pd.Series(s, index=self.grupos, name=t))
        df = pd.DataFrame(rows)
        df.index.name = "PERIODO"
        return df

    def prima_emitida(self) -> pd.DataFrame:
        pe = self.inp.pivot_table(index="Periodo", columns="Grupo", values="P_USD", aggfunc="sum")
        return pe.reindex(meses(T0, T1)).fillna(0.0)


# ============================================================================
# 3 · COMPARACIÓN
# ============================================================================
def largo(pnd: pd.DataFrame, nombre: str) -> pd.DataFrame:
    x = pnd.stack().rename(nombre).reset_index()
    x.columns = ["PERIODO", "Grupo", nombre]
    return x


def comparar(real: pd.DataFrame, pe: pd.DataFrame, pnds: dict, ces: pd.DataFrame, tc: pd.Series) -> pd.DataFrame:
    """Tabla larga por grupo × mes con real y cada variante: PND, RRC bruta/neta, PD tomada/retenida."""
    base = real[["Grupo", "PERIODO", "BEL", "BRUTO", "NETO", "IRR", "PND_real", "IS_eff", "g", "mr", "c"]].copy()
    base = base.merge(largo(pe, "PE"), on=["PERIODO", "Grupo"], how="left")
    base["TC"] = base["PERIODO"].map(tc)
    # Año CONTABLE: el de PERIODO, que es el mes de valuación de la RRC real y el mes
    # contable en que se registra la prima. No es el año de suscripción.
    base["Año"] = base["PERIODO"] // 100
    base = base.merge(ces, on=["Grupo", "Año"], how="left")
    base["ces"] = base.groupby("Grupo")["ces"].transform(lambda s: s.ffill().bfill())
    base["PR"] = base["PE"] * (1 - base["ces"])
    base = base.sort_values(["Grupo", "PERIODO"])
    base["dBRUTO_real"] = base.groupby("Grupo")["BRUTO"].diff()
    base["dNETO_real"] = base.groupby("Grupo")["NETO"].diff()
    base["PD_tom_real"] = base["PE"] - base["dBRUTO_real"]
    base["PD_ret_real"] = base["PR"] - base["dNETO_real"]
    for v, pnd in pnds.items():
        base = base.merge(largo(pnd, f"PND_{v}"), on=["PERIODO", "Grupo"], how="left")
        bel = base[f"PND_{v}"] * base["IS_eff"]
        base[f"BRUTO_{v}"] = bel * (1 + base["g"] + base["mr"])
        base[f"NETO_{v}"] = base[f"BRUTO_{v}"] - bel * base["c"]
        base[f"dBRUTO_{v}"] = base.groupby("Grupo")[f"BRUTO_{v}"].diff()
        base[f"dNETO_{v}"] = base.groupby("Grupo")[f"NETO_{v}"].diff()
        base[f"PD_tom_{v}"] = base["PE"] - base[f"dBRUTO_{v}"]
        base[f"PD_ret_{v}"] = base["PR"] - base[f"dNETO_{v}"]
    return base


def resumen_anual(cmp: pd.DataFrame, variantes: list) -> pd.DataFrame:
    """Prima devengada tomada y retenida por grupo y año: real vs variantes, con error."""
    c = cmp[cmp["PERIODO"] >= CAL0].copy()
    c["Año"] = c["PERIODO"] // 100          # año contable
    # equivalentes en MXN: Δreserva del mes × TC de cierre del mes (así lo lleva el ER:
    # el efecto cambiario del saldo no pasa por la variación de reserva)
    c["PE_MXN"] = c["PE"] * c["TC"]
    c["PD_tom_real_MXN"] = c["PD_tom_real"] * c["TC"]
    c["PD_ret_real_MXN"] = c["PD_ret_real"] * c["TC"]
    agg = {"PE": "sum", "PR": "sum", "PD_tom_real": "sum", "PD_ret_real": "sum",
           "dBRUTO_real": "sum", "dNETO_real": "sum", "PE_MXN": "sum",
           "PD_tom_real_MXN": "sum", "PD_ret_real_MXN": "sum"}
    for v in variantes:
        c[f"PD_tom_{v}_MXN"] = c[f"PD_tom_{v}"] * c["TC"]
        c[f"PD_ret_{v}_MXN"] = c[f"PD_ret_{v}"] * c["TC"]
        agg[f"PD_tom_{v}"] = "sum"; agg[f"PD_ret_{v}"] = "sum"
        agg[f"dBRUTO_{v}"] = "sum"; agg[f"dNETO_{v}"] = "sum"
        agg[f"PD_tom_{v}_MXN"] = "sum"; agg[f"PD_ret_{v}_MXN"] = "sum"
    a = c.groupby(["Grupo", "Año"]).agg(agg).reset_index()
    tot = c.groupby("Año").agg(agg).reset_index(); tot["Grupo"] = "TOTAL"
    a = pd.concat([a, tot], ignore_index=True)
    for v in variantes:
        a[f"err_tom_{v}"] = a[f"PD_tom_{v}"] - a["PD_tom_real"]
        a[f"err%_tom_{v}"] = a[f"err_tom_{v}"] / a["PD_tom_real"]
        a[f"err_ret_{v}"] = a[f"PD_ret_{v}"] - a["PD_ret_real"]
        a[f"err%_ret_{v}"] = a[f"err_ret_{v}"] / a["PD_ret_real"]
    return a


def metricas_pnd(cmp: pd.DataFrame, variantes: list) -> pd.DataFrame:
    """Ajuste de la prima no devengada (== RRC) por grupo: ratio medio modelo/real,
    MAPE mensual y R² sobre 202301–202605."""
    c = cmp[cmp["PERIODO"] >= CAL0]
    rows = []
    for g, d in list(c.groupby("Grupo")) + [("TOTAL", c.groupby("PERIODO").sum(numeric_only=True).reset_index())]:
        if g in CAL0_GRUPO:
            d = d[d["PERIODO"] >= CAL0_GRUPO[g]]
        r = d["PND_real"].to_numpy()
        for v in variantes:
            m = d[f"PND_{v}"].to_numpy()
            ok = r > 0
            rows.append(dict(Grupo=g, Variante=v,
                             ratio=m[ok].sum() / r[ok].sum(),
                             MAPE=np.mean(np.abs(m[ok] / r[ok] - 1)),
                             R2=1 - np.sum((m[ok] - r[ok]) ** 2) / np.sum((r[ok] - r[ok].mean()) ** 2),
                             PND_real_prom=r[ok].mean(), PND_mod_prom=m[ok].mean()))
    return pd.DataFrame(rows)


# ============================================================================
# 4 · CALIBRACIÓN: δ por grupo (regla M4 del MEC sobre antigüedad de registro)
# ============================================================================
def calibrar_delta(motor: Motor, real: pd.DataFrame, grid=None) -> tuple[dict, pd.DataFrame]:
    """Para cada grupo elige δ que minimiza Σ_t (PND_model − PND_real)² en la ventana
    de calibración. Como la contribución no proporcional no depende de δ, se separa."""
    if grid is None:
        grid = np.round(np.arange(-0.30, 0.601, 0.005), 3)
    tt = [t for t in meses(T0, T1) if t >= CAL0]
    prop = motor.tipo != 2
    # PND no proporcional (fijo) y prima proporcional por antigüedad de registro, por grupo y t
    fixed = {g: np.zeros(len(tt)) for g in motor.grupos}
    Pk = {g: np.zeros((len(tt), 12)) for g in motor.grupos}     # prima prop. registrada hace k meses
    for i, t in enumerate(tt):
        k_reg = int(midx(t)) - motor.mreg
        k_coh = int(midx(t)) - motor.mcoh
        fnp = np.zeros(len(motor.P))
        fnp[~prop] = _vec_lookup(motor.cart, k_coh[~prop])
        fnp[k_reg < 0] = 0
        s = np.bincount(motor.gidx, weights=motor.P * fnp, minlength=len(motor.grupos))
        for gi, g in enumerate(motor.grupos):
            fixed[g][i] = s[gi]
            m = prop & (motor.gidx == gi) & (k_reg >= 0) & (k_reg < 12)
            Pk[g][i] = np.bincount(k_reg[m], weights=motor.P[m], minlength=12)[:12]
    rr = real.set_index(["Grupo", "PERIODO"])["PND_real"]
    delta, filas = {}, []
    for g in motor.grupos:
        y = np.array([rr.get((g, t), np.nan) for t in tt])
        ok = ~np.isnan(y) & (np.array(tt) >= CAL0_GRUPO.get(g, CAL0))
        best = None
        for d in grid:
            f = nt_desp(d)
            m = fixed[g] + Pk[g] @ f
            sse = np.sum((m[ok] - y[ok]) ** 2)
            if best is None or sse < best[1]:
                best = (d, sse, m)
        d, sse, m = best
        delta[g] = float(d)
        r2 = 1 - sse / np.sum((y[ok] - y[ok].mean()) ** 2)
        filas.append(dict(Grupo=g, delta=d, meses_equiv_cuentas=1 + 2 * d * 365 / 30,
                          ratio=m[ok].sum() / y[ok].sum(), MAPE=np.mean(np.abs(m[ok] / y[ok] - 1)), R2=r2,
                          PND_real_prom=y[ok].mean(), ventana_desde=CAL0_GRUPO.get(g, CAL0)))
    return delta, pd.DataFrame(filas)


def tabla_fnd_calibrada(delta: dict) -> pd.DataFrame:
    rows = {}
    rows["NT mensual (referencia)"] = NT_M
    rows["NT trimestral (referencia)"] = nt_desp(DESP_TRIM)
    for g, d in delta.items():
        rows[f"{g} (ramo {GRUPOS[g]['ramo']}) δ={d:+.3f}"] = nt_desp(d)
    t = pd.DataFrame(rows).T
    t.columns = [f"k={k}" for k in range(12)]
    return t


# ============================================================================
# 5 · SALIDAS
# ============================================================================
def main():
    global T1
    tc = cargar_tc()
    real, is_ram, T1 = cargar_real()        # fija T1 al último mes con saldo real
    inp = cargar_input(tc)
    vec, cart, abiertos = cargar_vectores()
    ces = cargar_cesion_er()
    motor = Motor(inp, vec, cart, abiertos)

    variantes = ["NT_reg_m", "NT_reg_t", "NT_reg_susc", "MEC_pub", "MEC_prop", "MEC_reg"]
    pnds = {v: motor.pnd(v) for v in variantes}
    delta, cal = calibrar_delta(motor, real)
    pnds["CAL"] = motor.pnd("CAL", dict(delta=delta))
    variantes.append("CAL")

    pe = motor.prima_emitida()
    cmp = comparar(real, pe, pnds, ces, tc)
    anual = resumen_anual(cmp, variantes)
    met = metricas_pnd(cmp, variantes)
    tabla = tabla_fnd_calibrada(delta)

    # ---- consola ----
    pd.set_option("display.width", 250); pd.set_option("display.max_columns", 60)
    pd.set_option("display.float_format", lambda x: f"{x:,.4f}")
    print("\n=== δ calibrado por grupo (FND = max(0, NT(k_reg) − δ)) ===")
    print(cal.to_string(index=False))
    print(f"\n=== Ajuste de la prima no devengada (== RRC) por variante · ventana {CAL0}–{T1} ===")
    piv = met.pivot(index="Grupo", columns="Variante", values="ratio")[variantes]
    print("ratio Σmodelo/Σreal:\n" + piv.to_string())
    piv2 = met.pivot(index="Grupo", columns="Variante", values="MAPE")[variantes]
    print("MAPE mensual:\n" + piv2.to_string())
    print("\n=== Prima devengada TOMADA por año (USD) · real vs variantes · TOTAL ===")
    cols = ["Año", "PE", "PD_tom_real"] + [f"PD_tom_{v}" for v in variantes] + [f"err%_tom_{v}" for v in variantes]
    print(anual[anual["Grupo"] == "TOTAL"][cols].to_string(index=False))
    print("\n=== Prima devengada RETENIDA por año (USD) · TOTAL ===")
    cols = ["Año", "PR", "PD_ret_real"] + [f"PD_ret_{v}" for v in variantes] + [f"err%_ret_{v}" for v in variantes]
    print(anual[anual["Grupo"] == "TOTAL"][cols].to_string(index=False))

    # ---- archivos ----
    cmp.to_csv(os.path.join(OUT, "comparacion_mensual.csv"), index=False)
    anual.to_csv(os.path.join(OUT, "prima_devengada_anual.csv"), index=False)
    met.to_csv(os.path.join(OUT, "metricas_pnd.csv"), index=False)
    cal.to_csv(os.path.join(OUT, "calibracion_delta.csv"), index=False)
    tabla.to_csv(os.path.join(OUT, "tabla_fnd_calibrada.csv"))
    with open(os.path.join(OUT, "delta_calibrado.json"), "w") as f:
        json.dump(delta, f, indent=2)
    escribir_excel(cmp, anual, met, cal, tabla, variantes, real, is_ram, delta)
    print(f"\nSalidas en {OUT}")
    return cmp, anual, met, cal, tabla


def escribir_excel(cmp, anual, met, cal, tabla, variantes, real, is_ram, delta):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter as GCL
    VERDE, VERDE2, BLANCO, GRIS, CLARO, ROJO = 'FF00573F', 'FF2E7D53', 'FFFFFFFF', 'FFF2F2EE', 'FFE8F1EA', 'FFA6192E'
    AZUL, AZUL_CLARO = 'FF1F4E79', 'FFDDEBF7'        # PND real y PND calibrado: lo que se compara
    AMBAR, AMBAR_CLARO = 'FF7F6000', 'FFFFF2CC'       # celdas con fórmula Excel viva
    ETIQ = {"Año": "Año contable", "PERIODO": "Mes contable\n(AAAAMM)"}
    PCT = {"ratio", "MAPE", "R2", "IS_eff", "g", "mr", "c", "ces", "ratio_CAL_real"}
    wb = Workbook(); wb.remove(wb.active)

    def fmt(col):
        if col.startswith("err%") or col in PCT or col.startswith("k="):
            return '0.00%'
        if col == "delta":
            return '0.000'
        if col in ("PERIODO", "Año"):
            return '0'
        return '#,##0'

    def banda(ws, titulo, sub, n):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
        c = ws.cell(1, 1, titulo); c.font = Font(bold=True, color=BLANCO, size=12); c.fill = PatternFill('solid', fgColor=VERDE)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
        c = ws.cell(2, 1, sub); c.font = Font(italic=True, color=BLANCO, size=9); c.fill = PatternFill('solid', fgColor=VERDE2)
        c.alignment = Alignment('left', 'center', wrap_text=True)
        ws.row_dimensions[2].height = 30

    def celda(ws, i, j, v, col, azul=False, formula=False):
        if isinstance(v, (float, np.floating)) and (np.isnan(v) or np.isinf(v)):
            v = None
        c = ws.cell(i, j, v)
        if formula or (isinstance(v, (int, float, np.integer, np.floating)) and not isinstance(v, bool)):
            c.number_format = fmt(col)
            if not formula and col.startswith("err%") and abs(v) > 0.02:
                c.font = Font(color=ROJO, bold=True)
        if azul:
            c.fill = PatternFill('solid', fgColor=AZUL_CLARO)
        elif formula:
            c.fill = PatternFill('solid', fgColor=AMBAR_CLARO)
        elif i % 2 == 0:
            c.fill = PatternFill('solid', fgColor=GRIS)
        return c

    def cabecera(ws, j, col, azul=False, formula=False, comentario=None):
        c = ws.cell(3, j, ETIQ.get(col, col) + ("\n(fórmula)" if formula else ""))
        c.font = Font(bold=True, color=BLANCO)
        c.fill = PatternFill('solid', fgColor=(AZUL if azul else AMBAR if formula else VERDE))
        c.alignment = Alignment('center', 'center', wrap_text=True)
        if comentario:
            c.comment = Comment(comentario, "validar_prima_devengada")
        return c

    def hoja(nombre, titulo, sub, df, ancho=14, azules=(), comentarios=None):
        ws = wb.create_sheet(nombre)
        n = max(len(df.columns), 4)
        banda(ws, titulo, sub, n)
        for j, col in enumerate(df.columns, 1):
            col = str(col)
            cabecera(ws, j, col, azul=(col in azules), comentario=(comentarios or {}).get(col))
            ws.column_dimensions[GCL(j)].width = ancho
        ws.row_dimensions[3].height = 32
        for i, row in enumerate(df.itertuples(index=False), 4):
            for j, v in enumerate(row, 1):
                col = str(df.columns[j - 1])
                celda(ws, i, j, v, col, azul=(col in azules))
        ws.freeze_panes = 'A4'
        return ws

    # ---------------------------- Fórmulas (texto) ----------------------------
    NT_TXT = ", ".join(f"{x:.4f}" for x in NT_M)
    FORMULAS = [
        ("Año contable", "Año = ENTERO(PERIODO / 100)",
         "PERIODO es el mes CONTABLE: el mes de valuación del saldo de RRC en la base BEL-IRR-MR y el mes en que se "
         "registra la prima (aPog_MesProc / CALMONTH). No es el año de suscripción; ése queda como dato descriptivo (AñoSusc) en el input del MEC."),
        ("Antigüedad de registro k", "k = mes de valuación − mes de registro (CALMONTH), en meses",
         "Es el eje del FND para proporcional y facultativo. k = 0 es el propio mes en que entra la cuenta."),
        ("FND calibrado (TipoRea 1 y 3)", "FND = MIN(1, MAX(0, NT[k] − δ_ramo))  para k = 0…11;  FND = 0 si k < 0 o k ≥ 12",
         f"NT = recta de 24-avos de la Nota Técnica (tabla xPND 'NA'): [{NT_TXT}]. δ_ramo en delta_calibrado.json / hoja Calibracion."),
        ("FND no proporcional (TipoRea 2)", "FND = MIN(1, MAX(0, (FinVig − cierre del mes de valuación) / (FinVig − IniVig)))",
         "Prorrata exacta por fechas. Si el registro no trae fechas (como en la BD del MEC), se usa la curva PF+ de cartera por antigüedad de cohorte."),
        ("Prima en USD", "Prima_USD = PrimasNal / TC(mes de registro)",
         "TC de cierre mensual de la base BEL-IRR-MR (2022 en adelante); 2019–2021 promedios Banxico aproximados."),
        ("PND modelo (por ramo y mes)", "PND_modelo(ramo, t) = Σ_registros Prima_USD × FND(k, ramo)",
         "Suma sobre todos los registros de prima con mes de registro ≤ t. Columna PND_CAL (azul) en la hoja Mensual."),
        ("PND real implícita", "PND_real(ramo, t) = Σ_subramos BEL_subramo(t) / IS_subramo(t)",
         "BEL = 'RRC BEL' de BD_Montos_RRC_SONR; IS = 'Ind Sin RRC' de HParametros (Real). AyE suma 30+34+37; CAT suma 71+73. "
         "Columna PND_real (azul) en Mensual. En esa hoja, BEL_real / IS_eff reproduce PND_real."),
        ("Índice efectivo", "IS_eff = BEL_real / PND_real", "Índice de siniestralidad efectivo del grupo; en grupos de un solo subramo es el 'Ind Sin RRC' tal cual."),
        ("RRC bruta modelo", "BRUTO_modelo = PND_modelo × IS_eff × (1 + g + mr)",
         "g = 'RRC GTO' / 'RRC BEL' real del mes; mr = 'RRC MR' / 'RRC BEL' real del mes. Fórmula viva en la columna BRUTO_CAL."),
        ("IRR y RRC neta modelo", "IRR_modelo = PND_modelo × IS_eff × c ;  NETO_modelo = BRUTO_modelo − IRR_modelo",
         "c = 'RRC IRR' / 'RRC BEL' real del mes. Fórmula viva en NETO_CAL."),
        ("Prima retenida", "PR = PE × (1 − ces)", "ces = prima cedida / prima emitida del ER real, por ramo y año contable (2023, 2024, 2025 YTD oct; 2026 usa 2025)."),
        ("Prima devengada tomada", "PD_tomada(t) = PE(t) − [BRUTO(t) − BRUTO(t − 1)]",
         "Misma fórmula que Integración Dim (ER_2026: prima emitida − variación de reserva de primas s/prima tomada). Real usa BRUTO real; modelo usa BRUTO_CAL."),
        ("Prima devengada retenida", "PD_retenida(t) = PR(t) − [NETO(t) − NETO(t − 1)]", "Real usa NETO real; modelo usa NETO_CAL. La diferencia del mes anterior se toma dentro del mismo ramo."),
        ("Conversión a MXN", "Monto_MXN(t) = Monto_USD(t) × TC de cierre(t)", "Como el ER: la variación de reserva del mes al TC del mes; el efecto cambiario del saldo no pasa por la prima devengada."),
        ("Calibración de δ", "δ_ramo = argmin_δ Σ_t [ PND_modelo(ramo, t; δ) − PND_real(ramo, t) ]²",
         f"Mínimos cuadrados sobre la serie mensual {CAL0}–{T1} (CAT desde {CAL0_GRUPO.get('CAT', CAL0)}). Malla de δ en pasos de 0.005; la parte no proporcional no depende de δ."),
        ("Lectura de δ", "δ = (t − 1) / 2 × 30 / 365   ⇔   t = 1 + 2 δ × 365 / 30 meses de cuenta", "Regla M4 del MEC: δ = 0 es NT mensual; 0.082 trimestral; 0.123 cuatrimestral; 0.205 semestral."),
        ("Métricas de ajuste", "ratio = Σ_t PND_modelo / Σ_t PND_real ;  MAPE = media_t |PND_modelo / PND_real − 1| ;  R² sobre la serie mensual", "Hoja Ajuste_PND. Como IS, g, mr y c son los reales, ratio(RRC modelo / RRC real) = ratio(PND modelo / PND real)."),
        ("Error de prima devengada", "err% = (PD_modelo − PD_real) / PD_real", "Hojas Resumen y PD_anual_por_ramo, por año contable y por ramo. En rojo si |err%| > 2%."),
    ]
    ws = wb.create_sheet("Formulas")
    banda(ws, "FÓRMULAS · CÓMO SE CALCULA CADA COLUMNA", "Orden de cálculo: FND → PND modelo → RRC modelo → prima devengada. Azul = lo que se compara (PND real vs PND calibrado); ámbar = celdas con fórmula Excel viva en la hoja Mensual.", 3)
    for j, (h, w) in enumerate([("Concepto", 30), ("Fórmula", 78), ("Dónde sale y con qué datos", 110)], 1):
        cabecera(ws, j, h); ws.column_dimensions[GCL(j)].width = w
    for i, (a, b, cnota) in enumerate(FORMULAS, 4):
        for j, v in enumerate((a, b, cnota), 1):
            c = ws.cell(i, j, v); c.alignment = Alignment('left', 'top', wrap_text=True)
            if j == 2:
                c.font = Font(name='Consolas', size=9)
            if i % 2 == 0:
                c.fill = PatternFill('solid', fgColor=GRIS)
        ws.row_dimensions[i].height = 48
    ws.freeze_panes = 'A4'

    # ---------------------------- Resumen ----------------------------
    res = []
    tot = anual[anual["Grupo"] == "TOTAL"]
    for v in variantes:
        for lado, lab in [("tom", "Tomada"), ("ret", "Retenida")]:
            for _, r in tot.iterrows():
                res.append(dict(Variante=v, Lado=lab, **{"Año": int(r["Año"])},
                                PD_real=r[f"PD_{lado}_real"], PD_modelo=r[f"PD_{lado}_{v}"],
                                error=r[f"err_{lado}_{v}"], **{"err%": r[f"err%_{lado}_{v}"]},
                                PD_real_MXN=r[f"PD_{lado}_real_MXN"], PD_modelo_MXN=r[f"PD_{lado}_{v}_MXN"]))
    res = pd.DataFrame(res)
    hoja("Resumen", "PRIMA DEVENGADA · REAL vs MODELO · TOTAL CARTERA (USD y MXN, sin fianzas) · por AÑO CONTABLE",
         "Real: prima emitida − ΔRRC bruta (tomada); prima retenida − ΔRRC neta (retenida), saldos base BEL-IRR-MR. "
         "Modelo: misma fórmula con RRC reconstruida = PND_modelo · IS · (1+g+mr) e IRR = BEL·c, con IS, g, mr, c reales. "
         f"Año contable = año del mes de valuación (PERIODO); {T1 // 100} = enero–{['ene','feb','mar','abr','may','jun','jul','ago','sep','oct','nov','dic'][T1 % 100 - 1]}. MXN = Δ mensual × TC de cierre. Variante CAL = FND calibrado.",
         res)

    # ---------------------------- Gráficas ----------------------------
    from openpyxl.chart import LineChart, Reference, Series
    from openpyxl.drawing.line import LineProperties
    tt = cmp[cmp["PERIODO"] >= CAL0].groupby("PERIODO")[["PND_real", "PND_NT_reg_m", "PND_MEC_pub", "PND_CAL", "BRUTO", "BRUTO_CAL", "BRUTO_MEC_pub"]].sum().reset_index()
    tt.insert(1, "Año", tt["PERIODO"] // 100)
    tt.columns = ["PERIODO", "Año", "PND real (BEL/IS)", "PND NT registro mensual", "PND MEC publicado (cohorte vigencia)", "PND calibrado δ",
                  "RRC bruta real", "RRC bruta calibrado δ", "RRC bruta MEC publicado"]
    ws = hoja("Graficas", "SERIES MENSUALES · TOTAL CARTERA (USD)", "Prima no devengada implícita y RRC bruta: real vs variantes. Azul = PND real y PND calibrado (lo que se compara).", tt, ancho=16,
              azules=("PND real (BEL/IS)", "PND calibrado δ"))
    n = len(tt) + 3
    for (c1, c2, titulo, anchor) in [(3, 6, "Prima no devengada: real vs modelo", "L4"), (7, 9, "RRC bruta: real vs modelo", "L26")]:
        ch = LineChart(); ch.title = titulo; ch.height = 10; ch.width = 22
        ch.y_axis.numFmt = '#,##0'; ch.x_axis.delete = False; ch.y_axis.delete = False
        for col, color in zip(range(c1, c2 + 1), ["1F4E79", "C9A961", "A6192E", "00573F", "5B2C6F"]):
            sr = Series(Reference(ws, min_col=col, min_row=3, max_row=n), title_from_data=True)
            sr.graphicalProperties.line = LineProperties(solidFill=color, w=20000); sr.smooth = False
            ch.series.append(sr)
        ch.set_categories(Reference(ws, min_col=1, min_row=4, max_row=n))
        ws.add_chart(ch, anchor)

    # ---------------------------- Anuales y ajuste ----------------------------
    hoja("PD_anual_por_ramo", "PRIMA DEVENGADA POR RAMO Y AÑO CONTABLE · REAL vs VARIANTES (USD)",
         f"err% = (modelo − real) / real. Año contable = año del mes de valuación. {T1 // 100} = enero–{['ene','feb','mar','abr','may','jun','jul','ago','sep','oct','nov','dic'][T1 % 100 - 1]}.", anual, ancho=15)
    hoja("Ajuste_PND", f"AJUSTE DE LA PRIMA NO DEVENGADA (≡ RRC) POR VARIANTE · {CAL0}–{T1}",
         "ratio = Σ PND modelo / Σ PND real · MAPE = error absoluto medio mensual · R² sobre la serie mensual", met)
    hoja("Calibracion", "CALIBRACIÓN δ POR RAMO · FND = max(0, NT(k_registro) − δ) para proporcional/facultativo",
         "δ es el desplazamiento de la regla M4 del MEC (frecuencia de cuentas): δ = (t−1)/2·30/365 → t = 1 + 2δ·365/30 meses equivalentes",
         cal)
    t2 = tabla.reset_index().rename(columns={"index": "Vector"})
    hoja("TablaFND_calibrada", "TABLA FND CALIBRADA · % NO DEVENGADO POR ANTIGÜEDAD DE REGISTRO (k=0 mes de registro)",
         "Se aplica a proporcional y facultativo; el no proporcional sigue con prorrata exacta por fechas de vigencia.", t2, ancho=11)

    # ---------------------------- Mensual (con fórmulas vivas) ----------------------------
    d = cmp.sort_values(["Grupo", "PERIODO"]).reset_index(drop=True)
    otras = [v for v in variantes if v != "CAL"]
    # (columna, 'v' valor / 'f' fórmula, columna de origen en cmp)
    COLS = ([("Grupo", "v", "Grupo"), ("PERIODO", "v", "PERIODO"), ("Año", "v", "Año"), ("PE", "v", "PE"), ("ces", "v", "ces"),
             ("PR", "f", None),
             ("IS_eff", "v", "IS_eff"), ("g", "v", "g"), ("mr", "v", "mr"), ("c", "v", "c"), ("BEL_real", "v", "BEL"),
             ("PND_real", "v", "PND_real"), ("PND_CAL", "v", "PND_CAL"),
             ("ratio_CAL_real", "f", None),
             ("BRUTO_real", "v", "BRUTO"), ("BRUTO_CAL", "f", None),
             ("NETO_real", "v", "NETO"), ("NETO_CAL", "f", None),
             ("dBRUTO_real", "f", None), ("dBRUTO_CAL", "f", None), ("dNETO_real", "f", None), ("dNETO_CAL", "f", None),
             ("PD_tom_real", "f", None), ("PD_tom_CAL", "f", None), ("PD_ret_real", "f", None), ("PD_ret_CAL", "f", None)]
            + [(f"PND_{v}", "v", f"PND_{v}") for v in otras]
            + [(f"PD_tom_{v}", "v", f"PD_tom_{v}") for v in otras]
            + [(f"PD_ret_{v}", "v", f"PD_ret_{v}") for v in otras])
    L = {name: GCL(i + 1) for i, (name, _, _) in enumerate(COLS)}
    G, PEc, CES, ISc, gc, mrc, cc = L["Grupo"], L["PE"], L["ces"], L["IS_eff"], L["g"], L["mr"], L["c"]

    def dif(col):        # Δ contra el mes anterior del MISMO ramo; vacío en el primer mes de cada ramo
        return lambda r, p: f'=IF({G}{r}={G}{p},{L[col]}{r}-{L[col]}{p},"")'

    def pd_(base, dcol):  # prima devengada = base − Δreserva
        return lambda r, p: f'=IF({L[dcol]}{r}="","",{L[base]}{r}-{L[dcol]}{r})'
    FORM = {
        "PR": lambda r, p: f"={PEc}{r}*(1-{CES}{r})",
        "ratio_CAL_real": lambda r, p: f'=IF({L["PND_real"]}{r}=0,"",{L["PND_CAL"]}{r}/{L["PND_real"]}{r})',
        "BRUTO_CAL": lambda r, p: f"={L['PND_CAL']}{r}*{ISc}{r}*(1+{gc}{r}+{mrc}{r})",
        "NETO_CAL": lambda r, p: f"={L['BRUTO_CAL']}{r}-{L['PND_CAL']}{r}*{ISc}{r}*{cc}{r}",
        "dBRUTO_real": dif("BRUTO_real"), "dBRUTO_CAL": dif("BRUTO_CAL"),
        "dNETO_real": dif("NETO_real"), "dNETO_CAL": dif("NETO_CAL"),
        "PD_tom_real": pd_("PE", "dBRUTO_real"), "PD_tom_CAL": pd_("PE", "dBRUTO_CAL"),
        "PD_ret_real": pd_("PR", "dNETO_real"), "PD_ret_CAL": pd_("PR", "dNETO_CAL"),
    }
    COMENT = {
        "PND_real": "PND real implícita = Σ_subramos BEL_subramo / IS_subramo (base BEL-IRR-MR: 'RRC BEL' y 'Ind Sin RRC' de HParametros Real). "
                    "En esta hoja equivale a BEL_real / IS_eff. Es el REAL contra el que se compara el modelo.",
        "PND_CAL": "PND calibrado = Σ_registros Prima_USD × FND, con FND = MIN(1, MAX(0, NT[k] − δ_ramo)), k = mes de valuación − mes de registro "
                   "(proporcional y facultativo); prorrata por cohorte para el no proporcional. Es el MODELO.",
        "IS_eff": "IS_eff = BEL_real / PND_real (índice de siniestralidad efectivo del grupo, real del mes).",
        "g": "g = 'RRC GTO' / 'RRC BEL' real del mes.", "mr": "mr = 'RRC MR' / 'RRC BEL' real del mes.", "c": "c = 'RRC IRR' / 'RRC BEL' real del mes.",
        "ces": "ces = prima cedida / prima emitida del ER real, por ramo y año contable.",
        "BRUTO_CAL": "RRC bruta modelo = PND_CAL × IS_eff × (1 + g + mr).",
        "NETO_CAL": "RRC neta modelo = BRUTO_CAL − PND_CAL × IS_eff × c.",
        "PD_tom_real": "Prima devengada tomada real = PE − (BRUTO_real del mes − BRUTO_real del mes anterior).",
        "PD_tom_CAL": "Prima devengada tomada modelo = PE − (BRUTO_CAL del mes − BRUTO_CAL del mes anterior).",
        "PD_ret_real": "Prima devengada retenida real = PR − (NETO_real del mes − NETO_real del mes anterior).",
        "PD_ret_CAL": "Prima devengada retenida modelo = PR − (NETO_CAL del mes − NETO_CAL del mes anterior).",
        "ratio_CAL_real": "ratio = PND_CAL / PND_real. 1.00 = el modelo reproduce la reserva real del mes.",
        "Año": "Año contable = ENTERO(PERIODO/100). PERIODO es el mes de valuación, no el año de suscripción.",
    }
    ws = wb.create_sheet("Mensual")
    banda(ws, "DETALLE MENSUAL POR RAMO (USD) · AZUL = PND real vs PND calibrado · ÁMBAR = fórmulas Excel vivas",
          "Cada fila es un ramo en un mes contable. Las columnas ámbar son fórmulas que se pueden auditar celda por celda; "
          "las azules son las dos series que se comparan. Los Δ se calculan contra el mes anterior del mismo ramo (vacío en el primer mes). "
          "Las columnas al final (PND_*, PD_*) son las demás variantes, como valores. Ver hoja Formulas.", len(COLS))
    for j, (name, tipo, _) in enumerate(COLS, 1):
        cabecera(ws, j, name, azul=(name in ("PND_real", "PND_CAL")), formula=(tipo == "f"), comentario=COMENT.get(name))
        ws.column_dimensions[GCL(j)].width = 13 if name not in ("Grupo",) else 10
    ws.row_dimensions[3].height = 40
    for i, row in enumerate(d.itertuples(index=False), 4):
        rr = row._asdict()
        for j, (name, tipo, src) in enumerate(COLS, 1):
            if tipo == "f":
                celda(ws, i, j, FORM[name](i, i - 1), name, formula=True)
            else:
                celda(ws, i, j, rr[src], name, azul=(name in ("PND_real", "PND_CAL")))
    ws.freeze_panes = 'D4'

    hoja("Real_parametros", "PARÁMETROS REALES USADOS (por ramo × mes contable)",
         "IS_eff = BEL/PND implícita · g = gastos/BEL · mr = MR/BEL · c = IRR/BEL", real, ancho=13)
    sup = pd.DataFrame({"Supuesto": [
        f"Prima base: BD del MEC (PrimasNal, Tipo Póliza P*), fuente BD, origen Real, hasta {T1}; fianzas (130–170) excluidas por no tener RRC en la base.",
        "Conversión a USD al TC de cierre del mes de registro (base BEL-IRR-MR 2022+; 2019–2021 promedios Banxico aproximados).",
        "Real = saldos BD_Montos_RRC_SONR (USD). AyE = RAM_30+34+37; CAT = RAM_71+73. Índices de siniestralidad: HParametros 'Real' (31/35/39 y 30/34/37 según fecha; TEV/Hidro para CAT desde 2024, antes se rellenan).",
        "Prima no devengada real implícita = Σ BEL_subramo / IS_subramo. IS efectivo, gasto, MR y % cesión se toman del real mes a mes.",
        "Prima retenida = prima emitida × (1 − % cedido anual del ER real por ramo: 2023, 2024, 2025 YTD oct; 2026 usa 2025).",
        "No proporcional (TipoRea 2): la BD del MEC no trae fin de vigencia por registro; se usa la curva PF+ de cartera por antigüedad de cohorte como proxy de la prorrata exacta.",
        "Antigüedad de registro k = mes de valuación − mes de registro (Periodo). El año de los reportes es el CONTABLE (año de PERIODO), no el de suscripción.",
        f"Ventana de calibración y reporte: {CAL0}–{T1} (CAT desde {CAL0_GRUPO.get('CAT', CAL0)}). Vectores MEC: m2_fnd_prorrata sobre Registros_Vigencia_MEC.csv, horizonte {H} meses; abre sólo Vida (decisión M3).",
        "Las columnas ámbar de la hoja Mensual son fórmulas Excel; sus valores se recalculan al abrir el archivo. Si Excel muestra ceros, pulsa F9 o activa el cálculo automático.",
    ]})
    hoja("Supuestos", "SUPUESTOS Y FUENTES", "", sup, ancho=160)
    # orden de hojas: Resumen primero, Formulas después
    wb.move_sheet("Formulas", offset=-(len(wb.sheetnames) - 2))
    ruta = os.path.join(OUT, "Validacion_Prima_Devengada.xlsx")
    wb.save(ruta)
    return ruta

if __name__ == "__main__":
    main()
