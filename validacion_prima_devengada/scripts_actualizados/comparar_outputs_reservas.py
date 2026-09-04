# -*- coding: utf-8 -*-
"""
================================================================================
 comparar_outputs_reservas.py · v1 · Output anterior vs output con FND calibrado
--------------------------------------------------------------------------------
 Planeación Financiera (BP&A) · Reaseguradora Patria (GPV)

 Los dos reforecast de reservas (reforecastRRC_v11_Esc1_ocl.py y ReforecastSONR_v4.py)
 escriben un Excel de saldos con el mismo esquema:

     Reserva · Escenario · Tipo de Monto · Ramo · Periodo · Monto_MXN · [TC] · Monto_USD

 Entre la corrida anterior y la nueva lo ÚNICO que cambió es el factor de no
 devengamiento (FND). Este script pone lado a lado el output anterior («base») y el
 nuevo, cruza por llave y escribe  Comparativo_<etiqueta>.xlsx  con cuatro hojas:

     · Resumen        Tipo de Monto × Escenario: base, nuevo, dif, dif%  (MXN y USD)
     · Por_ramo       Tipo de Monto × Ramo, para todos los escenarios y, aparte,
                      sólo el escenario 2 (lo separa la columna «Alcance»)
     · Por_periodo    Tipo de Monto × Periodo del escenario 2
     · Detalle        el cruce completo por llave, con la columna «presente»
                      ('ambos' / 'solo_base' / 'solo_nuevo')

 Reglas del cruce
     · Llave = (Reserva, Escenario, Tipo de Monto, Ramo, Periodo).
     · Periodo se trata SIEMPRE como texto: en los escenarios 0-3 viene como entero
       AAAAMM y en el escenario 4 como '202512-<mes>'.
     · Si un archivo trae filas repetidas por llave se SUMAN antes de comparar
       (Detalle conserva n_filas_base / n_filas_nuevo para poder verlo).
     · Merge OUTER: lo que sólo está en un lado entra con 0 en el otro.
     · dif = nuevo - base;  dif% = dif / base  (vacío cuando base = 0).
     · En los agregados dif% se recalcula sobre los totales; no se promedian %.
     · Formato: encabezado verde Patria, montos '#,##0', porcentajes '0.00%' y
       |dif%| > 2% en rojo negrita; primera fila congelada y con autofiltro.

 Uso (los archivos se buscan en la carpeta del propio script y ahí se escribe la salida):
     python comparar_outputs_reservas.py
         RRC_esc.xlsx  vs RRC_esc_FNDcal.xlsx   -> Comparativo_RRC.xlsx
         SONR_esc.xlsx vs SONR_esc_FNDcal.xlsx  -> Comparativo_SONR.xlsx
         (compara cada par que esté completo y avisa del que falte)
     python comparar_outputs_reservas.py <base.xlsx> <nuevo.xlsx> [etiqueta]
     python comparar_outputs_reservas.py --demo
         arma un par sintético en _demo_comparador/, corre la comparación sobre él
         y verifica con asserts que el comparador cuadra.

 Sólo requiere pandas, numpy y openpyxl. No modifica los outputs que compara.
================================================================================
"""
import math
import os
import re
import sys

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CARPETA = os.path.dirname(os.path.abspath(__file__))

# Pares que se buscan cuando el script corre sin argumentos: (etiqueta, base, nuevo)
PARES_DEFAULT = [
    ("RRC",  "RRC_esc.xlsx",  "RRC_esc_FNDcal.xlsx"),
    ("SONR", "SONR_esc.xlsx", "SONR_esc_FNDcal.xlsx"),
]

# Esquema del output de los reforecast
COL_RESERVA, COL_ESC, COL_TIPO, COL_RAMO, COL_PER = "Reserva", "Escenario", "Tipo de Monto", "Ramo", "Periodo"
COL_MXN, COL_USD, COL_TC = "Monto_MXN", "Monto_USD", "TC"
LLAVE = [COL_RESERVA, COL_ESC, COL_TIPO, COL_RAMO, COL_PER]
MONTOS = [COL_MXN, COL_USD]
MONEDAS = ["MXN", "USD"]

ORDEN_TIPOS = ["BEL", "BELG", "IRR", "MR", "BRUTO", "NETO"]   # orden de presentación
ESC_FOCO = 2                    # escenario que se abre por ramo y por periodo
UMBRAL_ALERTA = 0.02            # |dif%| por encima de esto se pinta en rojo

# Formato Excel
COLOR_ENCABEZADO = "FF00573F"   # verde oscuro Patria
COLOR_ALERTA = "FFA6192E"       # rojo
FMT_NUM = "#,##0"
FMT_PCT = "0.00%"
FMT_TC = "0.0000"

TAG = "[comparador]"


# ============================ NORMALIZACIÓN DE LLAVE ==========================
def _es_nulo(v):
    if v is None or v is pd.NaT:
        return True
    try:
        return bool(pd.isna(v))
    except (TypeError, ValueError):
        return False


def _numero(v):
    """float si v es un número o un texto numérico; None en cualquier otro caso."""
    if isinstance(v, (bool, np.bool_)):
        return None
    if isinstance(v, (int, float, np.integer, np.floating)):
        f = float(v)
        return None if math.isnan(f) else f
    if isinstance(v, str):
        try:
            return float(v.strip().replace(",", ""))
        except ValueError:
            return None
    return None


def _entero_o_texto(v):
    """Escenario y Ramo: entero cuando el valor es numérico entero (10, 10.0, '10');
    en cualquier otro caso texto sin espacios. Así 10 y 10.0 caen en la misma llave."""
    if _es_nulo(v):
        return "NA"
    f = _numero(v)
    if f is not None and f.is_integer():
        return int(f)
    return str(v).strip()


def _periodo_texto(v):
    """Periodo SIEMPRE como texto: 202501 -> '202501', 202501.0 -> '202501',
    '202512-9' -> '202512-9'. Es la única forma de cruzar los escenarios 0-3 y el 4."""
    if _es_nulo(v):
        return "NA"
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    if isinstance(v, (float, np.floating)):
        return str(int(v)) if float(v).is_integer() else str(v)
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _texto(v):
    return "NA" if _es_nulo(v) else str(v).strip()


def _clave_orden(col, v):
    """Clave de texto para ordenar columnas que mezclan enteros y textos sin que
    pandas truene: primero los numéricos (por valor), luego los textos."""
    if col == COL_TIPO:
        s = str(v)
        return f"0{ORDEN_TIPOS.index(s):02d}" if s in ORDEN_TIPOS else f"1{s}"
    f = _numero(v)
    return f"0{f:020.4f}" if f is not None else f"1{v}"


def ordenar(df, cols):
    if df.empty:
        return df.reset_index(drop=True)
    aux = df.copy()
    claves = []
    for i, c in enumerate(cols):
        k = f"__orden_{i}"
        aux[k] = aux[c].map(lambda v, c=c: _clave_orden(c, v))
        claves.append(k)
    return aux.sort_values(claves, kind="mergesort").drop(columns=claves).reset_index(drop=True)


# ================================ LECTURA =====================================
def leer_output(ruta):
    """Lee un output de reservas (primera fila = encabezado), normaliza la llave y
    suma las filas repetidas por llave. Devuelve (df por llave, info de conteos)."""
    df = pd.read_excel(ruta, header=0)
    df.columns = [str(c).strip() for c in df.columns]

    # tolera diferencias de mayúsculas/minúsculas en los nombres de columna
    por_minuscula = {c.lower(): c for c in df.columns}
    renombres = {}
    for esperada in LLAVE + MONTOS + [COL_TC]:
        real = por_minuscula.get(esperada.lower())
        if real is not None and real != esperada:
            renombres[real] = esperada
    df = df.rename(columns=renombres)

    faltan = [c for c in LLAVE + MONTOS if c not in df.columns]
    if faltan:
        raise SystemExit(f"{TAG} A «{os.path.basename(ruta)}» le faltan columnas: {faltan}\n"
                         f"  Columnas encontradas: {list(df.columns)}")

    n_filas = len(df)
    df[COL_RESERVA] = df[COL_RESERVA].map(_texto).str.upper()
    df[COL_TIPO] = df[COL_TIPO].map(_texto).str.upper()
    df[COL_ESC] = df[COL_ESC].map(_entero_o_texto)
    df[COL_RAMO] = df[COL_RAMO].map(_entero_o_texto)
    df[COL_PER] = df[COL_PER].map(_periodo_texto)
    for c in MONTOS:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0).astype(float)

    agregacion = {COL_MXN: "sum", COL_USD: "sum"}
    if COL_TC in df.columns:
        df[COL_TC] = pd.to_numeric(df[COL_TC], errors="coerce")
        agregacion[COL_TC] = "mean"          # el TC no se suma; se conserva como referencia
    df["n_filas"] = 1
    agregacion["n_filas"] = "sum"

    # sort=False: las llaves mezclan enteros y textos y pandas no podría ordenarlas
    g = df.groupby(LLAVE, as_index=False, sort=False, dropna=False).agg(agregacion)
    info = {"filas": n_filas, "llaves": len(g), "duplicadas": int((g["n_filas"] > 1).sum())}
    return g, info


# ============================== COMPARACIÓN ===================================
def _agregar_difs(df, columnas):
    """columnas = {'MXN': (col_base, col_nuevo), 'USD': (...)}. Añade dif_<m> y
    dif%_<m>; dif% queda NaN cuando la base es 0."""
    for m, (cb, cn) in columnas.items():
        base = df[cb].to_numpy(dtype=float)
        nuevo = df[cn].to_numpy(dtype=float)
        dif = nuevo - base
        with np.errstate(divide="ignore", invalid="ignore"):
            pct = np.where(base != 0, dif / np.where(base != 0, base, 1.0), np.nan)
        df[f"dif_{m}"] = dif
        df[f"dif%_{m}"] = pct
    return df


def comparar(base, nuevo):
    """Merge OUTER por llave; faltantes a 0; columna 'presente'; diferencias."""
    b = base.rename(columns={c: f"{c}_base" for c in base.columns if c not in LLAVE})
    n = nuevo.rename(columns={c: f"{c}_nuevo" for c in nuevo.columns if c not in LLAVE})
    det = b.merge(n, on=LLAVE, how="outer", indicator=True)
    det["presente"] = det["_merge"].map({"both": "ambos", "left_only": "solo_base",
                                         "right_only": "solo_nuevo"}).astype(str)
    det = det.drop(columns="_merge")

    for m in MONTOS:
        for lado in ("base", "nuevo"):
            det[f"{m}_{lado}"] = det[f"{m}_{lado}"].fillna(0.0).astype(float)
    for lado in ("base", "nuevo"):
        det[f"n_filas_{lado}"] = det[f"n_filas_{lado}"].fillna(0).astype(int)

    det = _agregar_difs(det, {"MXN": (f"{COL_MXN}_base", f"{COL_MXN}_nuevo"),
                              "USD": (f"{COL_USD}_base", f"{COL_USD}_nuevo")})

    orden = LLAVE + ["presente"]
    for mon, col in (("MXN", COL_MXN), ("USD", COL_USD)):
        orden += [f"{col}_base", f"{col}_nuevo", f"dif_{mon}", f"dif%_{mon}"]
    orden += [c for c in (f"{COL_TC}_base", f"{COL_TC}_nuevo") if c in det.columns]
    orden += ["n_filas_base", "n_filas_nuevo"]
    return ordenar(det[orden], LLAVE)


def resumir(det, por):
    """Agrega el detalle por las columnas `por` y recalcula dif y dif% sobre los totales."""
    cols = {f"{COL_MXN}_base": "base_MXN", f"{COL_MXN}_nuevo": "nuevo_MXN",
            f"{COL_USD}_base": "base_USD", f"{COL_USD}_nuevo": "nuevo_USD"}
    r = det.groupby(por, as_index=False, sort=False, dropna=False)[list(cols)].sum()
    r = r.rename(columns=cols)
    r = _agregar_difs(r, {"MXN": ("base_MXN", "nuevo_MXN"), "USD": ("base_USD", "nuevo_USD")})
    salida = por + ["base_MXN", "nuevo_MXN", "dif_MXN", "dif%_MXN",
                    "base_USD", "nuevo_USD", "dif_USD", "dif%_USD"]
    return ordenar(r[salida], por)


# ================================ EXCEL =======================================
def _py(v):
    """Valor nativo de Python para openpyxl; NaN/NaT -> celda vacía."""
    if v is None or v is pd.NaT:
        return None
    if isinstance(v, np.bool_):
        return bool(v)
    if isinstance(v, np.integer):
        return int(v)
    if isinstance(v, (float, np.floating)):
        f = float(v)
        return None if math.isnan(f) else f
    return v


def _formato_columna(col):
    if col.startswith("dif%"):
        return FMT_PCT
    if col == COL_TC or col.startswith(f"{COL_TC}_"):
        return FMT_TC
    if "MXN" in col or "USD" in col:
        return FMT_NUM
    if col.startswith("n_filas"):
        return "0"
    return None


def _ancho_columna(col, valores):
    fmt = _formato_columna(col)
    n = len(str(col))
    for v in valores[:3000]:
        if v is None:
            continue
        if isinstance(v, (int, float)) and fmt == FMT_NUM:
            s = f"{v:,.0f}"
        elif isinstance(v, (int, float)) and fmt == FMT_PCT:
            s = f"{v:.2%}"
        elif isinstance(v, (int, float)) and fmt == FMT_TC:
            s = f"{v:.4f}"
        else:
            s = str(v)
        n = max(n, len(s))
    return min(max(n + 2, 9), 40)


def escribir_excel(ruta, hojas):
    """hojas = [(nombre, DataFrame, nota_si_vacia), ...]. Escribe con openpyxl:
    encabezado verde con letra blanca, '#,##0' en montos, '0.00%' en porcentajes,
    |dif%| > UMBRAL_ALERTA en rojo negrita, primera fila congelada y autofiltro."""
    wb = Workbook()
    wb.remove(wb.active)
    relleno = PatternFill(fill_type="solid", start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO)
    fuente_enc = Font(bold=True, color="FFFFFFFF")
    fuente_alerta = Font(bold=True, color=COLOR_ALERTA)
    alin_enc = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for nombre, df, nota in hojas:
        ws = wb.create_sheet(nombre)
        cols = [str(c) for c in df.columns]
        ws.append(cols)
        for celda in ws[1]:
            celda.fill = relleno
            celda.font = fuente_enc
            celda.alignment = alin_enc

        filas = [[_py(v) for v in fila] for fila in df.itertuples(index=False, name=None)]
        if not filas:
            ws.cell(row=2, column=1, value=nota)
        for fila in filas:
            ws.append(fila)

        for j, col in enumerate(cols, start=1):
            fmt = _formato_columna(col)
            es_pct = fmt == FMT_PCT
            if fmt is not None and filas:
                for i in range(2, len(filas) + 2):
                    celda = ws.cell(row=i, column=j)
                    celda.number_format = fmt
                    if es_pct and isinstance(celda.value, (int, float)) \
                            and abs(celda.value) > UMBRAL_ALERTA:
                        celda.font = fuente_alerta
            ws.column_dimensions[get_column_letter(j)].width = \
                _ancho_columna(col, [f[j - 1] for f in filas])

        ws.freeze_panes = "A2"
        if filas:
            ws.auto_filter.ref = ws.dimensions
    wb.save(ruta)


# ================================ CONSOLA =====================================
def _pct_txt(v):
    return "n/a" if v is None or (isinstance(v, float) and math.isnan(v)) else f"{v:+.2%}"


def imprimir_resumen(resumen, etiqueta):
    """Resumen en consola: BRUTO y NETO por escenario con separadores de miles."""
    sel = resumen[resumen[COL_TIPO].isin(["BRUTO", "NETO"])]
    if sel.empty:
        sel = resumen
        print(f"\n{TAG} {etiqueta}: no hay filas BRUTO/NETO; se muestran todos los tipos de monto")
    print(f"\n{TAG} Resumen {etiqueta} por escenario (dif = nuevo - base)")
    enc = (f"{'Tipo':<7}{'Esc':>4} | {'base MXN':>18}{'nuevo MXN':>18}{'dif MXN':>16}{'dif%':>9}"
           f" | {'base USD':>16}{'nuevo USD':>16}{'dif USD':>14}{'dif%':>9}")
    print(enc)
    print("-" * len(enc))
    for _, r in sel.iterrows():
        print(f"{str(r[COL_TIPO]):<7}{str(r[COL_ESC]):>4} | "
              f"{r['base_MXN']:>18,.0f}{r['nuevo_MXN']:>18,.0f}{r['dif_MXN']:>16,.0f}"
              f"{_pct_txt(r['dif%_MXN']):>9} | "
              f"{r['base_USD']:>16,.0f}{r['nuevo_USD']:>16,.0f}{r['dif_USD']:>14,.0f}"
              f"{_pct_txt(r['dif%_USD']):>9}")


def _imprimir_hallazgos(det):
    conteo = det["presente"].value_counts()
    print(f"\n{TAG} Llaves cruzadas: {len(det):,} | ambos {conteo.get('ambos', 0):,} | "
          f"solo_base {conteo.get('solo_base', 0):,} | solo_nuevo {conteo.get('solo_nuevo', 0):,}")
    alertas = int((det["dif%_MXN"].abs() > UMBRAL_ALERTA).sum())
    print(f"{TAG} Llaves con |dif%| > {UMBRAL_ALERTA:.0%} en MXN: {alertas:,}")
    top = det[det["presente"] == "ambos"].copy()
    if not top.empty:
        top["_abs"] = top["dif_MXN"].abs()
        top = top.nlargest(5, "_abs")
        print(f"{TAG} Mayores diferencias en MXN (llaves presentes en ambos):")
        for _, r in top.iterrows():
            print(f"   Esc {str(r[COL_ESC]):<3} {str(r[COL_TIPO]):<6} Ramo {str(r[COL_RAMO]):<5} "
                  f"{str(r[COL_PER]):<10} dif {r['dif_MXN']:>16,.0f}  ({_pct_txt(r['dif%_MXN'])})")


# ============================== ORQUESTACIÓN ==================================
def _etiqueta_limpia(etiqueta):
    e = re.sub(r"[^\w.\-]+", "_", str(etiqueta).strip())
    return e.strip("_") or "comparativo"


def _etiqueta_desde(ruta_base):
    stem = os.path.splitext(os.path.basename(ruta_base))[0]
    for pref in ("SONR", "RRC"):
        if stem.upper().startswith(pref):
            return pref
    return _etiqueta_limpia(stem)


def ejecutar_comparacion(ruta_base, ruta_nuevo, etiqueta, carpeta_salida=None):
    """Corre la comparación completa y escribe Comparativo_<etiqueta>.xlsx.
    Devuelve un dict con los DataFrames de cada hoja y la ruta del archivo."""
    carpeta_salida = carpeta_salida or CARPETA
    etiqueta = _etiqueta_limpia(etiqueta)
    print(f"\n{TAG} ===== {etiqueta} =====")

    base, ib = leer_output(ruta_base)
    nuevo, inuevo = leer_output(ruta_nuevo)
    for nombre, ruta, info in (("base ", ruta_base, ib), ("nuevo", ruta_nuevo, inuevo)):
        aviso = f", {info['duplicadas']:,} llaves con filas repetidas (sumadas)" if info["duplicadas"] else ""
        print(f"{TAG} {nombre}: {ruta}  ({info['filas']:,} filas, {info['llaves']:,} llaves{aviso})")

    det = comparar(base, nuevo)
    resumen = resumir(det, [COL_TIPO, COL_ESC])

    det_foco = det[det[COL_ESC] == ESC_FOCO]
    por_ramo_todos = resumir(det, [COL_TIPO, COL_RAMO])
    por_ramo_todos.insert(0, "Alcance", "Todos los escenarios")
    por_ramo_foco = resumir(det_foco, [COL_TIPO, COL_RAMO])
    por_ramo_foco.insert(0, "Alcance", f"Sólo escenario {ESC_FOCO}")
    por_ramo = pd.concat([por_ramo_todos, por_ramo_foco], ignore_index=True)
    por_periodo = resumir(det_foco, [COL_TIPO, COL_PER])
    if det_foco.empty:
        print(f"{TAG} Aviso: ningún archivo trae filas del escenario {ESC_FOCO}; "
              f"Por_periodo y el bloque 'Sólo escenario {ESC_FOCO}' quedan vacíos")

    ruta_salida = os.path.join(carpeta_salida, f"Comparativo_{etiqueta}.xlsx")
    escribir_excel(ruta_salida, [
        ("Resumen", resumen, "Sin datos"),
        ("Por_ramo", por_ramo, "Sin datos"),
        ("Por_periodo", por_periodo, f"Sin filas con Escenario = {ESC_FOCO}"),
        ("Detalle", det, "Sin datos"),
    ])

    imprimir_resumen(resumen, etiqueta)
    _imprimir_hallazgos(det)
    print(f"\n{TAG} Comparativo escrito en: {ruta_salida}")
    return {"Resumen": resumen, "Por_ramo": por_ramo, "Por_periodo": por_periodo,
            "Detalle": det, "ruta": ruta_salida}


# ================================ DEMO ========================================
def _fila_demo(esc, tipo, ramo, periodo, usd, tc):
    t = tc[str(periodo)]
    return ["RRC", esc, tipo, ramo, periodo, round(usd * t, 2), t, round(usd, 2)]


def crear_demo(carpeta):
    """Par sintético con esquema RRC: escenarios 2 y 4, cinco tipos de monto, tres
    ramos, tres periodos (más el '202512-9' del escenario 4), una fila sólo en base,
    una sólo en nuevo, una llave repetida en ambos y una base en 0 (dif% vacío)."""
    os.makedirs(carpeta, exist_ok=True)
    tipos = ["BRUTO", "NETO", "BEL", "IRR", "MR"]
    ramos = [10, 31, 40]
    periodos = [(2, 202510), (2, 202511), (2, 202512), (4, "202512-9")]
    tc = {"202510": 18.40, "202511": 18.55, "202512": 18.70, "202512-9": 18.70}
    escala = {"BRUTO": 1.00, "NETO": 0.80, "BEL": 0.95, "IRR": 0.03, "MR": 0.12}
    # efecto del FND nuevo por ramo: sube 3.5% en el 10, baja 2.5% en el 31, sube 1.2% en el 40
    factor_ramo = {10: 1.035, 31: 0.975, 40: 1.012}
    factor = {t: dict(factor_ramo) for t in ("BRUTO", "NETO", "BEL")}
    factor["IRR"] = {10: 1.0, 31: 1.0, 40: 1.0}
    factor["MR"] = {10: 1.05, 31: 1.0, 40: 1.0}

    filas_base, filas_nuevo = [], []
    for tipo in tipos:
        for ramo in ramos:
            for k, (esc, per) in enumerate(periodos):
                usd = 1_000_000 * escala[tipo] * (ramo / 10) * (1 + 0.02 * k)
                if tipo == "MR" and ramo == 31:
                    usd_b, usd_n = 0.0, 1500.0          # base en 0 -> dif% vacío
                else:
                    usd_b, usd_n = usd, usd * factor[tipo][ramo]
                filas_base.append(_fila_demo(esc, tipo, ramo, per, usd_b, tc))
                filas_nuevo.append(_fila_demo(esc, tipo, ramo, per, usd_n, tc))
    # llave repetida (en los dos archivos): debe sumarse antes de comparar
    filas_base.append(_fila_demo(2, "BRUTO", 10, 202510, 250_000, tc))
    filas_nuevo.append(_fila_demo(2, "BRUTO", 10, 202510, 260_000, tc))
    # una fila sólo en base y una sólo en nuevo
    filas_base.append(_fila_demo(2, "BRUTO", 50, 202512, 80_000, tc))
    filas_nuevo.append(_fila_demo(4, "NETO", 60, "202512-9", 45_000, tc))

    cols = [COL_RESERVA, COL_ESC, COL_TIPO, COL_RAMO, COL_PER, COL_MXN, COL_TC, COL_USD]
    ruta_base = os.path.join(carpeta, "RRC_esc.xlsx")
    ruta_nuevo = os.path.join(carpeta, "RRC_esc_FNDcal.xlsx")
    pd.DataFrame(filas_base, columns=cols).to_excel(ruta_base, index=False)
    pd.DataFrame(filas_nuevo, columns=cols).to_excel(ruta_nuevo, index=False)
    return ruta_base, ruta_nuevo


def correr_demo():
    carpeta = os.path.join(CARPETA, "_demo_comparador")
    print(f"{TAG} Modo demo: par sintético en {carpeta}")
    ruta_base, ruta_nuevo = crear_demo(carpeta)
    out = ejecutar_comparacion(ruta_base, ruta_nuevo, "DEMO_RRC", carpeta_salida=carpeta)

    # 1) el total BRUTO del Resumen coincide con la suma directa de cada archivo
    res = out["Resumen"]
    for ruta, lado in ((ruta_base, "base"), (ruta_nuevo, "nuevo")):
        directo = pd.read_excel(ruta)
        directo = directo[directo[COL_TIPO] == "BRUTO"]
        for mon in MONEDAS:
            esperado = float(directo[f"Monto_{mon}"].sum())
            obtenido = float(res.loc[res[COL_TIPO] == "BRUTO", f"{lado}_{mon}"].sum())
            assert math.isclose(esperado, obtenido, rel_tol=1e-9, abs_tol=1e-6), \
                f"Total BRUTO {lado} {mon}: Resumen {obtenido:,.2f} vs archivo {esperado:,.2f}"
    print(f"{TAG} [demo] OK  total BRUTO del Resumen = suma directa de los archivos (MXN y USD)")

    # 2) las filas solo_base / solo_nuevo aparecen en la hoja Detalle del Excel escrito
    det = pd.read_excel(out["ruta"], sheet_name="Detalle")
    det[COL_PER] = det[COL_PER].map(_periodo_texto)
    assert set(det["presente"]) == {"ambos", "solo_base", "solo_nuevo"}, set(det["presente"])
    sb = det[det["presente"] == "solo_base"]
    sn = det[det["presente"] == "solo_nuevo"]
    assert len(sb) == 1 and int(sb[COL_RAMO].iloc[0]) == 50 and sb[COL_TIPO].iloc[0] == "BRUTO" \
        and sb[COL_PER].iloc[0] == "202512", sb
    assert float(sb[f"{COL_MXN}_nuevo"].iloc[0]) == 0.0 and float(sb["dif_MXN"].iloc[0]) < 0
    assert len(sn) == 1 and int(sn[COL_RAMO].iloc[0]) == 60 and sn[COL_TIPO].iloc[0] == "NETO" \
        and sn[COL_PER].iloc[0] == "202512-9", sn
    assert float(sn[f"{COL_MXN}_base"].iloc[0]) == 0.0 and pd.isna(sn["dif%_MXN"].iloc[0])
    print(f"{TAG} [demo] OK  'solo_base' (BRUTO ramo 50 202512) y 'solo_nuevo' (NETO ramo 60 202512-9) en Detalle")

    # 3) la llave repetida se sumó (1,000,000 + 250,000 en base; 1,035,000 + 260,000 en nuevo)
    dup = det[(det[COL_ESC] == 2) & (det[COL_TIPO] == "BRUTO") & (det[COL_RAMO] == 10) & (det[COL_PER] == "202510")]
    assert len(dup) == 1, f"la llave repetida debe quedar en una sola fila: {len(dup)}"
    assert int(dup["n_filas_base"].iloc[0]) == 2 and int(dup["n_filas_nuevo"].iloc[0]) == 2
    assert math.isclose(float(dup[f"{COL_USD}_base"].iloc[0]), 1_250_000.0, abs_tol=0.01)
    assert math.isclose(float(dup[f"{COL_USD}_nuevo"].iloc[0]), 1_295_000.0, abs_tol=0.01)
    print(f"{TAG} [demo] OK  llave repetida sumada antes de comparar")

    # 4) base en 0 -> dif% vacío; y escenario 4 con periodo texto cruza bien
    mr31 = det[(det[COL_TIPO] == "MR") & (det[COL_RAMO] == 31)]
    assert len(mr31) == 4 and mr31["dif%_MXN"].isna().all() and (mr31["dif_MXN"] > 0).all()
    esc4 = det[det[COL_ESC] == 4]
    assert (esc4[COL_PER] == "202512-9").all() and (esc4["presente"] == "ambos").sum() == 15
    print(f"{TAG} [demo] OK  dif% vacío con base 0 y escenario 4 cruzado por periodo texto")

    # 5) las cuatro hojas existen y Por_periodo sólo trae el escenario 2
    hojas = pd.ExcelFile(out["ruta"]).sheet_names
    assert hojas == ["Resumen", "Por_ramo", "Por_periodo", "Detalle"], hojas
    pp = out["Por_periodo"]
    assert set(pp[COL_PER]) == {"202510", "202511", "202512"}, set(pp[COL_PER])
    assert set(out["Por_ramo"]["Alcance"]) == {"Todos los escenarios", f"Sólo escenario {ESC_FOCO}"}
    print(f"{TAG} [demo] OK  hojas Resumen / Por_ramo / Por_periodo / Detalle")
    print(f"\n{TAG} [demo] Todas las comprobaciones pasaron. Salida: {out['ruta']}")
    return 0


# ================================= MAIN =======================================
def _uso():
    print("Uso:\n"
          "  python comparar_outputs_reservas.py                       (busca los pares RRC y SONR en la carpeta del script)\n"
          "  python comparar_outputs_reservas.py <base.xlsx> <nuevo.xlsx> [etiqueta]\n"
          "  python comparar_outputs_reservas.py --demo                (par sintético + comprobaciones)")


def main(argv):
    args = argv[1:]
    if args and args[0] in ("-h", "--help", "/?"):
        print(__doc__)
        return 0
    if args and args[0] == "--demo":
        return correr_demo()

    if not args:
        hechos = 0
        for etiqueta, nombre_base, nombre_nuevo in PARES_DEFAULT:
            rb = os.path.join(CARPETA, nombre_base)
            rn = os.path.join(CARPETA, nombre_nuevo)
            faltan = [n for n, r in ((nombre_base, rb), (nombre_nuevo, rn)) if not os.path.isfile(r)]
            if faltan:
                print(f"{TAG} {etiqueta}: no se compara, falta {' y '.join(faltan)} en {CARPETA}")
                continue
            ejecutar_comparacion(rb, rn, etiqueta)
            hechos += 1
        if hechos == 0:
            print(f"{TAG} No hay ningún par completo en {CARPETA}.")
            _uso()
            return 1
        return 0

    if len(args) in (2, 3):
        rb, rn = os.path.abspath(args[0]), os.path.abspath(args[1])
        for r in (rb, rn):
            if not os.path.isfile(r):
                raise SystemExit(f"{TAG} No existe el archivo: {r}")
        etiqueta = args[2] if len(args) == 3 else _etiqueta_desde(rb)
        ejecutar_comparacion(rb, rn, etiqueta)
        return 0

    _uso()
    return 2


if __name__ == "__main__":
    sys.exit(main(sys.argv))
