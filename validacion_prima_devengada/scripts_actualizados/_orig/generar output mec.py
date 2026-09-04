# -*- coding: utf-8 -*-
"""
================================================================================
 generar_output_mec.py · v2 · Del archivo de vigencias al OUTPUT del FND
--------------------------------------------------------------------------------
 Planeación Financiera (BP&A) · Reaseguradora Patria (GPV)

 Lee Registros_Vigencia_MEC.csv (lo produce construir_input_mec.py), corre el
 framework modular (mec_devengamiento v2) y escribe:

     Output_MEC_Devengamiento.xlsx
       · Resumen                 alcance, parámetros y hallazgo que cambió la base
       · Comparativo (mes)       NT vs GS vs PF vs PF+ por antigüedad 0-11
       · Escenarios frecuencia   mensual · trimestral · cuatrimestral · semestral
       · PF+ por ramo            vector por ramo y decisión de apertura
       · Diagnóstico             por qué el triángulo de registro no es el FND

 SOLO devengamiento: FS/BELMEDIA, 1-LAG, MR, IRR e índice NO se tocan.
 Los scripts van juntos en la misma carpeta; se anclan a ella solos.
================================================================================
"""
import os, sys, glob, importlib.util, datetime as dt
import numpy as np
import pandas as pd

BASE = os.path.dirname(os.path.abspath(__file__))


def _cargar_mec():
    """Importa mec_devengamiento desde la carpeta del script. Tolera que el archivo
    se haya guardado como «mec devengamiento.py» (con espacio) al descargarlo."""
    if BASE not in sys.path:
        sys.path.insert(0, BASE)
    try:
        import mec_devengamiento as m
        return m
    except ModuleNotFoundError:
        pass
    cand = glob.glob(os.path.join(BASE, "mec*devengamiento*.py"))
    if not cand:
        raise SystemExit("[output] No encuentro el módulo del MEC en:\n"
                         f"  {BASE}\nDebe llamarse  mec_devengamiento.py  (con guion bajo).")
    spec = importlib.util.spec_from_file_location("mec_devengamiento", cand[0])
    m = importlib.util.module_from_spec(spec)
    sys.modules["mec_devengamiento"] = m
    spec.loader.exec_module(m)
    if os.path.basename(cand[0]) != "mec_devengamiento.py":
        print(f"[output] Aviso: cargué el módulo desde «{os.path.basename(cand[0])}». "
              f"Renómbralo a «mec_devengamiento.py» (los reforecast lo importan así).")
    return m


mec = _cargar_mec()
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.drawing.line import LineProperties
from openpyxl.utils import get_column_letter as GCL

# =============================== CONFIG =====================================
RUTA_INSUMOS = BASE          # aquí debe estar Registros_Vigencia_MEC.csv
RUTA_SALIDA = BASE

CFG = mec.ConfigMEC()
CFG.HORIZONTE = 24           # vector FND publicado
CFG.MARGEN_APERTURA = 0.10   # reducción relativa mínima del error para abrir el ramo

# --- Vectores oficiales de referencia (antigüedad 0..11, cierre = diciembre) ---
MESDIC = ['Dic','Nov','Oct','Sep','Ago','Jul','Jun','May','Abr','Mar','Feb','Ene']
NT_MENS = [0.95890411,0.873972603,0.791780822,0.706849315,0.624657534,0.539726027,
           0.454794521,0.37260274,0.287671233,0.205479452,0.120547945,0.043835616]
PF_MENS = [0.95890411,0.876712329,0.791780822,0.706849315,0.624657534,0.539726027,
           0.457534247,0.37260274,0.295890411,0.210958904,0.126027397,0.043835616]
GS_CART = [0.95833333,0.875,0.79166667,0.70833333,0.625,0.54166667,
           0.45833333,0.375,0.29166667,0.20833333,0.125,0.04166667]
GS_MYT  = [0.86635985,0.84143722,0.84538675,0.77002967,0.64978876,0.56176311,
           0.46060141,0.35943971,0.39117551,0.25590124,0.16905497,0.05844632]
GS_AGRO = [0.97373083,0.84776307,0.59944605,0.49999983,0.5,0.09093866,
           0.09093866,0.0,0.0,0.0,0.0,0.0]
FRECS = [("Mensual",1),("Trimestral",3),("Cuatrimestral",4),("Semestral",6),("Anual",12)]

NOMBRE_RAMO = {10:'Vida',30:'Accidentes y Enfermedades',40:'Automóviles',50:'Diversos',
               60:'Incendio',70:'Terremoto / Cat',80:'Agrícola',90:'Marítimo y Transportes',
               100:'Crédito',110:'Responsabilidad Civil',130:'Fianzas Fidelidad',
               140:'Fianzas Judicial',150:'Fianzas Administrativa',160:'Fianzas Crédito',
               170:'Crédito a la Vivienda'}

VERDE='FF00573F'; VERDE2='FF2E7D53'; CLARO='FFE8F1EA'; GRIS='FFF2F2EE'
ROJO='FFA6192E'; BLANCO='FFFFFFFF'; ORO='FFC9A961'; AZUL='FF1F4E79'; MORADO='FF5B2C6F'
CREMA='FFFDF6E3'
_t = Side(style='thin', color='FFDDDDDD'); BORD = Border(left=_t, right=_t, top=_t, bottom=_t)


def C(ws, r, c, v, bold=False, color='FF222222', fill=None, center=True,
      nf=None, size=9, border=False, italic=False, mono=False):
    x = ws.cell(r, c, v)
    x.font = Font(bold=bold, color=color, size=size, italic=italic,
                  name='Consolas' if mono else 'Calibri')
    if fill: x.fill = PatternFill('solid', fgColor=fill)
    x.alignment = Alignment('center' if center else 'left', 'center', wrap_text=not center)
    if nf: x.number_format = nf
    if border: x.border = BORD
    return x


def banda(ws, titulo, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    C(ws, 1, 1, titulo, bold=True, color=BLANCO, fill=VERDE, size=12)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    C(ws, 2, 1, sub, color=BLANCO, fill=VERDE2, size=9, italic=True)


def nota(ws, r, ncols, txt, fill=CREMA, color='FF222222', h=34):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    C(ws, r, 1, txt, color=color, fill=fill, size=8.5, center=False, italic=True)
    ws.row_dimensions[r].height = h


def _guardar_seguro(fn, ruta):
    try:
        fn(ruta); return ruta
    except PermissionError:
        raiz, ext = os.path.splitext(ruta)
        alt = f"{raiz}_{dt.datetime.now():%H%M%S}{ext}"
        fn(alt)
        print(f"[output] AVISO: '{os.path.basename(ruta)}' está ABIERTO. "
              f"Lo guardé como '{os.path.basename(alt)}'.")
        return alt


# ============================ CORRER EL FRAMEWORK ============================
csv_path = os.path.join(RUTA_INSUMOS, CFG.ARCHIVO_REGISTROS)
if not os.path.exists(csv_path):
    raise SystemExit(f"[output] No encuentro:\n  {csv_path}\n"
                     f"Corre primero construir_input_mec.py (misma carpeta).")

df = mec.m1_cargar_registros(csv_path, CFG)
print(f"[output] Registros de vigencia: {len(df):,} combinaciones · prima {df['Prima'].sum():,.0f}")
vectores, cartera, pesos = mec.m2_fnd_prorrata(df, CFG)
decisiones = mec.m3_decision_apertura(df, vectores, cartera, pesos, CFG)
tabla = mec.m6_publicar(vectores, cartera, decisiones, CFG)
ABIERTOS = sorted([r for r, d in decisiones.items() if d['abrir']], key=lambda x: -pesos[x])
print(f"[output] Ramos con vector propio: {ABIERTOS or 'ninguno'}")

orden = sorted(vectores.keys(), key=lambda r: -pesos[r])
wb = Workbook(); wb.remove(wb.active)

# --------------------------------- Resumen ----------------------------------
ws = wb.create_sheet("Resumen"); LC = 10
for c in range(1, LC + 1): ws.column_dimensions[GCL(c)].width = 14
banda(ws, "MEC · OUTPUT DEL FACTOR DE NO DEVENGAMIENTO (FND)",
      f"Base: prorrata exacta de las vigencias reales de Patria (PF+) · generado {dt.date.today()}", LC)
ws.merge_cells(start_row=4, start_column=1, end_row=6, end_column=LC)
C(ws, 4, 1, "HALLAZGO QUE CAMBIÓ LA BASE: PrimasNal es prima REGISTRADA (suscrita/causada en el mes "
             "contable), no devengada por riesgo corrido. Un triángulo de cohortes sobre esa columna mide "
             "REZAGO DE REGISTRO, no devengamiento. El FND ahora sale de la prorrata exacta de las fechas "
             "de vigencia. ALCANCE: sólo el FND; FS, 1−LAG, MR, IRR e índice se mantienen.",
  bold=True, color=BLANCO, fill=ROJO, size=10)
r = 8
C(ws, r, 1, "Parámetros de la corrida", bold=True, color=BLANCO, fill=VERDE); r += 1
prima_tot = df['Prima'].sum()
dur = (df['fin'] - df['ini']).dt.days
med = float(np.average(dur, weights=np.abs(df['Prima']))) if prima_tot else 0
for k, v in [("Fuente", csv_path),
             ("Registros (Ramo × vigencia)", f"{len(df):,}"),
             ("Prima total", f"{prima_tot:,.0f}"),
             ("Duración media ponderada", f"{med/30.4:.1f} meses"),
             ("Horizonte del vector FND", f"{CFG.HORIZONTE} meses"),
             ("Compuerta de apertura", f"back-test fuera de muestra en pesos de reserva · mejora ≥{CFG.MARGEN_APERTURA:.0%} en las {len(CFG.PARTICIONES)} particiones"),
             ("Ramos con vector propio", ", ".join(f"{a} {NOMBRE_RAMO.get(int(float(a)),'')}" for a in ABIERTOS) or "ninguno")]:
    C(ws, r, 1, k, bold=True, fill=GRIS, center=False)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=LC)
    C(ws, r, 2, str(v), fill=CLARO, center=False); r += 1
nota(ws, r + 1, LC, "Hojas: «Comparativo (mes)» NT/GS/PF/PF+ por antigüedad 0–11 · «Escenarios frecuencia» "
     "mensual, trimestral, cuatrimestral y semestral · «PF+ por ramo» vectores y decisión · "
     "«Diagnóstico» la evidencia del cambio de base y lo que falta para abrir Agro.", h=40)

# ---------------------------- Comparativo (mes) ------------------------------
ws = wb.create_sheet("Comparativo (mes)")
# columnas base + 2 por cada ramo abierto (valor y delta contra la cartera PF+)
NCOL = 10 + 2 * len(ABIERTOS)
anchos = [11,13,12,12,12,13,13,14,12,12] + [12,11] * len(ABIERTOS)
for i, w in enumerate(anchos): ws.column_dimensions[GCL(1+i)].width = w
banda(ws, "FND POR ANTIGÜEDAD (0–11 MESES) · % NO DEVENGADO · cierre = diciembre",
      "NT y PF en base mensual · PF+ Cartera y PF+ de cada ramo ABIERTO, con su delta contra la cartera PF+", NCOL)
r = 4
hdr = [("Antigüedad\n(meses)",VERDE),("Mes emisión\nequiv.",VERDE),("NT\n(lineal 24-avos)",ORO),
       ("GS Cartera",AZUL),("GS MyT (050)",AZUL),("GS Agro (081)",AZUL),("PF\n(tabla xPND)",MORADO),
       ("PF+ Cartera\n(vigencias reales)",VERDE),("Δ PF+ − NT",GRIS),("Δ PF+ − PF",GRIS)]
for i,(h,cl) in enumerate(hdr):
    C(ws, r, 1+i, h, bold=True, color=('FF222222' if cl==GRIS else BLANCO), fill=cl)
PAL_R = ['FFB5541C','FF1F4E79','FF5B2C6F','FF0F6B6B','FF7A3E9D','FF8A6D1D','FFA6192E','FF444444']
for j, ramo in enumerate(ABIERTOS):
    ri = int(float(ramo)); cl = PAL_R[j % len(PAL_R)]
    C(ws, r, 11+2*j, f"PF+ {ri}\n{NOMBRE_RAMO.get(ri,'')}", bold=True, color=BLANCO, fill=cl)
    C(ws, r, 12+2*j, f"Δ {ri} − Cartera", bold=True, color=BLANCO, fill=cl)
ws.row_dimensions[r].height = 38
r += 1; first = r
for k in range(12):
    pp = float(cartera[k]); bf = GRIS if k % 2 == 0 else CLARO
    vals = [k, MESDIC[k], NT_MENS[k], GS_CART[k], GS_MYT[k], GS_AGRO[k], PF_MENS[k],
            pp, pp - NT_MENS[k], pp - PF_MENS[k]]
    for i, v in enumerate(vals):
        nf = None if i == 1 else ('0' if i == 0 else '0.00%')
        C(ws, r, 1+i, v, bold=(i == 7), color=(VERDE if i == 7 else 'FF222222'),
          fill=bf, nf=nf, border=True, mono=(i >= 2))
    for j, ramo in enumerate(ABIERTOS):
        vr = float(vectores[ramo][k]); d = vr - pp
        C(ws, r, 11+2*j, vr, bold=True, color=PAL_R[j % len(PAL_R)], fill=bf, nf='0.00%', border=True, mono=True)
        C(ws, r, 12+2*j, d, color=(ROJO if abs(d) >= 0.02 else 'FF222222'),
          fill=bf, nf='0.00%', border=True, mono=True)
    r += 1
last = r - 1
nota(ws, r, NCOL, "PF+ va 2–3 pp por debajo de NT porque la vigencia real mediana es de 11 meses, no 12. Los ramos "
     "ABIERTOS son los que el back-test fuera de muestra validó: su columna Δ muestra cuánto se separan de la cartera "
     "PF+ en cada mes. GS Agro (081) sigue siendo el único vector claramente no lineal y esa forma no se reproduce con "
     "prorrata — requiere curva de exposición.", h=44)
r += 2
ch = LineChart(); ch.title = "FND por antigüedad — referencias y ramos abiertos"
ch.height = 10; ch.width = 20; ch.y_axis.numFmt = '0%'
ch.x_axis.delete = False; ch.y_axis.delete = False
ch.y_axis.title = '% no devengado'; ch.x_axis.title = 'antigüedad (meses)'
series = [(3,"NT lineal","C9A961"),(6,"GS Agro","A6192E"),(7,"PF (xPND)","5B2C6F"),(8,"PF+ Cartera","00573F")]
for j, ramo in enumerate(ABIERTOS):
    series.append((11+2*j, f"PF+ {int(float(ramo))}", PAL_R[j % len(PAL_R)][2:]))
for col, nm, color in series:
    sr = Series(Reference(ws, min_col=col, min_row=first, max_row=last), title=nm)
    sr.graphicalProperties.line = LineProperties(solidFill=color, w=20000); sr.smooth = False
    ch.series.append(sr)
ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=last))
ws.add_chart(ch, f"A{r}")

# -------------------------- Escenarios frecuencia ----------------------------
ws = wb.create_sheet("Escenarios frecuencia")
for i, w in enumerate([11,13,13,13,14,13,12]): ws.column_dimensions[GCL(1+i)].width = w
banda(ws, "ESCENARIOS POR FRECUENCIA DE CUENTAS · MENSUAL · TRIMESTRAL · CUATRIMESTRAL · SEMESTRAL",
      "FND_t(k) = max(0, FND_mensual(k) − (t−1)/2 × 30/365) · regla verificada EXACTA contra los vectores NT", 7)
r = 4
for nombre, cl, base in [("NT (Nota Técnica)", ORO, NT_MENS),
                         ("PF (tabla xPND)", MORADO, PF_MENS),
                         ("PF+ (vigencias reales)", VERDE, list(cartera[:12]))]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    C(ws, r, 1, f"{nombre} — % no devengado por antigüedad y frecuencia", bold=True, color=BLANCO, fill=cl)
    r += 1
    C(ws, r, 1, "Antigüedad", bold=True, color=BLANCO, fill=cl)
    C(ws, r, 2, "Mes equiv.", bold=True, color=BLANCO, fill=cl)
    for j, (fn, _) in enumerate(FRECS):
        C(ws, r, 3+j, fn, bold=True, color=BLANCO, fill=cl)
    r += 1
    for k in range(12):
        bf = GRIS if k % 2 == 0 else CLARO
        C(ws, r, 1, k, fill=bf, nf='0', border=True)
        C(ws, r, 2, MESDIC[k], fill=bf, border=True)
        for j, (fn, t) in enumerate(FRECS):
            v = float(mec.m4_escenario_frecuencia(base, t, CFG)[k])
            C(ws, r, 3+j, v, fill=bf, nf='0.00%', border=True, mono=True)
        r += 1
    r += 1
nota(ws, r, 7, "Con cuentas trimestrales (64% de la prima de Patria) la prima llega agrupada, así que al registrarla "
     "ya corrió en promedio 1 mes de riesgo y el FND baja frente al escenario mensual. Cuatrimestral y semestral "
     "acentúan el efecto (1.5 y 2.5 meses). Anual se incluye como referencia extrema.", h=36)

# ------------------------------ PF+ por ramo ---------------------------------
ws = wb.create_sheet("PF+ por ramo")
for i, w in enumerate([8,26,9] + [9]*12 + [11,11,11,9,12,11]): ws.column_dimensions[GCL(1+i)].width = w
banda(ws, "PF+ POR RAMO · FND DE PRORRATA EXACTA Y DECISIÓN DE APERTURA (back-test fuera de muestra)",
      f"La apertura es POR RAMO y la decide un back-test FUERA DE MUESTRA en pesos de reserva: abre si el vector "
      f"abre si su vector se separa MATERIALMENTE de la cartera (≥{CFG.UMBRAL_MATERIAL:.0%} en promedio) Y gana el "
      f"back-test fuera de muestra en todas las particiones", 21)
r = 4
C(ws, r, 1, "Ramo", bold=True, color=BLANCO, fill=VERDE)
C(ws, r, 2, "Nombre", bold=True, color=BLANCO, fill=VERDE)
C(ws, r, 3, "% prima", bold=True, color=BLANCO, fill=VERDE)
for k in range(12): C(ws, r, 4+k, f"k={k}", bold=True, color=BLANCO, fill=VERDE)
C(ws, r, 16, "dif media\nvs cartera", bold=True, color=BLANCO, fill=VERDE)
C(ws, r, 17, "err propio\n(fuera muestra)", bold=True, color=BLANCO, fill=VERDE)
C(ws, r, 18, "err cartera", bold=True, color=BLANCO, fill=VERDE)
C(ws, r, 19, "mejora", bold=True, color=BLANCO, fill=VERDE)
C(ws, r, 20, "estabilidad\n(particiones)", bold=True, color=BLANCO, fill=VERDE)
C(ws, r, 21, "Decisión", bold=True, color=BLANCO, fill=VERDE)
ws.row_dimensions[r].height = 30
r += 1
C(ws, r, 1, "—", bold=True, color=BLANCO, fill='FF808080')
C(ws, r, 2, "CARTERA (referencia)", bold=True, color=BLANCO, fill='FF808080', center=False)
C(ws, r, 3, 1.0, color=BLANCO, fill='FF808080', nf='0.0%')
for k in range(12): C(ws, r, 4+k, float(cartera[k]), bold=True, fill=CLARO, nf='0.0%', border=True, mono=True)
for cc in (16,17,18,19,20): C(ws, r, cc, "—", fill=CLARO, border=True)
C(ws, r, 21, "base", fill=CLARO, border=True)
r += 1
for ramo in orden:
    d = decisiones[ramo]; abre = d['abrir']
    bf = CREMA if abre else (GRIS if r % 2 == 0 else CLARO)
    ri = int(float(ramo))
    C(ws, r, 1, ri, bold=True, color=BLANCO, fill=(VERDE2 if abre else 'FF9AA5A0'), nf='0')
    C(ws, r, 2, NOMBRE_RAMO.get(ri, '—'), bold=abre, fill=bf, center=False)
    C(ws, r, 3, d['peso'], fill=bf, nf='0.0%', border=True)
    for k in range(12):
        C(ws, r, 4+k, float(vectores[ramo][k]), bold=abre,
          color=(VERDE if abre else 'FF222222'), fill=bf, nf='0.0%', border=True, mono=True)
    dm = d.get('dif_media', 0.0); mat = d.get('material', False)
    C(ws, r, 16, dm, bold=mat, color=(VERDE if mat else ROJO), fill=bf, nf='0.00%', border=True)
    C(ws, r, 17, d.get('err_propio'), fill=bf, nf='0.00%', border=True)
    C(ws, r, 18, d.get('err_cartera'), fill=bf, nf='0.00%', border=True)
    C(ws, r, 19, d.get('mejora'), fill=bf, nf='0.0%', border=True)
    C(ws, r, 20, f"{d.get('votos',0)}/{d.get('n_particiones',0)}", fill=bf, border=True)
    C(ws, r, 21, "ABRIR" if abre else "cartera", bold=True,
      color=(VERDE if abre else 'FF222222'), fill=bf, border=True)
    r += 1
nota(ws, r, 21, "AGRÍCOLA (80) no abre por prorrata: su vigencia es como la de la cartera. Su particularidad es "
     "ESTACIONAL (ciclo de cosecha) y la prorrata no la ve — requiere curva de exposición con fecha de ocurrencia "
     "de SIREC (ver Diagnóstico). Mientras tanto opera con cartera; no se inventa una curva.",
     fill=ROJO, color=BLANCO, h=40)

# ------------------------------- Diagnóstico ---------------------------------
ws = wb.create_sheet("Diagnóstico")
for c, w in [(1,6),(2,30),(3,68),(4,18)]: ws.column_dimensions[GCL(c)].width = w
banda(ws, "DIAGNÓSTICO · POR QUÉ CAMBIÓ LA BASE DEL FND Y QUÉ FALTA PARA ABRIR AGRO",
      "Evidencia verificada sobre la BD · alcance: solo devengamiento", 4)
r = 4
for i, h in enumerate(["#", "Hallazgo", "Evidencia", "Consecuencia"]):
    C(ws, r, 1+i, h, bold=True, color=BLANCO, fill=VERDE)
r += 1
HALL = [("1","PrimasNal es prima REGISTRADA, no devengada",
  "La BD se genera de dbo_aMOG_MovGonzalo (Tipo=5); en Gonz «devengado» es causación contable, cuyo opuesto es "
  "pagado/caja. La BD no tiene columna de prima devengada ni de reservas; sí tiene Cuentas Rendidas, Dentro de, "
  "A pagar en, Periodos y Meses Periodo.", "El triángulo mide rezago de registro"),
 ("2","El dato lo confirma empíricamente",
  "94% de la prima tiene vigencia ≤12 meses (mediana 11), pero el triángulo de cohortes llegaba a «100% emergido» "
  "hacia el mes 6. Eso es la cuenta que terminó de llegar, no el riesgo expirado.",
  "Ese FND liberaría reserva de más"),
 ("3","La BD sí permite el FND correcto",
  "Fecha Inicio y Fecha Fin de Vigencia están completas en ~99% de los registros de prima. Con ellas el FND se "
  "calcula por prorrata exacta registro por registro y se agrega por ramo (PF+), sin estimar nada.",
  "PF+ es la nueva base"),
 ("4","La prorrata SÓLO distingue por DURACIÓN, y aquí casi todo es a 12 meses",
  "Como el 94% del negocio dura ~12 meses, los vectores PF+ de casi todos los ramos quedan pegados a la cartera "
  "(diferencia media de 1–2 pp). Sólo Vida se separa de verdad (7 pp) por su cola multianual. Abrir un ramo cuyo "
  "vector es casi la cartera es sobre-ingeniería: no cambia la reserva.", "Se exige diferencia MATERIAL"),
 ("5","Criterio de apertura: MATERIAL + persistente",
  "Un back-test por mejora relativa abría ramos con vectores casi iguales a la cartera (mejora grande sobre errores "
  "diminutos). Ahora el ramo abre sólo si su vector se separa de la cartera ≥4 pp en promedio (en la ventana donde "
  "el FND es > 0) Y gana el back-test fuera de muestra en las 6 particiones. Con eso abre sólo Vida.",
  "Parsimonia con sustento"),
 ("6","Agro NO se resuelve con prorrata — su diferencia es ESTACIONAL",
  "Por DURACIÓN, Agrícola es casi igual a la cartera (1.8 pp), así que la prorrata no la distingue. Pero su riesgo se "
  "concentra en el ciclo de cosecha: el vector GS 081 cae a 9% en el mes 5, una diferencia de 42 pp contra la cartera. "
  "Esa forma depende de CUÁNDO ocurre el siniestro dentro de la vigencia, no de la duración. Se intentó con los "
  "siniestros de la BD y no es viable (mismo rezago de registro). Requiere FECHA DE OCURRENCIA de SIREC (M5).",
  "Traer ocurrencia de SIREC"),
 ("7","El rezago de registro no se tira",
  "El patrón que estimaba el MEC v1 es información válida, pero responde a otra pregunta: cuánta prima de un periodo "
  "ya iniciado aún no se ha reportado. Se conserva en mec.extra_triangulo_registro().", "Afecta a PT, no al FND")]
for n, h, e, c in HALL:
    bf = GRIS if int(n) % 2 == 1 else CLARO
    C(ws, r, 1, n, bold=True, color=BLANCO, fill=VERDE2)
    C(ws, r, 2, h, bold=True, fill=bf, center=False)
    C(ws, r, 3, e, fill=bf, center=False, size=8.5)
    C(ws, r, 4, c, bold=True, color=(ROJO if n in ('1','2','6') else VERDE), fill=bf)
    ws.row_dimensions[r].height = 62
    r += 1
nota(ws, r, 4, "Alcance intacto: este trabajo cambia sólo la fuente del FND. FS/BELMEDIA, el desarrollo de siniestros "
     "(1−LAG), el margen de riesgo, el IRR y el índice se mantienen con el método vigente.", fill=CLARO, h=30)

out = _guardar_seguro(wb.save, os.path.join(RUTA_SALIDA, "Output_MEC_Devengamiento.xlsx"))
print(f"[output] Escrito: {out}")
