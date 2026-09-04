# -*- coding: utf-8 -*-
"""
================================================================================
 construir_input_mec.py · Construye el INPUT canónico del MEC (devengamiento)
--------------------------------------------------------------------------------
 Planeación Financiera (BP&A) · Reaseguradora Patria (GPV)

 Fuentes (jerarquía del Layout del Input):
   1) PRINCIPAL  · BD_PptoTécnicoRPAT_GENERADA.xlsx (hoja «BD»)
   2) SECUNDARIA · ResumenPptoTécnico*.xlsm/.xlsx (hoja «BD_TXT») en la carpeta
   3) ALTERNA    · PptoTécnico*.txt (esquema 41 columnas de la herramienta)

 Produce:
   - Input_MEC_Devengamiento.xlsx   (hojas: Input · Cobertura · Validaciones)
   - TriangulosPrimaDevengada.csv   (lo que consume mec_devengamiento: unidad=LN2)

 Reglas clave:
   · SOLO primas (candado): BD Tipo Póliza ∈ {PD,PF,PV}; herramienta cuenta 61 ×(−1)
   · Origen 'Real' SOLO en la fuente BD con Periodo ≤ FRONTERA_REAL;
     todo lo de la herramienta es proyección ('Ppto') — el MEC estima con Real.
   · Cohorte MENSUAL desde Fecha Inicio de Vigencia (BD). La herramienta solo trae
     año (ZSUSCYEAR) → cohorte anclada a enero, y SOLO aporta la ventana jul–dic 2026.
================================================================================
"""
import os, re, glob, getpass
import datetime as _dt

def _guardar_seguro(guardar_fn, ruta, que="archivo"):
    """Guarda con guardar_fn(ruta). Si el archivo está ABIERTO/bloqueado en Windows
    (PermissionError), reintenta con un nombre alterno con hora y avisa, sin tronar."""
    try:
        guardar_fn(ruta)
        return ruta
    except PermissionError:
        raiz, ext = os.path.splitext(ruta)
        alt = f"{raiz}_{_dt.datetime.now():%H%M%S}{ext}"
        guardar_fn(alt)
        print(f"[input] AVISO: '{os.path.basename(ruta)}' está ABIERTO (probablemente en Excel).")
        print(f"[input]        Lo guardé como '{os.path.basename(alt)}'. Cierra el original y")
        print(f"[input]        renómbralo, o vuelve a correr con el archivo cerrado.")
        return alt
import numpy as np
import pandas as pd

# =============================== CONFIG =====================================
# El script se ancla a SU PROPIA CARPETA (ponlo en «…\OneDrive - GPV\Documents»).
# Así no depende del usuario, de acentos ni del directorio desde el que lo ejecutes.
BASE = os.path.dirname(os.path.abspath(__file__))

RUTA_BD_EXPLICITA = ""       # opcional: ruta completa de la BD si está en otro lado
RUTA_RESUMENES    = os.path.join(BASE, "Resumen Ppto Tecnico")
RUTA_SALIDA       = BASE

FRONTERA_REAL  = 202606      # último Periodo REAL de la BD (edítalo en cada cierre)
VENTANA_PPTO   = (202607, 202612)  # jul–dic 2026: meses que aporta la herramienta (Resumen/Txt)
MONEDA         = "Nal"       # "Nal" (PrimasNal, MXN) o "USD" (Primas USD)
COHORTE_MIN    = 1990        # cohortes por debajo se excluyen (V4) y se reportan
RUTA_SUBRAMO   = None        # opcional: CSV CeBe→Ramo para las fuentes de la herramienta


def localizar_bd():
    """Encuentra la BD en la carpeta del script (tolera acentos y variantes de nombre).
    Si no la encuentra, ABORTA con un mensaje claro (no inventa rutas)."""
    if RUTA_BD_EXPLICITA:
        if os.path.exists(RUTA_BD_EXPLICITA):
            return RUTA_BD_EXPLICITA
        raise SystemExit(f"[input] RUTA_BD_EXPLICITA no existe:\n  {RUTA_BD_EXPLICITA}")
    patrones = ["BD_Ppto*.xls*", "BD*Tecnico*RPAT*.xls*", "BD*Técnico*RPAT*.xls*"]
    cand = []
    for p in patrones:
        cand += glob.glob(os.path.join(BASE, p))
        cand += glob.glob(os.path.join(BASE, "*", p))      # una subcarpeta de profundidad
    cand = list(set(cand))
    # Prioridad: primero el AÑO más alto que aparezca en el nombre (2027 > 2026 > sin año),
    # y a igualdad de año, el archivo más reciente. Así "BD_...2027" gana a "BD_...GENERADA".
    def _anio(f):
        aa = re.findall(r'(20\d{2})', os.path.basename(f))
        return max(int(a) for a in aa) if aa else 0
    cand = sorted(cand, key=lambda f: (_anio(f), os.path.getmtime(f)), reverse=True)
    if cand:
        if len(cand) > 1:
            print("[input] Varias BD candidatas; uso la de año más reciente:")
            for c in cand[:5]:
                marca = f" (año {_anio(c)})" if _anio(c) else ""
                print(f"        {'-> ' if c == cand[0] else '   '}{os.path.basename(c)}{marca}")
        return cand[0]
    vistos = sorted(os.path.basename(f) for f in glob.glob(os.path.join(BASE, "*.xls*")))
    raise SystemExit(
        "[input] No encontré la BD (patrón «BD_Ppto*.xls*») en:\n"
        f"  {BASE}\n"
        "Archivos Excel que sí veo ahí:\n  " + ("\n  ".join(vistos) if vistos else "(ninguno)") + "\n"
        "Soluciones: (a) copia BD_PptoTécnicoRPAT_GENERADA.xlsx a esa carpeta, o\n"
        "            (b) escribe su ruta completa en RUTA_BD_EXPLICITA (arriba)."
    )


def leer_excel(path, **kw):
    """read_excel con el engine correcto según extensión (xlsb requiere pyxlsb)."""
    if os.path.splitext(path)[1].lower() == ".xlsb":
        kw.setdefault("engine", "pyxlsb")
    try:
        return pd.read_excel(path, **kw)
    except ImportError:
        raise SystemExit("[input] Falta el motor para .xlsb. Instala:  pip install pyxlsb")


_EPOCH_XL = pd.Timestamp('1899-12-30')

def _a_fecha_serie(serie):
    """Convierte una columna de fechas de vigencia a datetime, tolerando que la
    columna venga MEZCLADA. En la base 2027 conviven en la MISMA columna valores
    datetime (139k) y seriales de Excel como entero (56k); resolver la columna con
    un solo criterio malinterpreta los seriales y colapsa la vigencia a 0 días
    (se perdía el 30% de los registros). Por eso se resuelve valor por valor:
    los numéricos plausibles como serial (1000..80000) se convierten desde la época
    de Excel y el resto se parsea como fecha."""
    s = pd.Series(serie).reset_index(drop=True)
    if np.issubdtype(getattr(s, "dtype", object), np.datetime64):
        return pd.to_datetime(s, errors='coerce')
    num = pd.to_numeric(s, errors='coerce')
    es_serial = num.notna() & (num > 1000) & (num < 80000)
    out = pd.Series(pd.NaT, index=s.index, dtype='datetime64[ns]')
    if es_serial.any():
        out.loc[es_serial] = _EPOCH_XL + pd.to_timedelta(num[es_serial], unit='D')
    resto = ~es_serial
    if resto.any():
        out.loc[resto] = pd.to_datetime(s[resto], errors='coerce')
    out.index = pd.Series(serie).index
    return out


# columnas mínimas que definen la hoja de datos de la BD
_CLAVE_HOJA = {'Periodo', 'PrimasNal', 'Fecha Inicio de Vigencia', 'Ramo', 'Tipo Poliza'}

def hoja_datos(path):
    """Devuelve el nombre de la hoja que contiene los datos de la BD, buscándola por
    sus columnas. Así funciona igual si la hoja se llama «BD», «Hoja1» o cualquier
    otra (la base 2027 la trae como «Hoja1»)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        hojas = wb.sheetnames
        wb.close()
    except Exception:
        hojas = None
    candidatas = (['BD'] + [h for h in (hojas or []) if h != 'BD']) if hojas else ['BD', 'Hoja1', 0]
    for h in candidatas:
        try:
            cols = set(leer_excel(path, sheet_name=h, nrows=0).columns)
        except Exception:
            continue
        if _CLAVE_HOJA <= cols:
            return h
    # último recurso: la primera hoja
    return (hojas[0] if hojas else 0)

COLS41 = ['/ERP/CHRTACCT','/ERP/GL_ACCT','/ERP/CATEGORY','ZSOURCE','/ERP/CO_AREA',
 '/ERP/COMPCODE','/ERP/FUNCAREA','ZPAISCEDN','/ERP/PRODUCT','ZCUSTOMER',
 '0CURRENCY','/ERP/TWAERS','/ERP/COSTCNTR','/ERP/PROFTCTR','ZCONTRATO',
 'ZPLANYEAR','0FISCPER','0FISCPER3','0FISCYEAR','0FISCVARNT',
 '0CALMONTH','0CALMONTH2','0CALYEAR','ZINDICES','ZCONCEPTO',
 'ZDISTCHRP','ZDISTCHGS','ZTIPOREAS','ZTIPOCES','ZMGA',
 'ZOFICN_RP','ZOFICN_GS','ZSUSCYEAR','ZCEDENTE','ZCORREDOR',
 'ZTIPVENTA','ZSEGMENTO','ZCICCULTV','/ERP/AMOUNT_T','/ERP/AMOUNT',
 'MANDT','ZREGIONRP']

CANON = ['Fuente','LN2','Ramo','TipoRea','CohorteAnio','CohorteAAAAMM','Periodo',
         'Antiguedad','Origen','Moneda','PrimaDevMes','PrimaDevAcum']

REPORTE = []   # (regla, detalle, valor, severidad)
def rep(regla, detalle, valor, sev): REPORTE.append((regla, detalle, valor, sev))


# ============================ FUENTE 1 · BD =================================
def fuente_bd():
    ruta = localizar_bd()
    hoja = hoja_datos(ruta)
    print(f"[input] BD principal: {ruta}  (hoja «{hoja}»)")
    use = ['Periodo','Año Susc.','Fecha Inicio de Vigencia','LN2','Ramo','Tipo Poliza','Tipo Rea','PrimasNal','Primas USD']
    cols = list(leer_excel(ruta, sheet_name=hoja, nrows=0).columns)
    faltan = [c for c in use if c not in cols]
    if faltan:
        raise SystemExit(f"[input] A la hoja «BD» le faltan columnas: {faltan}\n"
                         f"        Columnas que sí tiene: {cols}")
    bd = leer_excel(ruta, sheet_name=hoja, usecols=use)
    n0 = len(bd)
    prim = bd[bd['Tipo Poliza'].astype(str).str.startswith('P')].copy()
    rep('V2','Filas excluidas por no ser primas (S*/R*/OD/«20») — fuente BD', n0-len(prim), 'OK (candado)')
    # COHORTE MENSUAL desde Fecha Inicio de Vigencia (viene como SERIAL de Excel -> fecha)
    iv = _a_fecha_serie(prim['Fecha Inicio de Vigencia'])
    sin_iv = iv.isna()
    if sin_iv.any():
        rep('V4b', 'Filas sin Fecha Inicio de Vigencia válida (excluidas)',
            f"{int(sin_iv.sum())} filas · monto {prim.loc[sin_iv, ('PrimasNal' if MONEDA=='Nal' else 'Primas USD')].sum():,.0f}",
            'Bloqueante aplicada')
    prim = prim[~sin_iv].copy(); iv = iv[~sin_iv]
    prim['CohorteAAAAMM'] = (iv.dt.year * 100 + iv.dt.month).astype(int)
    monto = 'PrimasNal' if MONEDA == 'Nal' else 'Primas USD'
    g = (prim.groupby(['LN2','Ramo','Tipo Rea','CohorteAAAAMM','Periodo'], as_index=False)[monto]
             .sum().rename(columns={'Tipo Rea':'TipoRea','Periodo':'Periodo', monto:'PrimaDevMes'}))
    g['Fuente'] = 'BD'
    print(f"[input]   {len(bd):,} filas · {len(prim):,} de primas · {len(g):,} filas agregadas · "
          f"cohortes MENSUALES por inicio de vigencia")
    return g


def registros_vigencia():
    """Genera Registros_Vigencia_MEC.csv: la fuente del FND por PRORRATA EXACTA.

    Una fila por (Ramo, inicio de vigencia, fin de vigencia) con la prima agregada.
    Es lo que consume mec_devengamiento (M2) para calcular el FND sin estimar nada:
    el devengamiento sale de las fechas reales de cobertura, no de un triángulo de
    registro (ver encabezado del módulo)."""
    ruta = localizar_bd()
    hoja = hoja_datos(ruta)
    use = ['Ramo','Tipo Poliza','Fecha Inicio de Vigencia','Fecha Fin de Vigencia',
           'PrimasNal','Primas USD']
    cols = list(leer_excel(ruta, sheet_name=hoja, nrows=0).columns)
    faltan = [c for c in use if c not in cols]
    if faltan:
        print(f"[input] Aviso: sin columnas {faltan}; no se puede generar el archivo de vigencias.")
        return None
    bd = leer_excel(ruta, sheet_name=hoja, usecols=use)
    p = bd[bd['Tipo Poliza'].astype(str).str.startswith('P')].copy()
    p['ini'] = _a_fecha_serie(p['Fecha Inicio de Vigencia'])
    p['fin'] = _a_fecha_serie(p['Fecha Fin de Vigencia'])
    monto = 'PrimasNal' if MONEDA == 'Nal' else 'Primas USD'
    dur = (p['fin'] - p['ini']).dt.days
    ok = p['ini'].notna() & p['fin'].notna() & (dur > 0) & (dur <= 3660)
    rep('V11', 'Registros de prima con vigencia válida (base del FND por prorrata)',
        f"{int(ok.sum()):,} de {len(p):,} ({ok.mean():.1%})",
        'OK' if ok.mean() >= 0.95 else 'Advertencia — revisar fechas')
    p = p[ok]
    g = (p.groupby(['Ramo', 'ini', 'fin'], as_index=False)[monto]
           .sum().rename(columns={monto: 'Prima'}))
    x = _guardar_seguro(lambda ruta_out: g.to_csv(ruta_out, index=False),
                        os.path.join(RUTA_SALIDA, 'Registros_Vigencia_MEC.csv'), "Vigencias")
    print(f"[input]   Vigencias: {len(g):,} combinaciones (Ramo × vigencia) · "
          f"duración mediana {dur[ok].median()/30.4:.1f} meses")
    return x


# ==================== FUENTES 2/3 · HERRAMIENTA =============================
def _detectar_ln(df, path):
    if '/ERP/FUNCAREA' in df.columns and df['/ERP/FUNCAREA'].notna().any():
        return str(df['/ERP/FUNCAREA'].mode().iloc[0])
    if 'ID' in df.columns and df['ID'].notna().any():          # p.ej. LN04004-1-1161-0-2023-...
        return str(df['ID'].dropna().iloc[0]).split('-')[0]
    m = re.search(r'LN\d{5}', os.path.basename(path))
    return m.group(0) if m else 'LN_DESCONOCIDA'

def _normaliza_herramienta(df, fuente_tag, path):
    # Primas = cuenta 61; signo invertido en la herramienta -> ×(−1)
    df = df[df['/ERP/GL_ACCT'].astype(str).str.startswith('61')].copy()
    df['PrimaDevMes'] = pd.to_numeric(df['/ERP/AMOUNT'], errors='coerce') * -1.0
    ln = _detectar_ln(df, path)
    df['LN2'] = ln
    if RUTA_SUBRAMO and os.path.exists(RUTA_SUBRAMO) and '/ERP/PROFTCTR' in df.columns:
        cat = pd.read_csv(RUTA_SUBRAMO)[['CeBe','Ramo']].drop_duplicates()
        df = df.merge(cat, how='left', left_on='/ERP/PROFTCTR', right_on='CeBe')
        df['Ramo'] = df['Ramo'].fillna(0).astype(int)
    else:
        df['Ramo'] = 0    # por asignar vía catálogo Subramo (CeBe) — no bloquea el input por LN2
    # La herramienta solo trae AÑO de suscripción -> cohorte anclada a ENERO (AAAA01).
    df['CohorteAAAAMM'] = pd.to_numeric(df['ZSUSCYEAR'], errors='coerce') * 100 + 1
    df['Periodo'] = pd.to_numeric(df['0CALMONTH'], errors='coerce')
    # SOLO se usa la proyección de suscripción para la VENTANA Ppto (jul–dic del año frontera).
    df = df[(df['Periodo'] >= VENTANA_PPTO[0]) & (df['Periodo'] <= VENTANA_PPTO[1])]
    g = (df.groupby(['LN2','Ramo','ZTIPOREAS','CohorteAAAAMM','Periodo'], as_index=False)['PrimaDevMes'].sum()
           .rename(columns={'ZTIPOREAS':'TipoRea'}))
    g['Fuente'] = f'{fuente_tag}-{ln}'
    print(f"[input]   {os.path.basename(path)} -> {ln} · {len(g)} filas en ventana Ppto {VENTANA_PPTO[0]}–{VENTANA_PPTO[1]}")
    return g

def fuente_xlsm(path):
    head = leer_excel(path, sheet_name='BD_TXT', nrows=0)
    quiero = ['/ERP/FUNCAREA','/ERP/GL_ACCT','ZSUSCYEAR','0CALMONTH','ZTIPOREAS',
              '/ERP/AMOUNT','0CURRENCY','/ERP/PROFTCTR','ID']
    use = [c for c in quiero if c in head.columns]
    df = leer_excel(path, sheet_name='BD_TXT', usecols=use)
    return _normaliza_herramienta(df, 'Resumen', path)

def fuente_txt(path):
    try:
        df = pd.read_csv(path, delimiter='\t', names=COLS41, header=None,
                         thousands=',', low_memory=False, skip_blank_lines=True)
    except UnicodeDecodeError:   # los export de la herramienta suelen venir en latin-1
        df = pd.read_csv(path, delimiter='\t', names=COLS41, header=None, encoding='latin-1',
                         thousands=',', low_memory=False, skip_blank_lines=True)
    return _normaliza_herramienta(df, 'Txt', path)

def fuentes_herramienta():
    piezas = []
    if not os.path.isdir(RUTA_RESUMENES):
        print(f"[input] Aviso: no veo la carpeta de resúmenes:\n         {RUTA_RESUMENES}\n         Sigo solo con la BD (el input queda válido, sin las LN de la herramienta).")
        return piezas
    for f in sorted(glob.glob(os.path.join(RUTA_RESUMENES, '*.xls[mx]'))):
        if re.search(r'vsReal', os.path.basename(f), re.I):
            continue      # productos de validación, no fuente (Layout · Fuentes)
        try:
            piezas.append(fuente_xlsm(f))
        except Exception as e:
            print(f"[input]   {os.path.basename(f)}: omitido ({e})")
    for f in sorted(glob.glob(os.path.join(RUTA_RESUMENES, '*.txt'))):
        try:
            piezas.append(fuente_txt(f))
        except Exception as e:
            print(f"[input]   {os.path.basename(f)}: omitido ({e})")
    return piezas


# ====================== DERIVADOS Y VALIDACIONES ============================
def derivar(df):
    df = df.copy()
    df['CohorteAAAAMM'] = pd.to_numeric(df['CohorteAAAAMM'], errors='coerce').fillna(0).astype(int)
    df['Periodo'] = pd.to_numeric(df['Periodo'], errors='coerce').fillna(0).astype(int)
    df['CohorteAnio'] = df['CohorteAAAAMM'] // 100
    # V4 cohortes inválidas (año fuera de rango)
    malas = df['CohorteAnio'] < COHORTE_MIN
    if malas.any():
        rep('V4', f'Filas con cohorte inválida (año <{COHORTE_MIN}) excluidas',
            f"{int(malas.sum())} filas · monto {df.loc[malas,'PrimaDevMes'].sum():,.0f}", 'Bloqueante aplicada')
        df = df[~malas]
    # Antigüedad MENSUAL exacta = meses entre inicio de cohorte (AAAAMM) y periodo (AAAAMM)
    df['Antiguedad'] = ((df['Periodo']//100 - df['CohorteAAAAMM']//100)*12
                        + (df['Periodo']%100 - df['CohorteAAAAMM']%100))
    # V3 antigüedad negativa
    neg = df['Antiguedad'] < 0
    if neg.any():
        rep('V3', 'Filas con antigüedad negativa excluidas (Periodo anterior a la cohorte)',
            f"{int(neg.sum())} filas · monto {df.loc[neg,'PrimaDevMes'].sum():,.0f}", 'Bloqueante aplicada')
        df = df[~neg]
    df['Origen'] = np.where((df['Fuente']=='BD') & (df['Periodo']<=FRONTERA_REAL), 'Real', 'Ppto')
    df['Moneda'] = 'MXN (Nal)' if MONEDA=='Nal' else 'USD'
    df = df.sort_values(['Fuente','LN2','Ramo','TipoRea','CohorteAAAAMM','Periodo'])
    df['PrimaDevAcum'] = df.groupby(['Fuente','LN2','Ramo','TipoRea','CohorteAAAAMM'])['PrimaDevMes'].cumsum()
    return df[CANON]

def validar(df):
    dup = df.duplicated(subset=['Fuente','LN2','Ramo','TipoRea','CohorteAAAAMM','Periodo']).sum()
    rep('V1','Llaves duplicadas', int(dup), 'OK' if dup==0 else 'BLOQUEANTE — revisar agregación')
    rep('V5','Moneda única de la corrida', df['Moneda'].nunique(), 'OK' if df['Moneda'].nunique()==1 else 'BLOQUEANTE')
    reales = df[df['Origen']=='Real']
    rep('V6','Filas Real (estiman el MEC) / Ppto (solo comparación)',
        f"{len(reales):,} / {len(df)-len(reales):,}", 'OK')
    caidas = (df.sort_values('Periodo').groupby(['Fuente','LN2','TipoRea','CohorteAAAAMM'])['PrimaDevAcum']
                .apply(lambda s: int((s.diff() < 0).sum())).sum())
    rep('V7','Meses con acumulado a la baja (devoluciones, signo natural)', int(caidas), 'Advertencia (se reporta)')
    # V9 cuadre BD vs herramienta en la ventana traslapada (proyección 2do semestre del año frontera)
    her = df[df['Fuente']!='BD']
    if len(her):
        vent = (FRONTERA_REAL//100)*100
        a = (df[(df['Fuente']=='BD') & (df['Periodo']>vent) & (df['Periodo']<=FRONTERA_REAL)]
             .groupby('LN2')['PrimaDevMes'].sum())
        b = (her[(her['Periodo']>vent) & (her['Periodo']<=FRONTERA_REAL)]
             .groupby('LN2')['PrimaDevMes'].sum())
        for ln in b.index:
            if ln in a.index and a[ln]!=0:
                d = b[ln]/a[ln]-1
                rep('V9', f'Cuadre BD vs herramienta {ln} ({vent+6}–{FRONTERA_REAL})',
                    f'{d:+.1%}', 'OK' if abs(d)<=0.05 else 'Advertencia — investigar')
    for (ln,), s in df[df['Origen']=='Real'].groupby(['LN2'])['PrimaDevMes']:
        if abs(s.sum()) < 1e6:
            rep('V10', f'Volumen Real bajo en {ln}', f"{s.sum():,.0f}", 'Informativa (irá a patrón de cartera)')


# ============================== SALIDAS =====================================
def escribir(df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    VERDE='FF00573F'; BLANCO='FFFFFFFF'; CLARO='FFE8F1EA'
    wb = Workbook(); wb.remove(wb.active)

    def banda(ws, titulo, sub, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c=ws.cell(1,1,titulo); c.font=Font(bold=True,color=BLANCO,size=12); c.fill=PatternFill('solid',fgColor=VERDE); c.alignment=Alignment('center','center')
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c=ws.cell(2,1,sub); c.font=Font(italic=True,color=BLANCO,size=9); c.fill=PatternFill('solid',fgColor='FF2E7D53'); c.alignment=Alignment('center','center')

    ws = wb.create_sheet('Input')
    banda(ws,'INPUT CANÓNICO DEL MEC · DEVENGAMIENTO (formato del Layout del Input)',
          f'Moneda {df["Moneda"].iloc[0]} · frontera real {FRONTERA_REAL} · Origen Real solo fuente BD · generado por construir_input_mec.py', len(CANON))
    ws.append(CANON)
    for c in ws[3]: c.font=Font(bold=True,color=BLANCO); c.fill=PatternFill('solid',fgColor=VERDE)
    for row in df.itertuples(index=False): ws.append(list(row))
    ws.freeze_panes='A4'

    ws = wb.create_sheet('Cobertura')
    banda(ws,'COBERTURA POR FUENTE × LN2','filas · cohortes · periodos · monto total del mes (PrimaDevMes)',7)
    ws.append(['Fuente','LN2','Filas','Cohorte min','Cohorte max','Periodo min–max','Σ PrimaDevMes'])
    for c in ws[3]: c.font=Font(bold=True,color=BLANCO); c.fill=PatternFill('solid',fgColor=VERDE)
    cov = (df.groupby(['Fuente','LN2']).agg(F=('PrimaDevMes','size'), c0=('CohorteAnio','min'),
           c1=('CohorteAnio','max'), p0=('Periodo','min'), p1=('Periodo','max'), M=('PrimaDevMes','sum')).reset_index())
    for r in cov.itertuples(index=False):
        ws.append([r.Fuente, r.LN2, r.F, r.c0, r.c1, f"{r.p0}–{r.p1}", round(r.M)])

    ws = wb.create_sheet('Validaciones')
    banda(ws,'RESULTADO DE VALIDACIONES DEL INPUT (V1–V10 del Layout)',
          f'frontera real = {FRONTERA_REAL} · cohorte mínima = {COHORTE_MIN}',4)
    ws.append(['Regla','Detalle','Valor','Estado'])
    for c in ws[3]: c.font=Font(bold=True,color=BLANCO); c.fill=PatternFill('solid',fgColor=VERDE)
    for r in REPORTE: ws.append(list(r))

    x1 = _guardar_seguro(wb.save, os.path.join(RUTA_SALIDA,'Input_MEC_Devengamiento.xlsx'), "Input")
    # CSV para el módulo MEC: unidad = LN2, SOLO Real, acumulado al grano LN2×cohorte
    # CSV para el MEC: Real (BD) + Ppto jul–dic 2026 (Resumen) para completar el año en curso.
    base = df[(df['Origen']=='Real') | ((df['Origen']=='Ppto') & (df['Periodo']>=VENTANA_PPTO[0]) & (df['Periodo']<=VENTANA_PPTO[1]))]
    g = (base.groupby(['LN2','CohorteAAAAMM','Periodo'], as_index=False)['PrimaDevMes'].sum()
               .sort_values(['LN2','CohorteAAAAMM','Periodo']))
    g['PrimaDevAcum'] = g.groupby(['LN2','CohorteAAAAMM'])['PrimaDevMes'].cumsum()
    g['Antiguedad'] = (g['Periodo']//100 - g['CohorteAAAAMM']//100)*12 + (g['Periodo']%100 - g['CohorteAAAAMM']%100)
    csv = g.rename(columns={'LN2':'Ramo'})[['Ramo','CohorteAAAAMM','Antiguedad','PrimaDevAcum']]
    x2 = _guardar_seguro(lambda p: csv.to_csv(p, index=False),
                         os.path.join(RUTA_SALIDA,'TriangulosPrimaDevengada.csv'), "CSV")
    print(f"[input] Escritos:\n  {x1}\n  {x2}")
    return x1, x2


if __name__ == '__main__':
    # 1) Fuente del FND (prorrata exacta por fechas de vigencia) — LA BASE del MEC v2
    registros_vigencia()
    # 2) Input canónico por cohorte (registro): trazabilidad, cobertura y análisis
    #    de prima aún no reportada. NO alimenta el FND (ver mec_devengamiento v2).
    piezas = [fuente_bd()] + fuentes_herramienta()
    inp = derivar(pd.concat(piezas, ignore_index=True))
    validar(inp)
    escribir(inp)
    print(f"[input] Input canónico: {len(inp):,} filas · fuentes: {sorted(inp['Fuente'].unique())}")
