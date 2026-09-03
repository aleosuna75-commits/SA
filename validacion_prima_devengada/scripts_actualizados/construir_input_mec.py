# -*- coding: utf-8 -*-
"""
================================================================================
 construir_input_mec.py · Construye el INPUT canónico del MEC (devengamiento)
--------------------------------------------------------------------------------
 Planeación Financiera (BP&A) · Reaseguradora Patria (GPV)

 Fuentes (jerarquía del Layout del Input):
   1) PRINCIPAL  · BD real (BDReal26.xlsx / BD_PptoTécnicoRPAT_GENERADA.xlsx)
                   -> aporta el REAL hasta FRONTERA_REAL
   2) PROYECCIÓN · FCST2026.xlsx (hoja «Ppto2026»)
                   -> aporta la VENTANA_PPTO (los meses que faltan del año)
   3) ALTERNAS   · ResumenPptoTécnico*.xlsm/.xlsx (hoja «BD_TXT») y PptoTécnico*.txt
                   (esquema 41 columnas). Sólo se usan si NO hay FCST.

 CORTE VIGENTE (edítalo en cada cierre): REAL enero–julio 2026 · FCST agosto–diciembre 2026

 Produce:
   - Input_MEC_Devengamiento.xlsx   (hojas: Input · Cobertura · Validaciones)
   - TriangulosPrimaDevengada.csv   (lo que consume mec_devengamiento: unidad=LN2)

 QUÉ CAMBIÓ EN ESTA VERSIÓN
   El Input ahora arrastra tres campos que antes se leían y se tiraban, y sin los
   cuales no se puede auditar el FND ni recalibrarlo:
     · AnioSusc      (Año Susc.)              -> permite verificar el filtro
                      «Val(left(aPog_MesProc,4)) <= Susc» del reforecast RRC, que en
                      la validación deja la RRC en el 29% de la real.
     · FinVigAAAAMM  (Fecha Fin de Vigencia)  -> separa la prorrata exacta del no
                      proporcional de la tabla por antigüedad de registro.
     · MesesPeriodo  (periodicidad de cuentas)-> separa la frecuencia del δ calibrado.
   Los tres son OPCIONALES: si la BD no trae la columna, el campo queda vacío y todo
   lo demás funciona igual.

 Reglas clave:
   · SOLO primas (candado): BD Tipo Póliza ∈ {PD,PF,PV}; herramienta cuenta 61 ×(−1)
   · Origen 'Real' SOLO en la fuente BD con Periodo ≤ FRONTERA_REAL;
     todo lo de la herramienta es proyección ('Ppto') — el MEC estima con Real.
   · Cohorte MENSUAL desde Fecha Inicio de Vigencia (BD). La herramienta solo trae
     año (ZSUSCYEAR) → cohorte anclada a enero, y SOLO aporta la ventana jul–dic 2026.
================================================================================
"""
import os, re, glob, getpass
import datetime as _dt

def _guardar_seguro(guardar_fn, ruta, que="archivo"):
    """Guarda con guardar_fn(ruta). Si el archivo está ABIERTO/bloqueado en Windows
    (PermissionError), reintenta con un nombre alterno con hora y avisa, sin tronar."""
    try:
        guardar_fn(ruta)
        return ruta
    except PermissionError:
        raiz, ext = os.path.splitext(ruta)
        alt = f"{raiz}_{_dt.datetime.now():%H%M%S}{ext}"
        guardar_fn(alt)
        print(f"[input] AVISO: '{os.path.basename(ruta)}' está ABIERTO (probablemente en Excel).")
        print(f"[input]        Lo guardé como '{os.path.basename(alt)}'. Cierra el original y")
        print(f"[input]        renómbralo, o vuelve a correr con el archivo cerrado.")
        return alt
import numpy as np
import pandas as pd

# =============================== CONFIG =====================================
# El script se ancla a SU PROPIA CARPETA (ponlo en «…\OneDrive - GPV\Documents»).
# Así no depende del usuario, de acentos ni del directorio desde el que lo ejecutes.
BASE = os.path.dirname(os.path.abspath(__file__))

RUTA_BD_EXPLICITA   = ""     # opcional: ruta completa de la BD si está en otro lado
RUTA_FCST_EXPLICITA = ""     # opcional: ruta completa del FCST si está en otro lado
RUTA_RESUMENES      = os.path.join(BASE, "Resumen Ppto Tecnico")
RUTA_SALIDA         = BASE

FRONTERA_REAL  = 202607      # último Periodo REAL de la BD (edítalo en cada cierre)
VENTANA_PPTO   = (202608, 202612)  # ago–dic 2026: meses que aporta el FCST
MONEDA         = "Nal"       # "Nal" (PrimasNal, MXN) o "USD" (Primas USD)
COHORTE_MIN    = 1990        # cohortes por debajo se excluyen (V4) y se reportan
RUTA_SUBRAMO   = None        # opcional: CSV CeBe→Ramo para las fuentes de la herramienta

MESES_MIN_VIGENCIAS = 24     # meses de registro que debe cubrir la BD para poder
                             # reescribir Registros_Vigencia_MEC.csv (curva PF+)

HOJA_FCST      = "Ppto2026"  # hoja del FCST; si no existe se busca por columnas
COL_MONTO_FCST = "PmasEmi"   # prima emitida del FCST
MONEDA_FCST    = "USD"       # el FCST viene en DÓLARES (Moneda=31 en todas sus filas);
                             # si MONEDA="Nal" se convierte con TC_MENSUAL

# Homologación de ramo: el FCST abre subramos (31/35/39 de A&E, 71/73 de CAT) y la BD
# real los trae agregados (30, 70). Se colapsa el FCST al grano de la BD para que las
# dos fuentes sumen en el mismo eje.
HOMOLOGA_RAMO = {31: 30, 34: 30, 35: 30, 37: 30, 39: 30, 71: 70, 72: 70, 73: 70, 20: 10}

# TC de cierre por mes (base BEL-IRR-MR). Se usa sólo para pasar el FCST de USD a MXN.
# Si existe un CSV «tc_mensual_bd.csv» (columnas Periodo,TC) en la carpeta, manda ese.
ARCHIVO_TC = "tc_mensual_bd.csv"
TC_MENSUAL = {202601: 17.4201, 202602: 17.2318, 202603: 17.9252, 202604: 17.4688,
              202605: 17.3401, 202606: 17.434371, 202607: 17.528643, 202608: 17.622914,
              202609: 17.717186, 202610: 17.811457, 202611: 17.905729, 202612: 18.000000}


def cargar_tc():
    """TC de cierre por Periodo. Prioriza el CSV de la carpeta sobre la tabla fija."""
    ruta = os.path.join(BASE, ARCHIVO_TC)
    tc = dict(TC_MENSUAL)
    if os.path.exists(ruta):
        try:
            t = pd.read_csv(ruta)
            tc.update({int(r.Periodo): float(r.TC) for r in t.itertuples()})
            print(f"[input] TC leído de {ARCHIVO_TC} ({len(t)} meses)")
        except Exception as e:
            print(f"[input] Aviso: no pude leer {ARCHIVO_TC} ({e}); uso la tabla fija.")
    return tc


def localizar_bd():
    """Encuentra la BD en la carpeta del script (tolera acentos y variantes de nombre).
    Si no la encuentra, ABORTA con un mensaje claro (no inventa rutas)."""
    if RUTA_BD_EXPLICITA:
        if os.path.exists(RUTA_BD_EXPLICITA):
            return RUTA_BD_EXPLICITA
        raise SystemExit(f"[input] RUTA_BD_EXPLICITA no existe:\n  {RUTA_BD_EXPLICITA}")
    patrones = ["BD_Ppto*.xls*", "BD*Tecnico*RPAT*.xls*", "BD*Técnico*RPAT*.xls*",
                "BDReal*.xls*", "BD*Real*.xls*"]
    cand = []
    for p in patrones:
        cand += glob.glob(os.path.join(BASE, p))
        cand += glob.glob(os.path.join(BASE, "*", p))      # una subcarpeta de profundidad
    cand = list(set(cand))
    # Prioridad: primero el AÑO más alto que aparezca en el nombre (2027 > 2026 > sin año),
    # y a igualdad de año, el archivo más reciente. Así "BD_...2027" gana a "BD_...GENERADA".
    def _anio(f):
        aa = re.findall(r'(20\d{2})', os.path.basename(f))
        return max(int(a) for a in aa) if aa else 0
    cand = sorted(cand, key=lambda f: (_anio(f), os.path.getmtime(f)), reverse=True)
    if cand:
        if len(cand) > 1:
            print("[input] Varias BD candidatas; uso la de año más reciente:")
            for c in cand[:5]:
                marca = f" (año {_anio(c)})" if _anio(c) else ""
                print(f"        {'-> ' if c == cand[0] else '   '}{os.path.basename(c)}{marca}")
        return cand[0]
    vistos = sorted(os.path.basename(f) for f in glob.glob(os.path.join(BASE, "*.xls*")))
    raise SystemExit(
        "[input] No encontré la BD (patrón «BD_Ppto*.xls*») en:\n"
        f"  {BASE}\n"
        "Archivos Excel que sí veo ahí:\n  " + ("\n  ".join(vistos) if vistos else "(ninguno)") + "\n"
        "Soluciones: (a) copia BD_PptoTécnicoRPAT_GENERADA.xlsx a esa carpeta, o\n"
        "            (b) escribe su ruta completa en RUTA_BD_EXPLICITA (arriba)."
    )


def localizar_fcst():
    """Encuentra el archivo del FCST en la carpeta del script. Devuelve None si no
    está: en ese caso el input se arma sólo con el real y las fuentes alternas."""
    if RUTA_FCST_EXPLICITA:
        if os.path.exists(RUTA_FCST_EXPLICITA):
            return RUTA_FCST_EXPLICITA
        raise SystemExit(f"[input] RUTA_FCST_EXPLICITA no existe:\n  {RUTA_FCST_EXPLICITA}")
    cand = []
    for p in ["FCST*.xls*", "Fcst*.xls*", "*Reforecast*.xls*"]:
        cand += glob.glob(os.path.join(BASE, p))
        cand += glob.glob(os.path.join(BASE, "*", p))
    cand = [c for c in set(cand) if not os.path.basename(c).startswith("~$")]
    if not cand:
        return None
    def _anio(f):
        aa = re.findall(r'(20\d{2})', os.path.basename(f))
        return max(int(a) for a in aa) if aa else 0
    cand = sorted(cand, key=lambda f: (_anio(f), os.path.getmtime(f)), reverse=True)
    if len(cand) > 1:
        print("[input] Varios FCST candidatos; uso el de año más reciente:")
        for c in cand[:5]:
            print(f"        {'-> ' if c == cand[0] else '   '}{os.path.basename(c)}")
    return cand[0]


def detectar_header(path, hoja, claves, max_filas=8):
    """Devuelve el índice de la fila que contiene los NOMBRES de las columnas.

    Hace falta porque las bases que exporta la casa suelen traer una o dos filas de
    títulos, totales o fórmulas antes del encabezado (BDReal26 lo trae en la fila 2 y
    FCST2026 en la 3). Leerlas con header=0 devuelve «Unnamed: N» y el script fallaba
    diciendo que a la hoja le faltaban columnas."""
    for h in range(max_filas):
        try:
            cols = set(str(c).strip() for c in leer_excel(path, sheet_name=hoja, nrows=0, header=h).columns)
        except Exception:
            continue
        if claves <= cols:
            return h
    return 0


def leer_excel(path, **kw):
    """read_excel con el engine correcto según extensión (xlsb requiere pyxlsb)."""
    if os.path.splitext(path)[1].lower() == ".xlsb":
        kw.setdefault("engine", "pyxlsb")
    try:
        return pd.read_excel(path, **kw)
    except ImportError:
        raise SystemExit("[input] Falta el motor para .xlsb. Instala:  pip install pyxlsb")


_EPOCH_XL = pd.Timestamp('1899-12-30')

def _a_fecha_serie(serie):
    """Convierte una columna de fechas de vigencia a datetime, tolerando que la
    columna venga MEZCLADA. En la base 2027 conviven en la MISMA columna valores
    datetime (139k) y seriales de Excel como entero (56k); resolver la columna con
    un solo criterio malinterpreta los seriales y colapsa la vigencia a 0 días
    (se perdía el 30% de los registros). Por eso se resuelve valor por valor:
    los numéricos plausibles como serial (1000..80000) se convierten desde la época
    de Excel y el resto se parsea como fecha."""
    s = pd.Series(serie).reset_index(drop=True)
    if np.issubdtype(getattr(s, "dtype", object), np.datetime64):
        return pd.to_datetime(s, errors='coerce')
    num = pd.to_numeric(s, errors='coerce')
    es_serial = num.notna() & (num > 1000) & (num < 80000)
    out = pd.Series(pd.NaT, index=s.index, dtype='datetime64[ns]')
    if es_serial.any():
        out.loc[es_serial] = _EPOCH_XL + pd.to_timedelta(num[es_serial], unit='D')
    resto = ~es_serial
    if resto.any():
        out.loc[resto] = pd.to_datetime(s[resto], errors='coerce')
    out.index = pd.Series(serie).index
    return out


# columnas mínimas que definen la hoja de datos de la BD
_CLAVE_HOJA = {'Periodo', 'PrimasNal', 'Fecha Inicio de Vigencia', 'Ramo', 'Tipo Poliza'}

def hoja_datos(path):
    """Devuelve el nombre de la hoja que contiene los datos de la BD, buscándola por
    sus columnas. Así funciona igual si la hoja se llama «BD», «Hoja1» o cualquier
    otra (la base 2027 la trae como «Hoja1»)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        hojas = wb.sheetnames
        wb.close()
    except Exception:
        hojas = None
    candidatas = (['BD'] + [h for h in (hojas or []) if h != 'BD']) if hojas else ['BD', 'Hoja1', 0]
    for h in candidatas:
        for hdr in range(4):        # el encabezado puede venir 1-3 filas abajo
            try:
                cols = set(str(c).strip() for c in
                           leer_excel(path, sheet_name=h, nrows=0, header=hdr).columns)
            except Exception:
                continue
            if _CLAVE_HOJA <= cols:
                return h, hdr
    # último recurso: la primera hoja con encabezado en la primera fila
    return (hojas[0] if hojas else 0), 0

COLS41 = ['/ERP/CHRTACCT','/ERP/GL_ACCT','/ERP/CATEGORY','ZSOURCE','/ERP/CO_AREA',
 '/ERP/COMPCODE','/ERP/FUNCAREA','ZPAISCEDN','/ERP/PRODUCT','ZCUSTOMER',
 '0CURRENCY','/ERP/TWAERS','/ERP/COSTCNTR','/ERP/PROFTCTR','ZCONTRATO',
 'ZPLANYEAR','0FISCPER','0FISCPER3','0FISCYEAR','0FISCVARNT',
 '0CALMONTH','0CALMONTH2','0CALYEAR','ZINDICES','ZCONCEPTO',
 'ZDISTCHRP','ZDISTCHGS','ZTIPOREAS','ZTIPOCES','ZMGA',
 'ZOFICN_RP','ZOFICN_GS','ZSUSCYEAR','ZCEDENTE','ZCORREDOR',
 'ZTIPVENTA','ZSEGMENTO','ZCICCULTV','/ERP/AMOUNT_T','/ERP/AMOUNT',
 'MANDT','ZREGIONRP']

CANON = ['Fuente','LN2','Ramo','TipoRea','AnioSusc','CohorteAnio','CohorteAAAAMM',
         'FinVigAAAAMM','MesesPeriodo','Periodo','Antiguedad','Origen','Moneda',
         'PrimaDevMes','PrimaDevAcum']

# Nombres posibles de la periodicidad de cuentas en la BD (se usa el primero que exista)
COLS_FREC = ['Meses Periodo', 'Meses Período', 'Periodos', 'Períodos', 'Período', 'Periodo Cuentas']

REPORTE = []   # (regla, detalle, valor, severidad)
def rep(regla, detalle, valor, sev): REPORTE.append((regla, detalle, valor, sev))


# ============================ FUENTE 1 · BD =================================
def fuente_bd():
    ruta = localizar_bd()
    hoja, hdr = hoja_datos(ruta)
    print(f"[input] BD principal (REAL): {ruta}  (hoja «{hoja}», encabezado en la fila {hdr + 1})")
    use = ['Periodo','Año Susc.','Fecha Inicio de Vigencia','LN2','Ramo','Tipo Poliza','Tipo Rea','PrimasNal','Primas USD']
    cols = list(leer_excel(ruta, sheet_name=hoja, nrows=0, header=hdr).columns)
    faltan = [c for c in use if c not in cols]
    if faltan:
        raise SystemExit(f"[input] A la hoja «BD» le faltan columnas: {faltan}\n"
                         f"        Columnas que sí tiene: {cols}")
    # opcionales: fin de vigencia y periodicidad de cuentas (si no están, quedan vacías)
    col_fin = 'Fecha Fin de Vigencia' if 'Fecha Fin de Vigencia' in cols else None
    col_frec = next((c for c in COLS_FREC if c in cols), None)
    if col_fin is None:
        rep('V12', 'Fecha Fin de Vigencia en la BD', 'ausente',
            'Advertencia — sin ella no se puede separar la prorrata del no proporcional')
    if col_frec is None:
        rep('V13', f'Periodicidad de cuentas en la BD (busqué {COLS_FREC})', 'ausente',
            'Advertencia — la frecuencia queda absorbida en el δ calibrado')
    bd = leer_excel(ruta, sheet_name=hoja, header=hdr, usecols=use + [c for c in (col_fin, col_frec) if c])
    n0 = len(bd)
    prim = bd[bd['Tipo Poliza'].astype(str).str.startswith('P')].copy()
    rep('V2','Filas excluidas por no ser primas (S*/R*/OD/«20») — fuente BD', n0-len(prim), 'OK (candado)')
    # COHORTE MENSUAL desde Fecha Inicio de Vigencia (viene como SERIAL de Excel -> fecha)
    iv = _a_fecha_serie(prim['Fecha Inicio de Vigencia'])
    sin_iv = iv.isna()
    if sin_iv.any():
        rep('V4b', 'Filas sin Fecha Inicio de Vigencia válida (excluidas)',
            f"{int(sin_iv.sum())} filas · monto {prim.loc[sin_iv, ('PrimasNal' if MONEDA=='Nal' else 'Primas USD')].sum():,.0f}",
            'Bloqueante aplicada')
    prim = prim[~sin_iv].copy(); iv = iv[~sin_iv]
    prim['CohorteAAAAMM'] = (iv.dt.year * 100 + iv.dt.month).astype(int)
    prim['AnioSusc'] = pd.to_numeric(prim['Año Susc.'], errors='coerce')
    if col_fin is not None:
        fv = _a_fecha_serie(prim[col_fin])
        prim['FinVigAAAAMM'] = (fv.dt.year * 100 + fv.dt.month)
    else:
        prim['FinVigAAAAMM'] = np.nan
    prim['MesesPeriodo'] = pd.to_numeric(prim[col_frec], errors='coerce') if col_frec else np.nan
    monto = 'PrimasNal' if MONEDA == 'Nal' else 'Primas USD'
    # dropna=False: si algún campo opcional viene vacío, la fila NO se pierde
    llaves = ['LN2','Ramo','Tipo Rea','AnioSusc','CohorteAAAAMM','FinVigAAAAMM','MesesPeriodo','Periodo']
    g = (prim.groupby(llaves, as_index=False, dropna=False)[monto]
             .sum().rename(columns={'Tipo Rea':'TipoRea', monto:'PrimaDevMes'}))
    g['Fuente'] = 'BD'
    print(f"[input]   {len(bd):,} filas · {len(prim):,} de primas · {len(g):,} filas agregadas · "
          f"cohortes MENSUALES por inicio de vigencia")
    return g


def registros_vigencia():
    """Genera Registros_Vigencia_MEC.csv: la fuente del FND por PRORRATA EXACTA.

    Una fila por (Ramo, inicio de vigencia, fin de vigencia) con la prima agregada.
    Es lo que consume mec_devengamiento (M2) para calcular el FND sin estimar nada:
    el devengamiento sale de las fechas reales de cobertura, no de un triángulo de
    registro (ver encabezado del módulo)."""
    ruta = localizar_bd()
    hoja, hdr = hoja_datos(ruta)
    # 'Periodo' se lee para saber cuántos meses de REGISTRO cubre la BD; de eso depende
    # si se puede reescribir el histórico de vigencias (ver candado más abajo).
    use = ['Periodo','Ramo','Tipo Poliza','Fecha Inicio de Vigencia','Fecha Fin de Vigencia',
           'PrimasNal','Primas USD']
    cols = list(leer_excel(ruta, sheet_name=hoja, nrows=0, header=hdr).columns)
    # Tipo Rea se arrastra para poder separar el no proporcional (prorrata exacta) del
    # proporcional y facultativo (tabla por antigüedad de registro). Es opcional: el
    # módulo MEC sólo exige Ramo | ini | fin | Prima.
    if 'Tipo Rea' in cols:
        use = use + ['Tipo Rea']
    faltan = [c for c in use if c not in cols]
    if faltan:
        print(f"[input] Aviso: sin columnas {faltan}; no se puede generar el archivo de vigencias.")
        return None
    bd = leer_excel(ruta, sheet_name=hoja, header=hdr, usecols=use)
    p = bd[bd['Tipo Poliza'].astype(str).str.startswith('P')].copy()
    p['ini'] = _a_fecha_serie(p['Fecha Inicio de Vigencia'])
    p['fin'] = _a_fecha_serie(p['Fecha Fin de Vigencia'])
    monto = 'PrimasNal' if MONEDA == 'Nal' else 'Primas USD'
    dur = (p['fin'] - p['ini']).dt.days
    ok = p['ini'].notna() & p['fin'].notna() & (dur > 0) & (dur <= 3660)
    rep('V11', 'Registros de prima con vigencia válida (base del FND por prorrata)',
        f"{int(ok.sum()):,} de {len(p):,} ({ok.mean():.1%})",
        'OK' if ok.mean() >= 0.95 else 'Advertencia — revisar fechas')
    p = p[ok]
    llaves = ['Ramo', 'ini', 'fin'] + (['Tipo Rea'] if 'Tipo Rea' in p.columns else [])
    g = (p.groupby(llaves, as_index=False, dropna=False)[monto]
           .sum().rename(columns={monto: 'Prima', 'Tipo Rea': 'TipoRea'}))

    # CANDADO: la curva PF+ de cartera (que usa el NO PROPORCIONAL) se estima de este
    # archivo. Una BD de pocos meses de registro produce una curva pobre, así que NO se
    # pisa el histórico: se guarda aparte y se avisa. Sólo se sobrescribe cuando la BD
    # cubre al menos MESES_MIN_VIGENCIAS meses de registro.
    meses = sorted(pd.to_numeric(bd['Periodo'], errors='coerce').dropna().astype(int).unique()) \
        if 'Periodo' in bd.columns else []
    n_meses = len(meses)
    destino = os.path.join(RUTA_SALIDA, 'Registros_Vigencia_MEC.csv')
    historico = os.path.exists(destino)
    if n_meses and n_meses < MESES_MIN_VIGENCIAS and historico:
        destino = os.path.join(RUTA_SALIDA,
                               f'Registros_Vigencia_MEC_{meses[0]}_{meses[-1]}.csv')
        rep('V18', 'Vigencias: la BD cubre pocos meses de registro',
            f"{n_meses} meses ({meses[0]}–{meses[-1]}) · guardado en "
            f"{os.path.basename(destino)}; se conserva el histórico",
            'Advertencia — la curva PF+ sigue con el archivo histórico')
        print(f"[input]   AVISO: la BD sólo cubre {n_meses} meses de registro. NO piso "
              f"Registros_Vigencia_MEC.csv; guardo «{os.path.basename(destino)}».")
    else:
        rep('V18', 'Vigencias: meses de registro que cubre la BD',
            f"{n_meses} meses" + (f" ({meses[0]}–{meses[-1]})" if meses else ""), 'OK')
    x = _guardar_seguro(lambda ruta_out: g.to_csv(ruta_out, index=False), destino, "Vigencias")
    print(f"[input]   Vigencias: {len(g):,} combinaciones (Ramo × vigencia) · "
          f"duración mediana {dur[ok].median()/30.4:.1f} meses")
    return x


# ========================= FUENTE 2 · FCST ==================================
_CLAVE_FCST = {'AñoPpto', 'MesPpto', 'AñoSusc', 'Ramo', 'TipoRea', 'PmasEmi'}
_TRASLAPE = {}      # lo llena fuente_fcst() y lo reporta validar() (regla V9)


def fuente_fcst(path=None):
    """Proyección del año en curso: aporta SOLO los meses de VENTANA_PPTO.

    Diferencias con la BD real, que se resuelven aquí:
      · viene en DÓLARES  -> se convierte a la moneda del input con el TC de cierre
      · abre subramos      -> se homologa al grano de la BD (31/35/39->30, 71/73->70)
      · no trae vigencia   -> FinVigAAAAMM y MesesPeriodo quedan vacíos

    COHORTE = MES CONTABLE. El input se organiza por FECHA CONTABLE: cada fila es un
    movimiento que ENTRA a los libros en un mes. El FCST no trae vigencia, así que su
    cohorte es su propio mes de registro y su antigüedad es 0. El año de suscripción se
    conserva como dato descriptivo (AnioSusc), no como ancla.

    Anclar la cohorte a enero del año de suscripción —como hacía la herramienta— es lo
    que NO hay que hacer: se midió que arroja la prima proyectada a antigüedades de 19
    a 71 meses e infla el triángulo en esos tramos, cuando el movimiento en realidad
    acaba de entrar.
    """
    path = path or localizar_fcst()
    if path is None:
        print("[input] Aviso: no encontré el FCST en la carpeta; sigo sin la proyección.")
        return None
    hoja = HOJA_FCST
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True); hojas = wb.sheetnames; wb.close()
        if hoja not in hojas:
            hoja = hojas[0]
    except Exception:
        pass
    hdr = detectar_header(path, hoja, _CLAVE_FCST)
    df = leer_excel(path, sheet_name=hoja, header=hdr)
    df.columns = [str(c).strip() for c in df.columns]
    faltan = _CLAVE_FCST - set(df.columns)
    if faltan:
        print(f"[input] Aviso: al FCST le faltan columnas {faltan}; lo omito.")
        return None
    print(f"[input] FCST (PROYECCIÓN): {path}  (hoja «{hoja}», encabezado en la fila {hdr + 1})")

    df['Periodo'] = (pd.to_numeric(df['AñoPpto'], errors='coerce') * 100
                     + pd.to_numeric(df['MesPpto'], errors='coerce')).astype('Int64')
    df['Ramo'] = pd.to_numeric(df['Ramo'], errors='coerce')
    colapsados = sorted(set(df['Ramo'].dropna().astype(int)) & set(HOMOLOGA_RAMO))
    df['Ramo'] = df['Ramo'].map(lambda r: HOMOLOGA_RAMO.get(int(r), int(r)) if pd.notna(r) else r)
    df['AnioSusc'] = pd.to_numeric(df['AñoSusc'], errors='coerce')   # descriptivo
    df['CohorteAAAAMM'] = df['Periodo']        # fecha contable: cuando entra el movimiento
    df['FinVigAAAAMM'] = np.nan
    df['MesesPeriodo'] = np.nan
    df['TipoRea'] = pd.to_numeric(df['TipoRea'], errors='coerce')
    if 'LN2' not in df.columns:
        df['LN2'] = df.get('LíneaNegocio', 'LN_DESCONOCIDA')

    monto = pd.to_numeric(df[COL_MONTO_FCST], errors='coerce').fillna(0.0)
    if MONEDA_FCST.upper() == MONEDA.upper() or (MONEDA_FCST == 'USD' and MONEDA == 'USD'):
        df['PrimaDevMes'] = monto
        factor = None
    else:                      # FCST en USD -> input en MXN
        tc = cargar_tc()
        factor = df['Periodo'].map(lambda p: tc.get(int(p)) if pd.notna(p) else np.nan)
        sin_tc = factor.isna() & df['Periodo'].notna()
        if sin_tc.any():
            rep('V15', 'Meses del FCST sin TC (se excluyen)',
                f"{sorted(df.loc[sin_tc, 'Periodo'].dropna().unique().tolist())}", 'Bloqueante aplicada')
        df['PrimaDevMes'] = monto * factor

    # traslape con el real: se guarda para la validación V9 (presupuesto vs realidad)
    ini, fin = VENTANA_PPTO
    tras = df[(df['Periodo'] < ini) & df['Periodo'].notna()]
    _TRASLAPE['fcst'] = tras.groupby('Periodo')['PrimaDevMes'].sum()
    _TRASLAPE['fcst_ln'] = tras.groupby('LN2')['PrimaDevMes'].sum()
    _TRASLAPE['factor'] = 'TC de cierre' if factor is not None else 'sin conversión'

    g = df[(df['Periodo'] >= ini) & (df['Periodo'] <= fin)].copy()
    if g.empty:
        print(f"[input] Aviso: el FCST no tiene filas en la ventana {ini}–{fin}.")
        return None
    g = (g.groupby(['LN2','Ramo','TipoRea','AnioSusc','CohorteAAAAMM','FinVigAAAAMM',
                    'MesesPeriodo','Periodo'], as_index=False, dropna=False)['PrimaDevMes'].sum())
    g['Fuente'] = 'FCST'
    rep('V14', f'Subramos del FCST colapsados al grano de la BD {HOMOLOGA_RAMO}',
        f"{colapsados}", 'OK (homologación)')
    # V19 · deja constancia del criterio: la cohorte del FCST es su MES CONTABLE, no el
    # año de suscripción. El año de suscripción queda como dato descriptivo.
    anios = [int(a) for a in sorted(g['AnioSusc'].dropna().unique())]
    rep('V19', 'FCST · cohorte = mes contable (fecha de ingreso del movimiento)',
        f"antigüedad 0 · años de suscripción presentes, sólo descriptivos: {anios}", 'OK')
    rep('V16', f'FCST · moneda de origen {MONEDA_FCST} -> input {MONEDA}',
        _TRASLAPE['factor'], 'OK')
    print(f"[input]   FCST -> {len(g):,} filas en la ventana {ini}–{fin} · "
          f"monto {g['PrimaDevMes'].sum():,.0f}")
    return g


# ==================== FUENTES 3 · HERRAMIENTA (alternas) ====================
def _detectar_ln(df, path):
    if '/ERP/FUNCAREA' in df.columns and df['/ERP/FUNCAREA'].notna().any():
        return str(df['/ERP/FUNCAREA'].mode().iloc[0])
    if 'ID' in df.columns and df['ID'].notna().any():          # p.ej. LN04004-1-1161-0-2023-...
        return str(df['ID'].dropna().iloc[0]).split('-')[0]
    m = re.search(r'LN\d{5}', os.path.basename(path))
    return m.group(0) if m else 'LN_DESCONOCIDA'

def _normaliza_herramienta(df, fuente_tag, path):
    # Primas = cuenta 61; signo invertido en la herramienta -> ×(−1)
    df = df[df['/ERP/GL_ACCT'].astype(str).str.startswith('61')].copy()
    df['PrimaDevMes'] = pd.to_numeric(df['/ERP/AMOUNT'], errors='coerce') * -1.0
    ln = _detectar_ln(df, path)
    df['LN2'] = ln
    if RUTA_SUBRAMO and os.path.exists(RUTA_SUBRAMO) and '/ERP/PROFTCTR' in df.columns:
        cat = pd.read_csv(RUTA_SUBRAMO)[['CeBe','Ramo']].drop_duplicates()
        df = df.merge(cat, how='left', left_on='/ERP/PROFTCTR', right_on='CeBe')
        df['Ramo'] = df['Ramo'].fillna(0).astype(int)
    else:
        df['Ramo'] = 0    # por asignar vía catálogo Subramo (CeBe) — no bloquea el input por LN2
    # COHORTE = MES CONTABLE, igual que el FCST: la herramienta tampoco trae vigencia,
    # así que el movimiento cohorta en el mes en que entra. (Antes se anclaba a enero
    # del año de suscripción, lo que enviaba la prima a antigüedades de varios años.)
    df['CohorteAAAAMM'] = pd.to_numeric(df['0CALMONTH'], errors='coerce')
    df['AnioSusc'] = pd.to_numeric(df['ZSUSCYEAR'], errors='coerce')   # descriptivo
    df['FinVigAAAAMM'] = np.nan      # la herramienta no trae vigencia por registro
    df['MesesPeriodo'] = np.nan      # ni periodicidad de cuentas
    df['Periodo'] = pd.to_numeric(df['0CALMONTH'], errors='coerce')
    # SOLO se usa la proyección de suscripción para la VENTANA Ppto (jul–dic del año frontera).
    df = df[(df['Periodo'] >= VENTANA_PPTO[0]) & (df['Periodo'] <= VENTANA_PPTO[1])]
    g = (df.groupby(['LN2','Ramo','ZTIPOREAS','AnioSusc','CohorteAAAAMM','FinVigAAAAMM',
                     'MesesPeriodo','Periodo'], as_index=False, dropna=False)['PrimaDevMes'].sum()
           .rename(columns={'ZTIPOREAS':'TipoRea'}))
    g['Fuente'] = f'{fuente_tag}-{ln}'
    print(f"[input]   {os.path.basename(path)} -> {ln} · {len(g)} filas en ventana Ppto {VENTANA_PPTO[0]}–{VENTANA_PPTO[1]}")
    return g

def fuente_xlsm(path):
    head = leer_excel(path, sheet_name='BD_TXT', nrows=0)
    quiero = ['/ERP/FUNCAREA','/ERP/GL_ACCT','ZSUSCYEAR','0CALMONTH','ZTIPOREAS',
              '/ERP/AMOUNT','0CURRENCY','/ERP/PROFTCTR','ID']
    use = [c for c in quiero if c in head.columns]
    df = leer_excel(path, sheet_name='BD_TXT', usecols=use)
    return _normaliza_herramienta(df, 'Resumen', path)

def fuente_txt(path):
    try:
        df = pd.read_csv(path, delimiter='\t', names=COLS41, header=None,
                         thousands=',', low_memory=False, skip_blank_lines=True)
    except UnicodeDecodeError:   # los export de la herramienta suelen venir en latin-1
        df = pd.read_csv(path, delimiter='\t', names=COLS41, header=None, encoding='latin-1',
                         thousands=',', low_memory=False, skip_blank_lines=True)
    return _normaliza_herramienta(df, 'Txt', path)

def fuentes_herramienta():
    piezas = []
    if not os.path.isdir(RUTA_RESUMENES):
        print(f"[input] Aviso: no veo la carpeta de resúmenes:\n         {RUTA_RESUMENES}\n         Sigo solo con la BD (el input queda válido, sin las LN de la herramienta).")
        return piezas
    for f in sorted(glob.glob(os.path.join(RUTA_RESUMENES, '*.xls[mx]'))):
        if re.search(r'vsReal', os.path.basename(f), re.I):
            continue      # productos de validación, no fuente (Layout · Fuentes)
        try:
            piezas.append(fuente_xlsm(f))
        except Exception as e:
            print(f"[input]   {os.path.basename(f)}: omitido ({e})")
    for f in sorted(glob.glob(os.path.join(RUTA_RESUMENES, '*.txt'))):
        try:
            piezas.append(fuente_txt(f))
        except Exception as e:
            print(f"[input]   {os.path.basename(f)}: omitido ({e})")
    return piezas


# ====================== DERIVADOS Y VALIDACIONES ============================
def derivar(df):
    df = df.copy()
    for c in ('AnioSusc', 'FinVigAAAAMM', 'MesesPeriodo'):   # por si una fuente no los trae
        if c not in df.columns:
            df[c] = np.nan
    df['CohorteAAAAMM'] = pd.to_numeric(df['CohorteAAAAMM'], errors='coerce').fillna(0).astype(int)
    df['Periodo'] = pd.to_numeric(df['Periodo'], errors='coerce').fillna(0).astype(int)
    df['CohorteAnio'] = df['CohorteAAAAMM'] // 100
    # V4 cohortes inválidas (año fuera de rango)
    malas = df['CohorteAnio'] < COHORTE_MIN
    if malas.any():
        rep('V4', f'Filas con cohorte inválida (año <{COHORTE_MIN}) excluidas',
            f"{int(malas.sum())} filas · monto {df.loc[malas,'PrimaDevMes'].sum():,.0f}", 'Bloqueante aplicada')
        df = df[~malas]
    # Antigüedad MENSUAL exacta = meses entre inicio de cohorte (AAAAMM) y periodo (AAAAMM)
    df['Antiguedad'] = ((df['Periodo']//100 - df['CohorteAAAAMM']//100)*12
                        + (df['Periodo']%100 - df['CohorteAAAAMM']%100))
    # V3 antigüedad negativa
    neg = df['Antiguedad'] < 0
    if neg.any():
        rep('V3', 'Filas con antigüedad negativa excluidas (Periodo anterior a la cohorte)',
            f"{int(neg.sum())} filas · monto {df.loc[neg,'PrimaDevMes'].sum():,.0f}", 'Bloqueante aplicada')
        df = df[~neg]
    df['Origen'] = np.where((df['Fuente']=='BD') & (df['Periodo']<=FRONTERA_REAL), 'Real', 'Ppto')
    df['Moneda'] = 'MXN (Nal)' if MONEDA=='Nal' else 'USD'
    df = df.sort_values(['Fuente','LN2','Ramo','TipoRea','CohorteAAAAMM','Periodo'])
    df['PrimaDevAcum'] = df.groupby(['Fuente','LN2','Ramo','TipoRea','CohorteAAAAMM'])['PrimaDevMes'].cumsum()
    return df[CANON]

def validar(df):
    dup = df.duplicated(subset=['Fuente','LN2','Ramo','TipoRea','AnioSusc','CohorteAAAAMM',
                                'FinVigAAAAMM','MesesPeriodo','Periodo']).sum()
    rep('V1','Llaves duplicadas', int(dup), 'OK' if dup==0 else 'BLOQUEANTE — revisar agregación')
    rep('V5','Moneda única de la corrida', df['Moneda'].nunique(), 'OK' if df['Moneda'].nunique()==1 else 'BLOQUEANTE')
    reales = df[df['Origen']=='Real']
    rep('V6','Filas Real (estiman el MEC) / Ppto (solo comparación)',
        f"{len(reales):,} / {len(df)-len(reales):,}", 'OK')
    caidas = (df.sort_values('Periodo').groupby(['Fuente','LN2','TipoRea','CohorteAAAAMM'])['PrimaDevAcum']
                .apply(lambda s: int((s.diff() < 0).sum())).sum())
    rep('V7','Meses con acumulado a la baja (devoluciones, signo natural)', int(caidas), 'Advertencia (se reporta)')
    # V9 · PRESUPUESTO vs REALIDAD en el traslape. El FCST proyecta los doce meses del
    # año, así que sus meses YA REALIZADOS (enero–FRONTERA_REAL) se pueden contrastar
    # contra el real. No es un cuadre contable —son dos cosas distintas— sino la medida
    # de qué tan buena fue la proyección en lo que ya se conoce; por eso el umbral es
    # holgado (10%) y sólo advierte.
    if len(_TRASLAPE.get('fcst', [])):
        real_m = (df[(df['Fuente'] == 'BD') & (df['Origen'] == 'Real')]
                  .groupby('Periodo')['PrimaDevMes'].sum())
        f_m = _TRASLAPE['fcst']
        comunes = sorted(set(real_m.index) & set(f_m.index))
        if comunes:
            ra, fa = float(real_m.reindex(comunes).sum()), float(f_m.reindex(comunes).sum())
            d = (fa / ra - 1) if ra else float('nan')
            rep('V9', f'FCST vs REAL en el traslape {comunes[0]}–{comunes[-1]} (acumulado)',
                f'real {ra:,.0f} · fcst {fa:,.0f} · {d:+.1%}',
                'OK' if abs(d) <= 0.10 else 'Advertencia — la proyección se separa del real')
            for p in comunes:
                if real_m[p]:
                    dm = float(f_m[p]) / float(real_m[p]) - 1
                    rep('V9', f'FCST vs REAL · {p}', f'{dm:+.1%}',
                        'OK' if abs(dm) <= 0.10 else 'Advertencia')
    # V9b · cuadre contra las fuentes alternas de la herramienta (si se usaron)
    her = df[~df['Fuente'].isin(['BD', 'FCST'])]
    if len(her):
        vent = (FRONTERA_REAL//100)*100
        a = (df[(df['Fuente']=='BD') & (df['Periodo']>vent) & (df['Periodo']<=FRONTERA_REAL)]
             .groupby('LN2')['PrimaDevMes'].sum())
        b = (her[(her['Periodo']>vent) & (her['Periodo']<=FRONTERA_REAL)]
             .groupby('LN2')['PrimaDevMes'].sum())
        for ln in b.index:
            if ln in a.index and a[ln]!=0:
                d = b[ln]/a[ln]-1
                rep('V9b', f'Cuadre BD vs herramienta {ln} ({vent+6}–{FRONTERA_REAL})',
                    f'{d:+.1%}', 'OK' if abs(d)<=0.05 else 'Advertencia — investigar')
    for (ln,), s in df[df['Origen']=='Real'].groupby(['LN2'])['PrimaDevMes']:
        if abs(s.sum()) < 1e6:
            rep('V10', f'Volumen Real bajo en {ln}', f"{s.sum():,.0f}", 'Informativa (irá a patrón de cartera)')
    # V17 · el corte del año: qué meses aporta cada fuente y con cuánto monto
    anio = FRONTERA_REAL // 100
    aa = df[(df['Periodo'] // 100 == anio)]
    for fuente, g in aa.groupby('Fuente'):
        ms = sorted(g['Periodo'].unique())
        rep('V17', f'Corte {anio} · fuente {fuente}',
            f"{ms[0]}–{ms[-1]} ({len(ms)} meses) · monto {g['PrimaDevMes'].sum():,.0f}", 'OK')
    faltan = sorted(set(anio * 100 + m for m in range(1, 13)) - set(aa['Periodo'].unique()))
    rep('V17', f'Meses de {anio} sin ninguna fuente', f"{faltan or 'ninguno'}",
        'OK' if not faltan else 'Advertencia — el año queda incompleto')


# ============================== SALIDAS =====================================
def escribir(df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    VERDE='FF00573F'; BLANCO='FFFFFFFF'; CLARO='FFE8F1EA'
    wb = Workbook(); wb.remove(wb.active)

    def banda(ws, titulo, sub, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c=ws.cell(1,1,titulo); c.font=Font(bold=True,color=BLANCO,size=12); c.fill=PatternFill('solid',fgColor=VERDE); c.alignment=Alignment('center','center')
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c=ws.cell(2,1,sub); c.font=Font(italic=True,color=BLANCO,size=9); c.fill=PatternFill('solid',fgColor='FF2E7D53'); c.alignment=Alignment('center','center')

    ws = wb.create_sheet('Input')
    banda(ws,'INPUT CANÓNICO DEL MEC · DEVENGAMIENTO (formato del Layout del Input)',
          f'Moneda {df["Moneda"].iloc[0]} · frontera real {FRONTERA_REAL} · Origen Real solo fuente BD · generado por construir_input_mec.py', len(CANON))
    ws.append(CANON)
    for c in ws[3]: c.font=Font(bold=True,color=BLANCO); c.fill=PatternFill('solid',fgColor=VERDE)
    for row in df.itertuples(index=False): ws.append(list(row))
    ws.freeze_panes='A4'

    ws = wb.create_sheet('Cobertura')
    banda(ws,'COBERTURA POR FUENTE × LN2','filas · cohortes · periodos · monto total del mes (PrimaDevMes)',7)
    ws.append(['Fuente','LN2','Filas','Cohorte min','Cohorte max','Periodo min–max','Σ PrimaDevMes'])
    for c in ws[3]: c.font=Font(bold=True,color=BLANCO); c.fill=PatternFill('solid',fgColor=VERDE)
    cov = (df.groupby(['Fuente','LN2']).agg(F=('PrimaDevMes','size'), c0=('CohorteAnio','min'),
           c1=('CohorteAnio','max'), p0=('Periodo','min'), p1=('Periodo','max'), M=('PrimaDevMes','sum')).reset_index())
    for r in cov.itertuples(index=False):
        ws.append([r.Fuente, r.LN2, r.F, r.c0, r.c1, f"{r.p0}–{r.p1}", round(r.M)])

    ws = wb.create_sheet('Validaciones')
    banda(ws,'RESULTADO DE VALIDACIONES DEL INPUT (V1–V10 del Layout)',
          f'frontera real = {FRONTERA_REAL} · cohorte mínima = {COHORTE_MIN}',4)
    ws.append(['Regla','Detalle','Valor','Estado'])
    for c in ws[3]: c.font=Font(bold=True,color=BLANCO); c.fill=PatternFill('solid',fgColor=VERDE)
    for r in REPORTE: ws.append(list(r))

    x1 = _guardar_seguro(wb.save, os.path.join(RUTA_SALIDA,'Input_MEC_Devengamiento.xlsx'), "Input")
    # CSV para el módulo MEC: unidad = LN2, SOLO Real, acumulado al grano LN2×cohorte
    # CSV para el MEC: Real (BD) + Ppto jul–dic 2026 (Resumen) para completar el año en curso.
    base = df[(df['Origen']=='Real') | ((df['Origen']=='Ppto') & (df['Periodo']>=VENTANA_PPTO[0]) & (df['Periodo']<=VENTANA_PPTO[1]))]
    g = (base.groupby(['LN2','CohorteAAAAMM','Periodo'], as_index=False)['PrimaDevMes'].sum()
               .sort_values(['LN2','CohorteAAAAMM','Periodo']))
    g['PrimaDevAcum'] = g.groupby(['LN2','CohorteAAAAMM'])['PrimaDevMes'].cumsum()
    g['Antiguedad'] = (g['Periodo']//100 - g['CohorteAAAAMM']//100)*12 + (g['Periodo']%100 - g['CohorteAAAAMM']%100)
    csv = g.rename(columns={'LN2':'Ramo'})[['Ramo','CohorteAAAAMM','Antiguedad','PrimaDevAcum']]
    x2 = _guardar_seguro(lambda p: csv.to_csv(p, index=False),
                         os.path.join(RUTA_SALIDA,'TriangulosPrimaDevengada.csv'), "CSV")
    print(f"[input] Escritos:\n  {x1}\n  {x2}")
    return x1, x2


if __name__ == '__main__':
    # 1) Fuente de la prorrata exacta por fechas de vigencia (no proporcional).
    registros_vigencia()
    # 2) Input canónico: REAL de la BD hasta FRONTERA_REAL + PROYECCIÓN del FCST en la
    #    VENTANA_PPTO. Si no hay FCST se cae a las fuentes alternas de la herramienta.
    piezas = [fuente_bd()]
    fcst = fuente_fcst()
    if fcst is not None:
        piezas.append(fcst)
    else:
        piezas += fuentes_herramienta()
    inp = derivar(pd.concat(piezas, ignore_index=True))
    validar(inp)
    escribir(inp)
    print(f"[input] Input canónico: {len(inp):,} filas · fuentes: {sorted(inp['Fuente'].unique())}")
